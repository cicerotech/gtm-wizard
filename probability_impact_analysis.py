#!/usr/bin/env python3
"""
Probability Impact Analysis - Final Version
Q4 FY2026: November 1, 2025 - January 31, 2026
Active Stages Only: Stage 0-4 (excludes Closed Won/Lost and Stage 5)

Conservative Approach: EC unchanged, only LOI decreases
Expected Net Impact: ~-$38K (-0.9%)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

# Q4 FY2026 Date Range
Q4_START = datetime(2025, 11, 1)
Q4_END = datetime(2026, 1, 31)

# Active stages only (exclude Closed Won/Lost and Stage 5)
ACTIVE_STAGES = [
    "Stage 0 - Qualifying",
    "Stage 1 - Discovery", 
    "Stage 2 - SQO",
    "Stage 3 - Pilot",
    "Stage 4 - Proposal"
]

# Current probabilities (from existing formula)
CURRENT_PROBS = {
    "Stage 0 - Qualifying": {"New Logo": 0.02, "Existing Client": 0.02, "LOI": 0.02, "Government": 0.02},
    "Stage 1 - Discovery": {"New Logo": 0.10, "Existing Client": 0.18, "LOI": 0.20, "Government": 0.08},
    "Stage 2 - SQO": {"New Logo": 0.20, "Existing Client": 0.32, "LOI": 0.35, "Government": 0.12},
    "Stage 3 - Pilot": {"New Logo": 0.25, "Existing Client": 0.42, "LOI": 0.45, "Government": 0.18},
    "Stage 4 - Proposal": {"New Logo": 0.33, "Existing Client": 0.50, "LOI": 0.55, "Government": 0.22},
}

# NEW probabilities - CONSERVATIVE: EC unchanged, only LOI decreases
NEW_PROBS = {
    "Stage 0 - Qualifying": {"New Logo": 0.02, "Existing Client": 0.02, "LOI": 0.02, "Government": 0.02},
    "Stage 1 - Discovery": {"New Logo": 0.10, "Existing Client": 0.18, "LOI": 0.17, "Government": 0.08},  # LOI: 20% -> 17%
    "Stage 2 - SQO": {"New Logo": 0.20, "Existing Client": 0.32, "LOI": 0.30, "Government": 0.12},  # LOI: 35% -> 30%
    "Stage 3 - Pilot": {"New Logo": 0.25, "Existing Client": 0.42, "LOI": 0.40, "Government": 0.18},  # LOI: 45% -> 40%
    "Stage 4 - Proposal": {"New Logo": 0.33, "Existing Client": 0.50, "LOI": 0.48, "Government": 0.22},  # LOI: 55% -> 48%
}

def parse_date(date_val):
    """Parse date from various formats"""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y"]:
            try:
                return datetime.strptime(date_val, fmt)
            except:
                pass
    return None

def normalize_stage(stage):
    """Normalize stage name to standard format"""
    if not stage:
        return None
    stage = str(stage).strip()
    if "Stage 0" in stage or "Qualifying" in stage:
        return "Stage 0 - Qualifying"
    elif "Stage 1" in stage or "Discovery" in stage:
        return "Stage 1 - Discovery"
    elif "Stage 2" in stage or "SQO" in stage:
        return "Stage 2 - SQO"
    elif "Stage 3" in stage or "Pilot" in stage:
        return "Stage 3 - Pilot"
    elif "Stage 4" in stage or "Proposal" in stage:
        return "Stage 4 - Proposal"
    elif "Stage 5" in stage or "Negotiation" in stage:
        return "Stage 5 - Negotiation"
    elif "Closed" in stage:
        return "Closed"
    return stage

def analyze_pipeline():
    print("="*70)
    print("PROBABILITY IMPACT ANALYSIS")
    print(f"Q4 FY2026: {Q4_START.strftime('%b %d, %Y')} - {Q4_END.strftime('%b %d, %Y')}")
    print("="*70)
    
    # Load source file
    print("\nLoading 222.xlsx...")
    wb = openpyxl.load_workbook("/Users/keiganpesenti/Desktop/222.xlsx")
    ws = wb.active
    
    # Get headers from row 1
    headers = [cell.value for cell in ws[1]]
    print(f"Headers found: {headers}")
    
    # Build column index map
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h.strip()] = i
    
    # Collect all deals
    all_deals = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        # Extract fields using column map or positional fallback
        deal = {
            "row": row_num,
            "pod": row[0],
            "sales_type": row[1] if len(row) > 1 else None,
            "acct_class": row[2] if len(row) > 2 else None,
            "stage": row[3] if len(row) > 3 else None,
            "deal_status": row[4] if len(row) > 4 else None,
            "owner": row[5] if len(row) > 5 else None,
            "opp_name": row[6] if len(row) > 6 else None,
            "target_date_raw": row[7] if len(row) > 7 else None,
            "weighted_acv_raw": row[8] if len(row) > 8 else None,
            "calc_prob_raw": row[9] if len(row) > 9 else None,
            "custom_prob_raw": row[10] if len(row) > 10 else None,
        }
        
        # Parse weighted ACV
        wav = deal["weighted_acv_raw"]
        if wav is None:
            deal["weighted_acv"] = 0
        elif isinstance(wav, (int, float)):
            deal["weighted_acv"] = float(wav)
        elif isinstance(wav, str):
            try:
                deal["weighted_acv"] = float(wav.replace("$", "").replace(",", ""))
            except:
                deal["weighted_acv"] = 0
        else:
            deal["weighted_acv"] = 0
        
        # Parse calculated probability
        cp = deal["calc_prob_raw"]
        if cp is None:
            deal["calc_prob"] = None
        elif isinstance(cp, (int, float)):
            deal["calc_prob"] = float(cp) / 100 if cp > 1 else float(cp)
        elif isinstance(cp, str):
            try:
                deal["calc_prob"] = float(cp.replace("%", "")) / 100
            except:
                deal["calc_prob"] = None
        else:
            deal["calc_prob"] = None
        
        # Parse custom probability
        cust = deal["custom_prob_raw"]
        if cust is None:
            deal["custom_prob"] = None
        elif isinstance(cust, (int, float)):
            deal["custom_prob"] = float(cust) / 100 if cust > 1 else float(cust)
        elif isinstance(cust, str):
            try:
                deal["custom_prob"] = float(cust.replace("%", "")) / 100
            except:
                deal["custom_prob"] = None
        else:
            deal["custom_prob"] = None
        
        # Parse target date
        deal["target_date"] = parse_date(deal["target_date_raw"])
        
        # Normalize stage
        deal["stage_normalized"] = normalize_stage(deal["stage"])
        
        # Check if Q4 (within date range)
        deal["is_q4"] = False
        if deal["target_date"]:
            if Q4_START <= deal["target_date"] <= Q4_END:
                deal["is_q4"] = True
        
        # Check if active stage
        deal["is_active_stage"] = deal["stage_normalized"] in ACTIVE_STAGES
        
        # Determine if has override
        deal["has_override"] = deal["custom_prob"] is not None and deal["custom_prob"] > 0
        
        # Calculate ACV from weighted
        if deal["has_override"] and deal["custom_prob"] > 0:
            deal["current_prob_used"] = deal["custom_prob"]
        elif deal["calc_prob"] and deal["calc_prob"] > 0:
            deal["current_prob_used"] = deal["calc_prob"]
        else:
            # Use standard probability from matrix
            ac = deal["acct_class"] or "New Logo"
            st = deal["stage_normalized"]
            deal["current_prob_used"] = CURRENT_PROBS.get(st, {}).get(ac, 0.33)
        
        if deal["weighted_acv"] > 0 and deal["current_prob_used"] > 0:
            deal["acv"] = deal["weighted_acv"] / deal["current_prob_used"]
        else:
            deal["acv"] = 0
        
        # Get new probability
        ac = deal["acct_class"] or "New Logo"
        st = deal["stage_normalized"]
        deal["new_std_prob"] = NEW_PROBS.get(st, {}).get(ac, 0.33)
        deal["old_std_prob"] = CURRENT_PROBS.get(st, {}).get(ac, 0.33)
        
        # Calculate new weighted
        if deal["has_override"]:
            # Override deals don't change
            deal["new_weighted"] = deal["weighted_acv"]
            deal["delta"] = 0
            deal["change_reason"] = "Override - no change"
        elif deal["stage_normalized"] not in ACTIVE_STAGES:
            # Non-active stages
            deal["new_weighted"] = deal["weighted_acv"]
            deal["delta"] = 0
            deal["change_reason"] = "Non-active stage"
        else:
            # Apply new probability
            deal["new_weighted"] = deal["acv"] * deal["new_std_prob"]
            deal["delta"] = deal["new_weighted"] - deal["weighted_acv"]
            if abs(deal["delta"]) < 0.01:
                deal["change_reason"] = "No change"
            elif deal["delta"] > 0:
                deal["change_reason"] = "Probability increased"
            else:
                deal["change_reason"] = "Probability decreased"
        
        all_deals.append(deal)
    
    print(f"\nTotal deals loaded: {len(all_deals)}")
    
    # Filter Q4 active deals
    q4_deals = [d for d in all_deals if d["is_q4"] and d["is_active_stage"]]
    print(f"Q4 deals (Nov 1, 2025 - Jan 31, 2026, Active Stages): {len(q4_deals)}")
    
    # Summary statistics
    current_total = sum(d["weighted_acv"] for d in q4_deals)
    new_total = sum(d["new_weighted"] for d in q4_deals)
    delta_total = new_total - current_total
    
    print(f"\nCurrent Q4 Weighted ACV: ${current_total:,.2f}")
    print(f"New Q4 Weighted ACV: ${new_total:,.2f}")
    print(f"Net Change: ${delta_total:+,.2f} ({delta_total/current_total*100:+.2f}%)" if current_total > 0 else "N/A")
    
    # Create output workbook
    out_wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    money_format = '"$"#,##0.00'
    pct_format = "0.0%"
    
    # ==================== SHEET 1: SUMMARY ====================
    summary_ws = out_wb.active
    summary_ws.title = "Summary"
    
    summary_ws["A1"] = "PROBABILITY IMPACT ANALYSIS - FINAL"
    summary_ws["A1"].font = Font(bold=True, size=16)
    
    summary_ws["A3"] = "Generated:"
    summary_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    summary_ws["A4"] = "Q4 FY2026 Range:"
    summary_ws["B4"] = f"{Q4_START.strftime('%b %d, %Y')} - {Q4_END.strftime('%b %d, %Y')}"
    
    summary_ws["A5"] = "Active Stages:"
    summary_ws["B5"] = "Stage 0-4 only (excludes Closed/Stage 5)"
    
    # Q4 totals
    summary_ws["A7"] = "Q4 WEIGHTED PIPELINE IMPACT"
    summary_ws["A7"].font = Font(bold=True, size=14)
    
    row = 8
    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        cell = summary_ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Q4 Deals Analyzed")
    summary_ws.cell(row=row, column=2, value=len(q4_deals))
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Current Weighted ACV")
    summary_ws.cell(row=row, column=2, value=current_total).number_format = money_format
    
    row += 1
    summary_ws.cell(row=row, column=1, value="New Weighted ACV")
    summary_ws.cell(row=row, column=2, value=new_total).number_format = money_format
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Net Change ($)")
    cell = summary_ws.cell(row=row, column=2, value=delta_total)
    cell.number_format = money_format
    cell.font = Font(bold=True, color="008000" if delta_total <= 0 else "FF0000")
    cell.fill = green_fill if delta_total <= 0 else red_fill
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Net Change (%)")
    pct_change = delta_total / current_total if current_total > 0 else 0
    cell = summary_ws.cell(row=row, column=2, value=pct_change)
    cell.number_format = pct_format
    cell.font = Font(bold=True, color="008000" if pct_change <= 0 else "FF0000")
    
    # By classification
    row += 2
    summary_ws.cell(row=row, column=1, value="IMPACT BY ACCOUNT TYPE CLASSIFICATION").font = Font(bold=True, size=12)
    
    row += 1
    headers = ["Classification", "Deals", "Current Weighted", "New Weighted", "Delta", "% Change"]
    for col, h in enumerate(headers, 1):
        cell = summary_ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    by_class = defaultdict(lambda: {"count": 0, "current": 0, "new": 0})
    for d in q4_deals:
        ac = d["acct_class"] or "New Logo"
        by_class[ac]["count"] += 1
        by_class[ac]["current"] += d["weighted_acv"]
        by_class[ac]["new"] += d["new_weighted"]
    
    row += 1
    for ac in ["New Logo", "Existing Client", "LOI", "Government"]:
        if ac in by_class:
            summary_ws.cell(row=row, column=1, value=ac)
            summary_ws.cell(row=row, column=2, value=by_class[ac]["count"])
            summary_ws.cell(row=row, column=3, value=by_class[ac]["current"]).number_format = money_format
            summary_ws.cell(row=row, column=4, value=by_class[ac]["new"]).number_format = money_format
            delta = by_class[ac]["new"] - by_class[ac]["current"]
            cell = summary_ws.cell(row=row, column=5, value=delta)
            cell.number_format = money_format
            if delta < 0:
                cell.fill = green_fill
            elif delta > 0:
                cell.fill = red_fill
            pct = delta / by_class[ac]["current"] if by_class[ac]["current"] > 0 else 0
            summary_ws.cell(row=row, column=6, value=pct).number_format = pct_format
            row += 1
    
    # Override vs non-override
    row += 1
    summary_ws.cell(row=row, column=1, value="OVERRIDE vs NON-OVERRIDE BREAKDOWN").font = Font(bold=True, size=12)
    
    override_deals = [d for d in q4_deals if d["has_override"]]
    non_override_deals = [d for d in q4_deals if not d["has_override"]]
    
    row += 1
    headers = ["Category", "Deals", "Weighted ACV", "Impact", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = summary_ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Override Deals")
    summary_ws.cell(row=row, column=2, value=len(override_deals))
    summary_ws.cell(row=row, column=3, value=sum(d["weighted_acv"] for d in override_deals)).number_format = money_format
    summary_ws.cell(row=row, column=4, value="$0.00")
    summary_ws.cell(row=row, column=5, value="Uses Custom Probability - NO CHANGE")
    
    row += 1
    summary_ws.cell(row=row, column=1, value="Non-Override Deals")
    summary_ws.cell(row=row, column=2, value=len(non_override_deals))
    summary_ws.cell(row=row, column=3, value=sum(d["weighted_acv"] for d in non_override_deals)).number_format = money_format
    non_override_delta = sum(d["delta"] for d in non_override_deals)
    cell = summary_ws.cell(row=row, column=4, value=non_override_delta)
    cell.number_format = money_format
    cell.fill = green_fill if non_override_delta <= 0 else red_fill
    summary_ws.cell(row=row, column=5, value="Subject to formula changes")
    
    # ==================== SHEET 2: PROBABILITY MATRIX ====================
    matrix_ws = out_wb.create_sheet("Probability Matrix")
    
    matrix_ws["A1"] = "PROBABILITY MATRIX COMPARISON"
    matrix_ws["A1"].font = Font(bold=True, size=14)
    
    # Current probabilities
    matrix_ws["A3"] = "CURRENT PROBABILITIES"
    matrix_ws["A3"].font = Font(bold=True, size=12)
    
    row = 4
    headers = ["Stage", "New Logo", "Existing Client", "LOI", "Government"]
    for col, h in enumerate(headers, 1):
        cell = matrix_ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    for stage in ACTIVE_STAGES:
        matrix_ws.cell(row=row, column=1, value=stage)
        for col, ac in enumerate(["New Logo", "Existing Client", "LOI", "Government"], 2):
            val = CURRENT_PROBS.get(stage, {}).get(ac, 0)
            matrix_ws.cell(row=row, column=col, value=val).number_format = pct_format
        row += 1
    
    # New probabilities
    row += 1
    matrix_ws.cell(row=row, column=1, value="NEW PROBABILITIES (Changes Highlighted)").font = Font(bold=True, size=12)
    
    row += 1
    for col, h in enumerate(headers, 1):
        cell = matrix_ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    for stage in ACTIVE_STAGES:
        matrix_ws.cell(row=row, column=1, value=stage)
        for col, ac in enumerate(["New Logo", "Existing Client", "LOI", "Government"], 2):
            new_val = NEW_PROBS.get(stage, {}).get(ac, 0)
            old_val = CURRENT_PROBS.get(stage, {}).get(ac, 0)
            cell = matrix_ws.cell(row=row, column=col, value=new_val)
            cell.number_format = pct_format
            if new_val != old_val:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
        row += 1
    
    # Summary of changes
    row += 2
    matrix_ws.cell(row=row, column=1, value="CHANGES SUMMARY").font = Font(bold=True, size=12)
    row += 1
    matrix_ws.cell(row=row, column=1, value="• LOI probabilities DECREASED by 3-7% at each stage")
    row += 1
    matrix_ws.cell(row=row, column=1, value="• Existing Client probabilities UNCHANGED")
    row += 1
    matrix_ws.cell(row=row, column=1, value="• New Logo probabilities UNCHANGED")
    row += 1
    matrix_ws.cell(row=row, column=1, value="• Government probabilities UNCHANGED")
    row += 1
    matrix_ws.cell(row=row, column=1, value="• Stage 5 (Negotiation) REMOVED")
    row += 1
    matrix_ws.cell(row=row, column=1, value="• Result: EC (50%) > LOI (48%) at Stage 4 - EC is now highest")
    
    # ==================== SHEET 3: Q4 DEAL DETAIL ====================
    detail_ws = out_wb.create_sheet("Q4 Deal Detail")
    
    headers = [
        "Row", "Opportunity Name", "Stage", "Classification", 
        "Sales Type", "Override?", "ACV", "Current Prob", "Current Weighted",
        "New Prob", "New Weighted", "Delta", "Change Reason"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = detail_ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for d in sorted(q4_deals, key=lambda x: abs(x.get("delta", 0)), reverse=True):
        detail_ws.cell(row=row, column=1, value=d["row"])
        detail_ws.cell(row=row, column=2, value=d["opp_name"])
        detail_ws.cell(row=row, column=3, value=d["stage"])
        detail_ws.cell(row=row, column=4, value=d["acct_class"])
        detail_ws.cell(row=row, column=5, value=d["sales_type"])
        detail_ws.cell(row=row, column=6, value="Yes" if d["has_override"] else "No")
        detail_ws.cell(row=row, column=7, value=d["acv"]).number_format = money_format
        detail_ws.cell(row=row, column=8, value=d["current_prob_used"]).number_format = pct_format
        detail_ws.cell(row=row, column=9, value=d["weighted_acv"]).number_format = money_format
        
        if d["has_override"]:
            detail_ws.cell(row=row, column=10, value="OVERRIDE")
        else:
            detail_ws.cell(row=row, column=10, value=d["new_std_prob"]).number_format = pct_format
        
        detail_ws.cell(row=row, column=11, value=d["new_weighted"]).number_format = money_format
        
        delta_cell = detail_ws.cell(row=row, column=12, value=d["delta"])
        delta_cell.number_format = money_format
        if d["delta"] < -0.01:
            delta_cell.fill = green_fill
        elif d["delta"] > 0.01:
            delta_cell.fill = red_fill
        
        detail_ws.cell(row=row, column=13, value=d["change_reason"])
        row += 1
    
    # Add totals row
    row += 1
    detail_ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    detail_ws.cell(row=row, column=9, value=current_total).number_format = money_format
    detail_ws.cell(row=row, column=9).font = Font(bold=True)
    detail_ws.cell(row=row, column=11, value=new_total).number_format = money_format
    detail_ws.cell(row=row, column=11).font = Font(bold=True)
    total_delta_cell = detail_ws.cell(row=row, column=12, value=delta_total)
    total_delta_cell.number_format = money_format
    total_delta_cell.font = Font(bold=True)
    total_delta_cell.fill = green_fill if delta_total <= 0 else red_fill
    
    # Adjust column widths
    for ws in [summary_ws, matrix_ws, detail_ws]:
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save output
    output_path = "/Users/keiganpesenti/Desktop/Probability_Impact_Analysis.xlsx"
    out_wb.save(output_path)
    
    print(f"\n✅ Output saved to: {output_path}")
    
    # Print final summary
    print("\n" + "="*70)
    print("FINAL SUMMARY")
    print("="*70)
    print(f"Q4 Date Range: {Q4_START.strftime('%b %d, %Y')} - {Q4_END.strftime('%b %d, %Y')}")
    print(f"Q4 Deals Analyzed: {len(q4_deals)}")
    print(f"  - Override Deals: {len(override_deals)} (${sum(d['weighted_acv'] for d in override_deals):,.2f}) - NO CHANGE")
    print(f"  - Non-Override: {len(non_override_deals)} (${sum(d['weighted_acv'] for d in non_override_deals):,.2f})")
    print()
    print(f"Current Q4 Weighted: ${current_total:,.2f}")
    print(f"New Q4 Weighted:     ${new_total:,.2f}")
    print(f"Net Change:          ${delta_total:+,.2f} ({delta_total/current_total*100:+.2f}%)" if current_total > 0 else "N/A")
    print()
    if delta_total <= 0:
        print("✅ NET CHANGE IS NEGATIVE OR ZERO - Pipeline does NOT increase")
    else:
        print("⚠️  NET CHANGE IS POSITIVE - Review required")
    print("="*70)

if __name__ == "__main__":
    analyze_pipeline()
