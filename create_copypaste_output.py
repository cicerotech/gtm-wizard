"""
Create Copy-Paste Ready Output
==============================
Matches exact column structure of existing Inputs_formula tab
to feed into Summary formulas
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from collections import defaultdict

print("="*80)
print("CREATING COPY-PASTE OUTPUT")
print("Matching existing Inputs_formula structure")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[1] Loading data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
print("\n[2] Classifying meetings...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

# =============================================================================
# BUILD OUTPUT MATCHING INPUTS_FORMULA STRUCTURE
# =============================================================================
print("\n[3] Building output matching Inputs_formula structure...")

# Target column order based on your screenshots:
# A: Owner
# B: Account  
# C: Include
# D: Verified
# E: TotalMeetings
# F: FirstMeetingDate
# G: CloseDate
# H: SalesCycleDays
# I: DaysFirstToSecond
# J: DaysToDemo
# K: DemoCount
# L: DaysToCAB
# M: CABCount
# N: DaysToScoping
# O: ScopingCount
# P: DaysToCompliance
# Q: ComplianceCount
# R: DaysToContract
# S: ContractingCount
# T: Notes

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
    return None

output_records = []

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) == 0:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    # SF match
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    # First meeting date
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min()
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        first_mtg = first_meeting_sf
    else:
        first_mtg = first_meeting_audit
    
    # Classify meetings
    meeting_classes = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_mtg) if first_mtg else (idx == 0)
        cls = classify_meeting(row['Subject'], is_first)
        days_from_first = (row['Date'] - first_mtg).days if first_mtg and pd.notna(row['Date']) else 0
        meeting_classes[cls].append({
            'date': row['Date'],
            'days': days_from_first
        })
    
    # Calculate metrics
    total_meetings = len(acct_meetings)
    
    # Days first to second
    if len(acct_meetings) >= 2:
        days_first_to_second = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        days_first_to_second = 0
    
    # Demo
    demos = meeting_classes.get('Demo', [])
    demo_count = len(demos)
    days_to_demo = min([d['days'] for d in demos]) if demos else 0
    
    # CAB
    cabs = meeting_classes.get('CAB', [])
    cab_count = len(cabs)
    days_to_cab = min([d['days'] for d in cabs]) if cabs else 0
    
    # Scoping
    scopings = meeting_classes.get('Scoping', [])
    scoping_count = len(scopings)
    days_to_scoping = min([d['days'] for d in scopings]) if scopings else 0
    
    # Compliance
    compliances = meeting_classes.get('Compliance', [])
    compliance_count = len(compliances)
    days_to_compliance = min([d['days'] for d in compliances]) if compliances else 0
    
    # Contracting
    contractings = meeting_classes.get('Contracting', [])
    contracting_count = len(contractings)
    days_to_contracting = min([d['days'] for d in contractings]) if contractings else 0
    
    # Close date
    if has_no_close(orig_name):
        close_date = ''
        sales_cycle = 0
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close'].strftime('%m/%d/%Y')
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            sales_cycle = int(sf_data['days_to_close'])
        elif first_mtg:
            sales_cycle = (sf_data['first_close'] - first_mtg).days
        else:
            sales_cycle = 0
    else:
        close_date = ''
        sales_cycle = 0
    
    output_records.append({
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Account': orig_name,
        'Include': 'Y',
        'Verified': 0,
        'TotalMeetings': total_meetings,
        'FirstMeetingDate': first_mtg.strftime('%m/%d/%Y') if first_mtg else '',
        'CloseDate': close_date,
        'SalesCycleDays': sales_cycle,
        'DaysFirstToSecond': days_first_to_second,
        'DaysToDemo': days_to_demo,
        'DemoCount': demo_count,
        'DaysToCAB': days_to_cab,
        'CABCount': cab_count,
        'DaysToScoping': days_to_scoping,
        'ScopingCount': scoping_count,
        'DaysToCompliance': days_to_compliance,
        'ComplianceCount': compliance_count,
        'DaysToContract': days_to_contracting,
        'ContractingCount': contracting_count,
        'Notes': '',
    })

output_df = pd.DataFrame(output_records)

# Sort by Owner (with owner first), then Account
output_df['_has_owner'] = output_df['Owner'].apply(lambda x: 0 if pd.notna(x) and x != '' else 1)
output_df = output_df.sort_values(['_has_owner', 'Owner', 'Account']).reset_index(drop=True)
output_df = output_df.drop(columns=['_has_owner'])

print(f"  Created {len(output_df)} records")

# =============================================================================
# GENERATE OUTPUT
# =============================================================================
print("\n[4] Generating Excel output...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

# === COPY_PASTE_TO_INPUTS tab (main output) ===
ws = wb.active
ws.title = 'COPY_PASTE_TO_INPUTS'

# Exact column order matching your Inputs_formula
columns = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount',
    'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance',
    'ComplianceCount', 'DaysToContract', 'ContractingCount', 'Notes'
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")

for col_idx, col_name in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font_white
    cell.fill = header_fill

for row_idx, (_, row) in enumerate(output_df.iterrows(), 2):
    for col_idx, col_name in enumerate(columns, 1):
        val = row.get(col_name, '')
        ws.cell(row=row_idx, column=col_idx, value=val)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12

# === SOURCE_MEETINGS tab for reference ===
ws_source = wb.create_sheet('SOURCE_MEETINGS')

source_data = []
for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) == 0:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    first_date = acct_meetings['Date'].min()
    
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_date)
        cls = classify_meeting(row['Subject'], is_first)
        days = (row['Date'] - first_date).days if pd.notna(row['Date']) else 0
        
        source_data.append({
            'Account': orig_name,
            'Date': row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',
            'Subject': str(row['Subject'])[:80] if pd.notna(row['Subject']) else '',
            'Classification': cls,
            'DaysFromFirst': days,
        })

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'DaysFromFirst']
for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font_white
    cell.fill = header_fill

for row_idx, row_data in enumerate(source_data, 2):
    for col_idx, col_name in enumerate(source_cols, 1):
        ws_source.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 60

# Save
output_path = '/Users/keiganpesenti/Desktop/COPY_PASTE_READY.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("HOW TO USE")
print("="*80)
print("""
1. Open COPY_PASTE_READY.xlsx
2. Go to 'COPY_PASTE_TO_INPUTS' tab
3. Select A2 through T{} (all data rows, no header)
4. Copy (Ctrl+C)
5. Go to your existing workbook
6. Go to 'Inputs_formula' or 'BL Review' tab
7. Select cell A2
8. Paste (Ctrl+V)
9. Your Summary formulas will auto-update
""".format(len(output_df) + 1))

print("\nCOLUMN MAPPING:")
print("-" * 50)
for i, col in enumerate(columns, 1):
    print(f"  Column {chr(64+i)}: {col}")

