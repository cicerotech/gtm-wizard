"""
Validated Sales Data Generator v6
=================================
- Formula-driven metrics from Source_Meetings
- Override columns for user corrections
- Final values use override if present, otherwise formula
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from collections import defaultdict

print("="*80)
print("VALIDATED SALES DATA GENERATOR v6")
print("Formula-driven with override capability")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# STEP 1: Load Reference Data
# =============================================================================
print("\n[STEP 1] Loading reference data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }
print(f"  SF reference: {len(sf_lookup)} accounts")

# =============================================================================
# STEP 2: Load Meeting Audit Data
# =============================================================================
print("\n[STEP 2] Loading meeting audit data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# STEP 3: Classify All Meetings (Source Data)
# =============================================================================
print("\n[STEP 3] Classifying all meetings...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

source_meetings = []
account_list = []

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    account_list.append(orig_name)
    first_date = acct_meetings['Date'].min()
    
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_date)
        classification = classify_meeting(row['Subject'], is_first)
        
        source_meetings.append({
            'Account': orig_name,
            'Date': row['Date'],
            'DateStr': row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',
            'Subject': str(row['Subject'])[:80] if pd.notna(row['Subject']) else '',
            'Classification': classification,
            'IsFirstMeeting': 'Y' if is_first else '',
            'DaysFromFirst': (row['Date'] - first_date).days if pd.notna(row['Date']) else '',
        })

source_df = pd.DataFrame(source_meetings)
print(f"  Created source data for {len(source_df)} meetings across {len(account_list)} accounts")

# =============================================================================
# STEP 4: Build Account List with SF Data
# =============================================================================
print("\n[STEP 4] Building account list...")

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
    return None

account_data = []
for orig_name in account_list:
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    # Get first meeting date
    acct_meetings = df_meetings[df_meetings['Company / Account'] == orig_name]
    first_meeting_audit = acct_meetings['Date'].min() if len(acct_meetings) > 0 else None
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        first_mtg = first_meeting_sf
    elif first_meeting_audit and pd.notna(first_meeting_audit):
        first_mtg = first_meeting_audit
    else:
        first_mtg = None
    
    # Close date
    if has_no_close(orig_name):
        close_date = None
        sales_cycle = ''
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            sales_cycle = int(sf_data['days_to_close'])
        elif first_mtg and close_date > first_mtg:
            sales_cycle = (close_date - first_mtg).days
        else:
            sales_cycle = ''
    else:
        close_date = None
        sales_cycle = ''
    
    account_data.append({
        'Account': orig_name,
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'FirstMeetingDate': first_mtg.strftime('%m/%d/%Y') if first_mtg else '',
        'CloseDate': close_date.strftime('%m/%d/%Y') if close_date else '',
        'SalesCycleDays': sales_cycle,
    })

account_df = pd.DataFrame(account_data)
account_df['_has_owner'] = account_df['Owner'].apply(lambda x: 0 if pd.notna(x) and x != '' else 1)
account_df = account_df.sort_values(['_has_owner', 'Owner', 'Account']).reset_index(drop=True)
account_df = account_df.drop(columns=['_has_owner'])

print(f"  {len(account_df)} accounts ready")

# =============================================================================
# STEP 5: Generate Excel with Formulas
# =============================================================================
print("\n[STEP 5] Generating Excel with formulas...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# === TAB 1: Source_Meetings (must be first for formulas to reference) ===
ws_source = wb.active
ws_source.title = 'Source_Meetings'

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'IsFirstMeeting', 'DaysFromFirst']

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

class_fills = {
    'Demo': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    'CAB': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'Scoping': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    'Compliance': PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
    'Contracting': PatternFill(start_color="D5A6BD", end_color="D5A6BD", fill_type="solid"),
}

for row_idx, (_, row) in enumerate(source_df.iterrows(), 2):
    ws_source.cell(row=row_idx, column=1, value=row['Account'])
    ws_source.cell(row=row_idx, column=2, value=row['DateStr'])
    ws_source.cell(row=row_idx, column=3, value=row['Subject'])
    ws_source.cell(row=row_idx, column=4, value=row['Classification'])
    ws_source.cell(row=row_idx, column=5, value=row['IsFirstMeeting'])
    ws_source.cell(row=row_idx, column=6, value=row['DaysFromFirst'])
    
    classification = row['Classification']
    if classification in class_fills:
        for col_idx in range(1, 7):
            ws_source.cell(row=row_idx, column=col_idx).fill = class_fills[classification]

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 60
ws_source.column_dimensions['D'].width = 15

source_last_row = len(source_df) + 1

# === TAB 2: BL_Review_Inputs with formulas ===
ws_inputs = wb.create_sheet('BL_Review_Inputs')

# Column structure:
# A: Account
# B: Owner
# C: Include
# D: Verified
# E: FirstMeetingDate
# F: TotalMeetings (formula)
# G: DemoCount_Calc (formula)
# H: DemoCount_Override (user input)
# I: DemoCount_Final (formula: uses override if present)
# J: DaysToDemo_Calc (formula)
# K: DaysToDemo_Override
# L: DaysToDemo_Final
# ... similar for CAB, Scoping, Compliance, Contracting
# Then CloseDate, SalesCycleDays at end

input_headers = [
    'Account', 'Owner', 'Include', 'Verified', 'FirstMeetingDate',
    'TotalMtgs_Calc', 'TotalMtgs_Override', 'TotalMtgs',
    'DemoCount_Calc', 'DemoCount_Override', 'DemoCount',
    'DaysToDemo_Calc', 'DaysToDemo_Override', 'DaysToDemo',
    'CABCount_Calc', 'CABCount_Override', 'CABCount',
    'DaysToCAB_Calc', 'DaysToCAB_Override', 'DaysToCAB',
    'ScopingCount_Calc', 'ScopingCount_Override', 'ScopingCount',
    'DaysToScoping_Calc', 'DaysToScoping_Override', 'DaysToScoping',
    'ComplianceCount_Calc', 'ComplianceCount_Override', 'ComplianceCount',
    'DaysToCompliance_Calc', 'DaysToCompliance_Override', 'DaysToCompliance',
    'ContractCount_Calc', 'ContractCount_Override', 'ContractCount',
    'DaysToContract_Calc', 'DaysToContract_Override', 'DaysToContract',
    'CloseDate', 'SalesCycleDays'
]

calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for formulas
override_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for override
final_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue for final

for col_idx, col_name in enumerate(input_headers, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    
    if '_Calc' in col_name:
        cell.fill = calc_fill
    elif '_Override' in col_name:
        cell.fill = override_fill
    elif col_name in ['TotalMtgs', 'DemoCount', 'DaysToDemo', 'CABCount', 'DaysToCAB', 
                      'ScopingCount', 'DaysToScoping', 'ComplianceCount', 'DaysToCompliance',
                      'ContractCount', 'DaysToContract']:
        cell.fill = final_fill

for row_idx, (_, row) in enumerate(account_df.iterrows(), 2):
    account_name = row['Account']
    
    # Static columns
    ws_inputs.cell(row=row_idx, column=1, value=account_name)  # Account
    ws_inputs.cell(row=row_idx, column=2, value=row['Owner'])  # Owner
    ws_inputs.cell(row=row_idx, column=3, value='Y')  # Include
    ws_inputs.cell(row=row_idx, column=4, value='')  # Verified
    ws_inputs.cell(row=row_idx, column=5, value=row['FirstMeetingDate'])  # FirstMeetingDate
    
    # TotalMtgs_Calc (F) - COUNTIF on Source_Meetings
    ws_inputs.cell(row=row_idx, column=6, 
                   value=f'=COUNTIF(Source_Meetings!A:A,A{row_idx})').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=7, value='').fill = override_fill  # TotalMtgs_Override
    # TotalMtgs_Final (H) - use override if present
    ws_inputs.cell(row=row_idx, column=8, 
                   value=f'=IF(G{row_idx}<>"",G{row_idx},F{row_idx})').fill = final_fill
    
    # DemoCount_Calc (I)
    ws_inputs.cell(row=row_idx, column=9, 
                   value=f'=COUNTIFS(Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Demo")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=10, value='').fill = override_fill  # DemoCount_Override
    ws_inputs.cell(row=row_idx, column=11, 
                   value=f'=IF(J{row_idx}<>"",J{row_idx},I{row_idx})').fill = final_fill
    
    # DaysToDemo_Calc (L) - MINIFS for first demo
    ws_inputs.cell(row=row_idx, column=12, 
                   value=f'=IFERROR(MINIFS(Source_Meetings!F:F,Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Demo"),"")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=13, value='').fill = override_fill  # DaysToDemo_Override
    ws_inputs.cell(row=row_idx, column=14, 
                   value=f'=IF(M{row_idx}<>"",M{row_idx},L{row_idx})').fill = final_fill
    
    # CABCount_Calc (O)
    ws_inputs.cell(row=row_idx, column=15, 
                   value=f'=COUNTIFS(Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"CAB")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=16, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=17, 
                   value=f'=IF(P{row_idx}<>"",P{row_idx},O{row_idx})').fill = final_fill
    
    # DaysToCAB_Calc (R)
    ws_inputs.cell(row=row_idx, column=18, 
                   value=f'=IFERROR(MINIFS(Source_Meetings!F:F,Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"CAB"),"")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=19, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=20, 
                   value=f'=IF(S{row_idx}<>"",S{row_idx},R{row_idx})').fill = final_fill
    
    # ScopingCount_Calc (U)
    ws_inputs.cell(row=row_idx, column=21, 
                   value=f'=COUNTIFS(Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Scoping")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=22, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=23, 
                   value=f'=IF(V{row_idx}<>"",V{row_idx},U{row_idx})').fill = final_fill
    
    # DaysToScoping_Calc (X)
    ws_inputs.cell(row=row_idx, column=24, 
                   value=f'=IFERROR(MINIFS(Source_Meetings!F:F,Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Scoping"),"")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=25, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=26, 
                   value=f'=IF(Y{row_idx}<>"",Y{row_idx},X{row_idx})').fill = final_fill
    
    # ComplianceCount_Calc (AA)
    ws_inputs.cell(row=row_idx, column=27, 
                   value=f'=COUNTIFS(Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Compliance")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=28, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=29, 
                   value=f'=IF(AB{row_idx}<>"",AB{row_idx},AA{row_idx})').fill = final_fill
    
    # DaysToCompliance_Calc (AD)
    ws_inputs.cell(row=row_idx, column=30, 
                   value=f'=IFERROR(MINIFS(Source_Meetings!F:F,Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Compliance"),"")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=31, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=32, 
                   value=f'=IF(AE{row_idx}<>"",AE{row_idx},AD{row_idx})').fill = final_fill
    
    # ContractCount_Calc (AG)
    ws_inputs.cell(row=row_idx, column=33, 
                   value=f'=COUNTIFS(Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Contracting")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=34, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=35, 
                   value=f'=IF(AH{row_idx}<>"",AH{row_idx},AG{row_idx})').fill = final_fill
    
    # DaysToContract_Calc (AJ)
    ws_inputs.cell(row=row_idx, column=36, 
                   value=f'=IFERROR(MINIFS(Source_Meetings!F:F,Source_Meetings!A:A,A{row_idx},Source_Meetings!D:D,"Contracting"),"")').fill = calc_fill
    ws_inputs.cell(row=row_idx, column=37, value='').fill = override_fill
    ws_inputs.cell(row=row_idx, column=38, 
                   value=f'=IF(AK{row_idx}<>"",AK{row_idx},AJ{row_idx})').fill = final_fill
    
    # CloseDate, SalesCycleDays (static)
    ws_inputs.cell(row=row_idx, column=39, value=row['CloseDate'])
    ws_inputs.cell(row=row_idx, column=40, value=row['SalesCycleDays'])

ws_inputs.column_dimensions['A'].width = 25
ws_inputs.column_dimensions['B'].width = 18

# === TAB 3: Instructions ===
ws_instructions = wb.create_sheet('Instructions')

instructions = [
    ['HOW TO USE THIS FILE', '', ''],
    ['', '', ''],
    ['COLUMN COLOR LEGEND:', '', ''],
    ['Green (_Calc)', 'Formula-calculated from Source_Meetings - DO NOT EDIT', ''],
    ['Yellow (_Override)', 'Enter your correction here if the formula is wrong', ''],
    ['Blue (Final)', 'Shows Override if entered, otherwise shows Calc value', ''],
    ['', '', ''],
    ['WORKFLOW:', '', ''],
    ['1.', 'Review the blue "Final" columns for each account', ''],
    ['2.', 'If a value looks wrong, check Source_Meetings tab to see why', ''],
    ['3.', 'To correct: enter the right value in the yellow Override column', ''],
    ['4.', 'The blue Final column will automatically use your override', ''],
    ['', '', ''],
    ['EXAMPLE - Correcting DemoCount:', '', ''],
    ['', 'If Coherent shows DemoCount=6 but you know it should be 8:', ''],
    ['', '1. Go to Coherent row, find DemoCount_Override (yellow column)', ''],
    ['', '2. Enter "8" in that cell', ''],
    ['', '3. DemoCount (blue) will now show 8 instead of 6', ''],
    ['', '', ''],
    ['TO ADD A MISSING MEETING:', '', ''],
    ['', '1. Go to Source_Meetings tab', ''],
    ['', '2. Add a new row with Account, Date, Subject, Classification', ''],
    ['', '3. The formulas will automatically recalculate', ''],
    ['', '', ''],
    ['CLASSIFICATION KEYWORDS:', '', ''],
    ['Demo:', 'demo, sigma, cortex, platform, walkthrough, product overview', ''],
    ['CAB:', 'cab, customer advisory, advisory board', ''],
    ['Scoping:', 'scoping, scope, pricing, proposal', ''],
    ['Compliance:', 'infosec, security, compliance', ''],
    ['Contracting:', 'contract, redline, msa, negotiation, legal', ''],
    ['Intro:', 'First meeting in sequence OR intro, introduction in subject', ''],
    ['Followup:', 'Default if no keywords match', ''],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_instructions.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif 'Green' in str(val):
            cell.fill = calc_fill
        elif 'Yellow' in str(val):
            cell.fill = override_fill
        elif 'Blue' in str(val):
            cell.fill = final_fill

ws_instructions.column_dimensions['A'].width = 20
ws_instructions.column_dimensions['B'].width = 60

# Save
output_path = '/Users/keiganpesenti/Desktop/validated_inputs_v6.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total accounts: {len(account_df)}")
print(f"Total source meetings: {len(source_df)}")
print("\nColumn structure for each metric:")
print("  _Calc (Green)    = Formula from Source_Meetings - don't edit")
print("  _Override (Yellow) = User can enter correction here")
print("  Final (Blue)     = Uses Override if present, otherwise Calc")

