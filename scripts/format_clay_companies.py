#!/usr/bin/env python3
"""
Format Clay Companies Excel file with:
- Black header with white text
- Times New Roman, 11pt, black text
- Combined summary column with bullets
- Spacer rows between companies
- No duplicate entries
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# File paths
INPUT_FILE = "/Users/keiganpesenti/Desktop/clay companies list.xlsx"
OUTPUT_FILE = "/Users/keiganpesenti/Desktop/clay_companies_formatted.xlsx"

# Values to filter out (metadata, not real content)
FILTER_VALUES = {'response', '', 'none', 'n/a'}

def should_include(value):
    """Check if a value should be included in the summary."""
    if not value:
        return False
    normalized = str(value).strip().lower()
    return normalized not in FILTER_VALUES

def load_workbook_data(filepath):
    """Load and extract data from the source Excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    companies = []
    current_company = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        # Col 0: Account Name
        company_name = row[0] if len(row) > 0 else None
        
        # Col 6: Legal Entity Name (not Col 1 which is Website)
        legal_entity = row[6] if len(row) > 6 else None
        
        # Collect summary info from columns 3, 4, 5 only (indices 3, 4, 5)
        # Skip column 2 (just "Response" marker)
        # Skip columns 7, 8 (aggregated - duplicates of 3, 4, 5)
        summary_parts = []
        for idx in [3, 4, 5]:
            if len(row) > idx and row[idx]:
                val = str(row[idx]).strip()
                if should_include(val):
                    summary_parts.append(val)
        
        if company_name:  # New company entry
            if current_company:
                companies.append(current_company)
            current_company = {
                'name': str(company_name).strip(),
                'legal_entity': str(legal_entity).strip() if legal_entity else '',
                'summaries': summary_parts
            }
        elif current_company and summary_parts:
            # Additional rows for current company
            current_company['summaries'].extend(summary_parts)
    
    # Don't forget the last company
    if current_company:
        companies.append(current_company)
    
    wb.close()
    return companies

def deduplicate_summaries(summaries):
    """Remove duplicate entries while preserving order."""
    seen = set()
    unique = []
    for item in summaries:
        # Normalize for comparison (lowercase, stripped, first 100 chars)
        normalized = item.lower().strip()[:100]
        if normalized not in seen:
            seen.add(normalized)
            unique.append(item)
    return unique

def format_summary_with_bullets(summaries):
    """Format summaries with bullet points, removing duplicates."""
    unique_summaries = deduplicate_summaries(summaries)
    if not unique_summaries:
        return ""
    
    # Add bullet points to each line
    bulleted = ["• " + s for s in unique_summaries]
    return "\n".join(bulleted)

def create_formatted_workbook(companies):
    """Create a new formatted workbook with the company data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    
    # Define styles
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    body_font = Font(name='Times New Roman', size=11, color='000000')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Company Name', 'Legal Entity', 'Summary']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    
    # Add company data with spacer rows
    current_row = 2
    for i, company in enumerate(companies):
        # Format summary with bullets and deduplication
        summary = format_summary_with_bullets(company['summaries'])
        
        # Company name
        cell_a = ws.cell(row=current_row, column=1, value=company['name'])
        cell_a.font = body_font
        cell_a.alignment = wrap_alignment
        cell_a.border = thin_border
        
        # Legal entity
        cell_b = ws.cell(row=current_row, column=2, value=company['legal_entity'])
        cell_b.font = body_font
        cell_b.alignment = wrap_alignment
        cell_b.border = thin_border
        
        # Summary with bullets
        cell_c = ws.cell(row=current_row, column=3, value=summary)
        cell_c.font = body_font
        cell_c.alignment = wrap_alignment
        cell_c.border = thin_border
        
        # Calculate row height based on number of lines
        num_lines = max(1, summary.count('\n') + 1)
        ws.row_dimensions[current_row].height = max(30, num_lines * 15)
        
        current_row += 1
        
        # Add spacer row (except after last company)
        if i < len(companies) - 1:
            ws.row_dimensions[current_row].height = 10
            current_row += 1
    
    return wb

def main():
    print(f"Reading from: {INPUT_FILE}")
    
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: File not found: {INPUT_FILE}")
        return
    
    # Load data
    companies = load_workbook_data(INPUT_FILE)
    print(f"Found {len(companies)} companies")
    
    # Create formatted workbook
    wb = create_formatted_workbook(companies)
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"Saved formatted file to: {OUTPUT_FILE}")
    
    # Preview first few entries
    print("\n--- Preview of formatted data ---")
    for company in companies[:3]:
        print(f"\nCompany: {company['name']}")
        print(f"Legal Entity: {company['legal_entity']}")
        unique = deduplicate_summaries(company['summaries'])
        print(f"Summary points ({len(unique)} unique):")
        for s in unique:
            print(f"  • {s[:70]}...")

if __name__ == "__main__":
    main()
