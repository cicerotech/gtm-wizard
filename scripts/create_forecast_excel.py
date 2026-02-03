#!/usr/bin/env python3
"""
Create Q1 2026 Forecast Excel Workbook with formulas
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Read the raw data (try different encodings)
try:
    df = pd.read_csv('/Users/keiganpesenti/revops_weekly_update/gtm-brain/data/opps_to_review.csv', encoding='utf-8')
except UnicodeDecodeError:
    try:
        df = pd.read_csv('/Users/keiganpesenti/revops_weekly_update/gtm-brain/data/opps_to_review.csv', encoding='latin-1')
    except:
        df = pd.read_csv('/Users/keiganpesenti/revops_weekly_update/gtm-brain/data/opps_to_review.csv', encoding='cp1252')

# Clean column names
df.columns = df.columns.str.strip()

# Create workbook
wb = Workbook()

# Styles
header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
normal_font = Font(name='Times New Roman', size=11, color='000000')
bold_font = Font(name='Times New Roman', size=11, bold=True, color='000000')
title_font = Font(name='Times New Roman', size=14, bold=True, color='000000')
currency_format = '"$"#,##0.00'
percent_format = '0%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================
# SHEET 1: Raw Data
# ============================================
ws_raw = wb.active
ws_raw.title = "Raw Data"

# Write headers and data
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

# Auto-width columns
for col in range(1, len(df.columns) + 1):
    ws_raw.column_dimensions[get_column_letter(col)].width = 18

# Get data range
data_rows = len(df) + 1  # +1 for header

# ============================================
# SHEET 2: Summary
# ============================================
ws_summary = wb.create_sheet("Summary")

# Title
ws_summary['A1'] = "Q1 2026 FORECAST SNAPSHOT"
ws_summary['A1'].font = title_font
ws_summary['A2'] = "February 2, 2026 | All formulas reference Raw Data tab"
ws_summary['A2'].font = Font(name='Times New Roman', size=10, italic=True)

# Summary Metrics Section
row = 4
ws_summary.cell(row=row, column=1, value="SUMMARY METRICS").font = bold_font

row = 5
headers = ["Metric", "Total (All)", "AI-Enabled Only", "AI % of Total"]
for c, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Column mapping from actual CSV:
# D = Net ACV
# E = BL Forecast Category  
# F = Quarterly Forecast Net
# G = Quarterly Forecast Net (AI-Enabled)
# H = Weighted ACV
# I = Weighted ACV (AI-Enabled)
# J = Quarterly Commit (Net New)
# K = Quarterly Commit Net (AI-Enabled)
# L = Blended Forecast (base)
# M = Blended Forecast (AI-Enabled)
# N = Pod
# O = Opportunity Owner

# Metrics with formulas
metrics = [
    ("Total Records", f"=COUNTA('Raw Data'!A:A)-1", "-", "-"),
    ("Net ACV (Pipeline)", f"=SUM('Raw Data'!D:D)", "-", "-"),
    ("Quarterly Forecast Net", f"=SUM('Raw Data'!F:F)", f"=SUM('Raw Data'!G:G)", f"=C8/B8"),
    ("Weighted ACV", f"=SUM('Raw Data'!H:H)", f"=SUM('Raw Data'!I:I)", f"=C9/B9"),
    ("Quarterly Commit (Net New)", f"=SUM('Raw Data'!J:J)", f"=SUM('Raw Data'!K:K)", f"=C10/B10"),
    ("Blended Forecast (base)", f"=SUM('Raw Data'!L:L)", f"=SUM('Raw Data'!M:M)", f"=C11/B11"),
]

for i, (label, total, ai, pct) in enumerate(metrics, 6):
    ws_summary.cell(row=i, column=1, value=label).font = normal_font
    ws_summary.cell(row=i, column=2, value=total).font = normal_font
    ws_summary.cell(row=i, column=3, value=ai).font = normal_font
    ws_summary.cell(row=i, column=4, value=pct).font = normal_font
    for c in range(1, 5):
        ws_summary.cell(row=i, column=c).border = thin_border
    if label == "Quarterly Commit (Net New)":
        ws_summary.cell(row=i, column=1).font = bold_font
        ws_summary.cell(row=i, column=2).font = bold_font
    # Format currency
    if i > 6:
        ws_summary.cell(row=i, column=2).number_format = currency_format
        if ai != "-":
            ws_summary.cell(row=i, column=3).number_format = currency_format
        if pct != "-":
            ws_summary.cell(row=i, column=4).number_format = percent_format

# By Pod Section
row = 14
ws_summary.cell(row=row, column=1, value="BY POD BREAKDOWN").font = bold_font

row = 15
pod_headers = ["Pod", "Opps", "Net ACV", "Forecast Net", "Forecast (AI-E)", "Weighted", "Weighted (AI-E)", "Commit (Net New)", "Blended", "Blended (AI-E)"]
for c, h in enumerate(pod_headers, 1):
    cell = ws_summary.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# US Row
row = 16
us_formulas = [
    "US",
    f"=COUNTIF('Raw Data'!N:N,\"US\")",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!D:D)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!F:F)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!G:G)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!H:H)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!I:I)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!J:J)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!L:L)",
    f"=SUMIF('Raw Data'!N:N,\"US\",'Raw Data'!M:M)",
]
for c, val in enumerate(us_formulas, 1):
    cell = ws_summary.cell(row=row, column=c, value=val)
    cell.font = normal_font
    cell.border = thin_border
    if c >= 3:
        cell.number_format = currency_format

# EU Row
row = 17
eu_formulas = [
    "EU",
    f"=COUNTIF('Raw Data'!N:N,\"EU\")",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!D:D)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!F:F)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!G:G)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!H:H)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!I:I)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!J:J)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!L:L)",
    f"=SUMIF('Raw Data'!N:N,\"EU\",'Raw Data'!M:M)",
]
for c, val in enumerate(eu_formulas, 1):
    cell = ws_summary.cell(row=row, column=c, value=val)
    cell.font = normal_font
    cell.border = thin_border
    if c >= 3:
        cell.number_format = currency_format

# Total Row
row = 18
total_formulas = [
    "TOTAL",
    f"=SUM(B16:B17)",
    f"=SUM(C16:C17)",
    f"=SUM(D16:D17)",
    f"=SUM(E16:E17)",
    f"=SUM(F16:F17)",
    f"=SUM(G16:G17)",
    f"=SUM(H16:H17)",
    f"=SUM(I16:I17)",
    f"=SUM(J16:J17)",
]
for c, val in enumerate(total_formulas, 1):
    cell = ws_summary.cell(row=row, column=c, value=val)
    cell.font = bold_font
    cell.border = thin_border
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    if c >= 3:
        cell.number_format = currency_format

# Forecast Ranges Section
row = 21
ws_summary.cell(row=row, column=1, value="CONSERVATIVE FORECAST RANGES").font = bold_font

row = 22
range_headers = ["Call", "Amount", "Confidence", "Description"]
for c, h in enumerate(range_headers, 1):
    cell = ws_summary.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

ranges = [
    ("Floor (Commit)", "=B10", "90%+", "BL-committed deals - highest confidence"),
    ("Most Likely", "=AVERAGE(B10,B9)", "70-80%", "Between commit and weighted"),
    ("Expected (Weighted)", "=B9", "60-70%", "Statistical probability-adjusted"),
    ("Target (Blended)", "=B11", "50-60%", "Combo of BL forecast + weighted"),
    ("Upside (Forecast)", "=B8", "40-50%", "Full BL-submitted pipeline"),
]

for i, (call, amount, conf, desc) in enumerate(ranges, 23):
    ws_summary.cell(row=i, column=1, value=call).font = normal_font
    ws_summary.cell(row=i, column=2, value=amount).font = normal_font
    ws_summary.cell(row=i, column=2).number_format = currency_format
    ws_summary.cell(row=i, column=3, value=conf).font = normal_font
    ws_summary.cell(row=i, column=4, value=desc).font = normal_font
    for c in range(1, 5):
        ws_summary.cell(row=i, column=c).border = thin_border
    if call == "Floor (Commit)":
        for c in range(1, 5):
            ws_summary.cell(row=i, column=c).fill = PatternFill(start_color='E8F4E8', end_color='E8F4E8', fill_type='solid')
            ws_summary.cell(row=i, column=c).font = bold_font

# Key Ratios Section
row = 30
ws_summary.cell(row=row, column=1, value="KEY RATIOS").font = bold_font

row = 31
ratio_headers = ["Ratio", "Formula", "Value"]
for c, h in enumerate(ratio_headers, 1):
    cell = ws_summary.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

ratios = [
    ("Commit / Weighted", "=B10/B9", "=B10/B9"),
    ("Commit / Forecast", "=B10/B8", "=B10/B8"),
    ("Weighted / Forecast", "=B9/B8", "=B9/B8"),
]

for i, (ratio, formula, value) in enumerate(ratios, 32):
    ws_summary.cell(row=i, column=1, value=ratio).font = normal_font
    ws_summary.cell(row=i, column=2, value=formula).font = Font(name='Times New Roman', size=9, italic=True)
    ws_summary.cell(row=i, column=3, value=value).font = normal_font
    ws_summary.cell(row=i, column=3).number_format = percent_format
    for c in range(1, 4):
        ws_summary.cell(row=i, column=c).border = thin_border

# Column widths for Summary
ws_summary.column_dimensions['A'].width = 28
ws_summary.column_dimensions['B'].width = 18
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 40
for col in ['E', 'F', 'G', 'H', 'I', 'J']:
    ws_summary.column_dimensions[col].width = 15

# ============================================
# SHEET 3: By Rep
# ============================================
ws_rep = wb.create_sheet("By Rep")

ws_rep['A1'] = "BY REP BREAKDOWN"
ws_rep['A1'].font = title_font

row = 3
rep_headers = ["Owner", "Pod", "Opps", "Net ACV", "Forecast Net", "Forecast (AI-E)", "Weighted", "Weighted (AI-E)", "Commit (Net New)", "Blended", "Blended (AI-E)"]
for c, h in enumerate(rep_headers, 1):
    cell = ws_rep.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Get unique owners
owners = df['Opportunity Owner'].str.strip().unique()

for i, owner in enumerate(owners, 4):
    owner_clean = owner.strip() if pd.notna(owner) else ""
    pod = df[df['Opportunity Owner'].str.strip() == owner_clean]['Pod'].iloc[0] if len(df[df['Opportunity Owner'].str.strip() == owner_clean]) > 0 else ""
    
    formulas = [
        owner_clean,
        pod,
        f"=COUNTIF('Raw Data'!O:O,\"{owner_clean}\")",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!D:D)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!F:F)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!G:G)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!H:H)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!I:I)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!J:J)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!L:L)",
        f"=SUMIF('Raw Data'!O:O,\"{owner_clean}\",'Raw Data'!M:M)",
    ]
    
    for c, val in enumerate(formulas, 1):
        cell = ws_rep.cell(row=i, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if c >= 4:
            cell.number_format = currency_format

# Column widths
ws_rep.column_dimensions['A'].width = 22
ws_rep.column_dimensions['B'].width = 8
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws_rep.column_dimensions[col].width = 15

# Save
output_path = '/Users/keiganpesenti/Desktop/Q1_2026_Forecast_Workbook.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Raw data rows: {len(df)}")
print(f"Columns: {list(df.columns)}")
