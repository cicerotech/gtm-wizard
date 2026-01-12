import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy

# Load the workbook
wb = openpyxl.load_workbook('/Users/keiganpesenti/Desktop/TopCo Pipeline Data.xlsx')
ws = wb['Month-over-Month']

# Find the notes row
notes_row = None
for row in range(1, 50):
    cell_val = ws.cell(row=row, column=1).value
    if cell_val and 'Notes' in str(cell_val) and 'Total' not in str(cell_val):
        notes_row = row
        break

print(f"Found Notes row: {notes_row}")

if notes_row:
    # Insert a new row below for originals
    ws.insert_rows(notes_row + 1)
    
    # Label the original row
    ws.cell(row=notes_row + 1, column=1).value = "Original Notes"
    ws.cell(row=notes_row + 1, column=1).font = Font(italic=True, color="808080")
    
    # Copy originals to the new row and update main row with clean versions
    clean_notes = {
        'Sept': 'Closed $3.1m incl. ESB ($493k), Coimisín na Meán ($404k)',
        'Oct': 'Closed $2.3m incl. Coherent ($1.1m), DHL ($250k), Bayer ($300k)',
        'Nov': 'Closed $1.5m incl. IQVIA ($300k), Delinea ($120k), Asana ($75k). Cargill ACV revised down ~$660k.',
        'Dec': 'Closed $2.8m incl. ServiceNow ($500k), Cargill ($521k), Intuit ($485k), Peregrine ($275k)'
    }
    
    # Map columns to months based on header row
    col_to_month = {}
    for col in range(2, 10):
        header = ws.cell(row=1, column=col).value
        if header:
            header_str = str(header)
            if 'Sep' in header_str:
                col_to_month[col] = 'Sept'
            elif 'Oct' in header_str:
                col_to_month[col] = 'Oct'
            elif 'Nov' in header_str:
                col_to_month[col] = 'Nov'
            elif 'Dec' in header_str:
                col_to_month[col] = 'Dec'
    
    print(f"Column mapping: {col_to_month}")
    
    for col in range(2, 10):
        original_val = ws.cell(row=notes_row, column=col).value
        if original_val:
            # Copy original to new row (greyed out)
            ws.cell(row=notes_row + 1, column=col).value = original_val
            ws.cell(row=notes_row + 1, column=col).font = Font(italic=True, color="808080", size=9)
            
            # Update main row with clean version if we have one
            month = col_to_month.get(col)
            if month and month in clean_notes:
                ws.cell(row=notes_row, column=col).value = clean_notes[month]
                print(f"Updated column {col} ({month})")

# Also find and clean up the methodology note
for row in range(1, 60):
    cell_val = ws.cell(row=row, column=1).value
    if cell_val and 'Total Wtd' in str(cell_val):
        # Check next columns for the methodology note
        for col in range(2, 5):
            val = ws.cell(row=row, column=col).value
            if val and 'Weighted pipeline' in str(val):
                # Insert row for original
                ws.insert_rows(row + 1)
                ws.cell(row=row + 1, column=1).value = "Original Methodology"
                ws.cell(row=row + 1, column=1).font = Font(italic=True, color="808080")
                ws.cell(row=row + 1, column=col).value = val
                ws.cell(row=row + 1, column=col).font = Font(italic=True, color="808080", size=9)
                
                # Clean version
                ws.cell(row=row, column=col).value = "Weighted pipeline reflects closed-won/lost activity and ACV adjustments. Historical data restated to align with current methodology."
                print(f"Updated methodology note at row {row}")
                break
        break

# Save to new file
output_path = '/Users/keiganpesenti/Desktop/TopCo Pipeline Data - Investor Clean.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")
print("Original notes preserved in greyed-out rows below clean versions")

