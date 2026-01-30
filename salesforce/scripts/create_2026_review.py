#!/usr/bin/env python3
"""Quick script to create 2026 Marketing Review Excel from screenshot data."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

wb = Workbook()
ws = wb.active
ws.title = "2026 Marketing Review"

# Headers
headers = ["Category", "Account", "Amount", "Deal", "", "", "", "", "", "Category", "Account", "Amount", "Deal"]
ws.append(headers)

# Main data (from screenshot)
main_data = [
    ["COMMITMENT (LOI)", "Bayer", 3000000, "Bayer - LOI"],
    ["COMMITMENT (LOI)", "Dolby", 2000000, "Dolby - LOI"],
    ["COMMITMENT (LOI)", "The Weir Group PLC", 1000000, "Weir Group - LOI"],
    ["COMMITMENT (LOI)", "Cox Media Group", 350000, "Cox Media Group - LOI"],
    ["COMMITMENT (LOI)", "AES", 333333, "AES - LOI"],
    ["COMMITMENT (LOI)", "Western Digital", 250000, "Western Digital - LOI"],
    ["COMMITMENT (LOI)", "CHS", 250000, "CHS - LOI"],
    ["COMMITMENT (LOI)", "GE Vernova", 200000, "GE Vernova - LOI"],
    ["COMMITMENT (LOI)", "Asana", 65000, "Asana - LOI"],
    ["COMMITMENT (LOI)", "Delinea", 0, "Delinea - LOI"],
    ["COMMITMENT (LOI)", "BNY Mellon", 0, "BNY Mellon - LOI"],
    ["COMMITMENT (LOI)", "Tailored Brands", 0, "Tailored Brands - LOI"],
    ["COMMITMENT (LOI)", "Amazon", 0, "Amazon - EL"],
    ["COMMITMENT (LOI)", "DHL", 0, "DHL - LOI"],
    ["COMMITMENT (LOI)", "Wealth Partners", 0, "Wealth Partners - LOI"],
    ["", "", "", ""],
    ["RECURRING REVENUE", "IQVIA", 300000, "IQVIA - Recurring"],
    ["RECURRING REVENUE", "Bayer", 300000, "Bayer - Marketing Compliance"],
    ["RECURRING REVENUE", "Peregrine Hospitality", 275000, "Peregrine Hospitality - Contracting"],
    ["RECURRING REVENUE", "DHL", 250000, "DHL - Insights"],
    ["RECURRING REVENUE", "CHS", 120000, "CHS - Contracting"],
    ["RECURRING REVENUE", "Delinea", 120000, "Delinea - Recurring (sigma)"],
    ["RECURRING REVENUE", "Duracell", 120000, "Duracell - G2N"],
    ["RECURRING REVENUE", "Samsara", 40000, "Samsara - Negotiation"],
    ["", "", "", ""],
    ["PROJECT REVENUE", "Intuit", 75000, "Intuit - Managed Services"],
    ["PROJECT REVENUE", "Asana", 75000, "Asana - Contracting & Insights Pilot"],
    ["PROJECT REVENUE", "Peregrine Hospitality", 42000, "Peregrine - Eudia Counsel"],
    ["PROJECT REVENUE", "Chevron", 0, "Chevron - Marketing Compliance"],
    ["", "", "", ""],
    ["TOTAL COMMITMENT", "", 7448333, ""],
    ["TOTAL REVENUE", "", 1525000, ""],
    ["GRAND TOTAL", "", 192000, ""],
]

# Isolated clients (yellow highlighted - excluded from totals)
isolated_data = [
    ["RECURRING", "Coherent", 1150000, "Coherent - Contracting"],
    ["RECURRING", "Cargill", 521000, "Cargill - Contracting"],
    ["RECURRING", "Intuit", 410000, "Intuit - Contracting, Marketing, Sigma"],
    ["RECURRING", "Meta", 350000, "Meta - 2026 Extension"],
    ["PROJECT", "Cargill", 150000, "Cargill - Cortex Pilot"],
]

# Write main data
for i, row in enumerate(main_data, start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

# Write isolated data (columns J-M)
for i, row in enumerate(isolated_data, start=2):
    for j, val in enumerate(row):
        cell = ws.cell(row=i, column=10+j, value=val)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Format headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col in range(1, 14):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

# Column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 35
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 18
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 40

# Add EOY Review Answers sheet
ws2 = wb.create_sheet("EOY Review Answers")
answers = [
    ["MARKETING EOY 2026 REVIEW - SPECIFIC ANSWERS", ""],
    ["", ""],
    ["AI SUPPER CLUB SERIES", ""],
    ["Revenue Contribution (Actual):", "$1,525,000"],
    ["Direct Close (ARR):", "$1,525,000"],
    ["Commitment (LOI) from attendees:", "$7,448,333"],
    ["", ""],
    ["LIGHTHOUSE EVENT", ""],
    ["Revenue Contribution:", "$485,000"],
    ["Recurring Revenue Closed:", "$410,000 (Intuit isolated)"],
    ["Project Revenue:", "$75,000 (Intuit Managed Services)"],
    ["Note:", "Intuit & Pure Storage converted within 12 months"],
    ["", ""],
    ["AUGMENTED INTELLIGENCE SUMMIT", ""],
    ["Recurring Revenue:", "$1,525,000"],
    ["Project Revenue:", "$192,000"],
    ["Commitment (LOI) Dollars:", "$7,448,333"],
    ["Net-New Opportunities:", "65"],
    ["", ""],
    ["IQPC (Bayer Sourced)", ""],
    ["Revenue from Bayer:", "$300,000 (Recurring)"],
    ["Commitment from Bayer:", "$3,000,000 (LOI)"],
    ["", ""],
    ["ISOLATED CLIENTS (Reference Only - Not in Totals)", ""],
    ["Coherent:", "$1,150,000 (Recurring - Contracting)"],
    ["Cargill:", "$671,000 ($521K Recurring + $150K Project)"],
    ["Intuit:", "$410,000 (Recurring - Contracting, Marketing, Sigma)"],
    ["Meta:", "$350,000 (Recurring - 2026 Extension)"],
    ["Isolated Total:", "$2,581,000"],
]

for row in answers:
    ws2.append(row)

ws2['A1'].font = Font(bold=True, size=14)
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 50

output_path = os.path.expanduser("~/Desktop/MARKETING_2026_EOY_REVIEW.xlsx")
wb.save(output_path)
print(f"✅ Saved: {output_path}")
