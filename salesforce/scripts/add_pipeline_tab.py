#!/usr/bin/env python3
"""Add active pipeline analysis tab to Marketing 2026 Review."""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import re

# Load the original marketing data
excel_path = os.path.expanduser("~/Desktop/MARKETING DATA REFERENCE.xlsx")
attendees = pd.read_excel(excel_path, sheet_name="all campaign attendees")
active = pd.read_excel(excel_path, sheet_name="active in pipeline")

print(f"Campaign attendees: {len(attendees)}")
print(f"Active opportunities: {len(active)}")

# Normalize company names for matching
def normalize(name):
    if pd.isna(name):
        return ""
    return str(name).strip().lower()

# Extract account name from opportunity name (e.g., "Vista Equity Partners - Proposal" → "vista equity partners")
def extract_account(opp_name):
    if pd.isna(opp_name):
        return ""
    # Split on common patterns: " - ", " – ", etc.
    parts = re.split(r'\s*[-–]\s*', str(opp_name))
    if parts:
        return normalize(parts[0])
    return normalize(opp_name)

attendees['Company_Norm'] = attendees['Company'].apply(normalize)
active['Account_Norm'] = active['Opportunity Name'].apply(extract_account)

# Create unique company list from campaigns
campaign_companies = set(attendees['Company_Norm'].unique())
print(f"Unique campaign companies: {len(campaign_companies)}")

# Create company-to-campaign mapping
company_campaigns = {}
for _, row in attendees.iterrows():
    company = row['Company_Norm']
    campaign = row['Campaign Name'] if pd.notna(row.get('Campaign Name')) else ''
    if company not in company_campaigns:
        company_campaigns[company] = set()
    if campaign:
        company_campaigns[company].add(campaign)

# Match active pipeline to campaign attendees using fuzzy matching
def find_campaign_match(account_norm):
    """Find if this account matches any campaign attendee company."""
    if not account_norm:
        return set()
    
    # Exact match
    if account_norm in company_campaigns:
        return company_campaigns[account_norm]
    
    # Partial match - check if account contains or is contained in campaign company
    for camp_company in campaign_companies:
        if len(camp_company) >= 4 and len(account_norm) >= 4:
            if camp_company in account_norm or account_norm in camp_company:
                return company_campaigns.get(camp_company, set())
    
    return set()

active['Campaigns'] = active['Account_Norm'].apply(find_campaign_match)
active['From_Campaign'] = active['Campaigns'].apply(lambda x: len(x) > 0)

# Filter to campaign-related opportunities
campaign_pipeline = active[active['From_Campaign']].copy()
print(f"Campaign-related active opportunities: {len(campaign_pipeline)}")

if len(campaign_pipeline) == 0:
    print("\n⚠️  No direct matches found. Showing all active pipeline instead.")
    campaign_pipeline = active.copy()
    campaign_pipeline['Campaigns'] = campaign_pipeline['Account_Norm'].apply(lambda x: {'(All Active)'})

# Build summary by account
summary = []
for account_norm in campaign_pipeline['Account_Norm'].unique():
    acct_opps = campaign_pipeline[campaign_pipeline['Account_Norm'] == account_norm]
    
    # Get original account name from opportunity name
    opp_name = acct_opps['Opportunity Name'].iloc[0]
    account_name = re.split(r'\s*[-–]\s*', str(opp_name))[0] if pd.notna(opp_name) else account_norm
    
    campaigns = list(acct_opps['Campaigns'].iloc[0])
    total_acv = acct_opps['ACV'].sum() if 'ACV' in acct_opps.columns else 0
    opp_count = len(acct_opps)
    stages = acct_opps['Stage'].unique().tolist() if 'Stage' in acct_opps.columns else []
    
    summary.append({
        'Account': account_name,
        'Campaigns': ', '.join(str(c) for c in campaigns[:2]) + ('...' if len(campaigns) > 2 else ''),
        'Open Opps': opp_count,
        'Active ACV': total_acv,
        'Stages': ', '.join(str(s) for s in stages[:3]),
    })

summary_df = pd.DataFrame(summary)
if len(summary_df) > 0:
    summary_df = summary_df.sort_values('Active ACV', ascending=False)

print("\nTop 15 by Active ACV:")
print(summary_df.head(15).to_string(index=False))

# Calculate totals
total_acv = summary_df['Active ACV'].sum() if len(summary_df) > 0 else 0
total_opps = summary_df['Open Opps'].sum() if len(summary_df) > 0 else 0
total_accounts = len(summary_df)

print(f"\n📊 SUMMARY:")
print(f"   Campaign-related accounts in active pipeline: {total_accounts}")
print(f"   Total open opportunities: {total_opps}")
print(f"   Total active ACV: ${total_acv:,.0f}")

# Load the existing 2026 review workbook and add new sheet
review_path = os.path.expanduser("~/Desktop/MARKETING_2026_EOY_REVIEW.xlsx")
wb = load_workbook(review_path)

# Create new sheet for pipeline analysis
if "Active Pipeline Analysis" in wb.sheetnames:
    del wb["Active Pipeline Analysis"]

ws = wb.create_sheet("Active Pipeline Analysis")

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Title
ws['A1'] = "MARKETING CAMPAIGN → ACTIVE PIPELINE ANALYSIS"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

ws['A3'] = "Campaign Attendees Currently in Active Pipeline"
ws['A3'].font = Font(bold=True, size=11)

# Summary stats
ws['A4'] = f"Total Accounts: {total_accounts}"
ws['B4'] = f"Open Opportunities: {total_opps}"
ws['C4'] = f"Total Active ACV: ${total_acv:,.0f}"

# Headers
headers = ["Account", "Campaigns Attended", "Open Opps", "Active ACV", "Current Stages"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Data rows
for i, row in enumerate(summary_df.head(50).itertuples(index=False), start=7):
    ws.cell(row=i, column=1, value=row[0])  # Account
    ws.cell(row=i, column=2, value=row[1])  # Campaigns
    ws.cell(row=i, column=3, value=row[2])  # Open Opps
    ws.cell(row=i, column=4, value=row[3])  # Active ACV
    ws.cell(row=i, column=5, value=row[4])  # Stages

# Column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 40

# Save
wb.save(review_path)
print(f"\n✅ Added 'Active Pipeline Analysis' tab to: {review_path}")
