#!/usr/bin/env python3
"""
Marketing 2026 EOY Review Analysis
Cross-references campaign attendees with closed won deals and active pipeline.
Isolates specific clients (per user request) from main totals.
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the marketing data
excel_path = os.path.expanduser("~/Desktop/MARKETING DATA REFERENCE.xlsx")

# Read all sheets
attendees = pd.read_excel(excel_path, sheet_name="all campaign attendees")
closed_won = pd.read_excel(excel_path, sheet_name="closed won since sept.")
active = pd.read_excel(excel_path, sheet_name="active in pipeline")

print("="*80)
print("MARKETING ANALYSIS - 2026 EOY REVIEW")
print("="*80)

# Normalize company names
def normalize(name):
    if pd.isna(name):
        return ""
    return str(name).strip().lower()

attendees['Company_Norm'] = attendees['Company'].apply(normalize)
closed_won['Account_Norm'] = closed_won['Account Name'].apply(normalize)
active['Account_Norm'] = active['Account Name'].apply(normalize)

# Get campaign groups
print("\n📋 CAMPAIGN ATTENDEE BREAKDOWN:")
campaign_counts = attendees['Campaign Name'].value_counts()
for camp, count in campaign_counts.items():
    print(f"  • {camp}: {count} attendees")

# Create company-to-campaign mapping
company_campaigns = {}
for _, row in attendees.iterrows():
    company = row['Company_Norm']
    campaign = row['Campaign Name'] if pd.notna(row['Campaign Name']) else ''
    if company not in company_campaigns:
        company_campaigns[company] = set()
    company_campaigns[company].add(campaign)

# Categorize deals
def categorize_deal(row):
    deal_type = str(row.get('Deal Type', '')).strip()
    if 'Commitment' in deal_type:
        return 'COMMITMENT (LOI)'
    elif 'Recurring' in deal_type:
        return 'RECURRING REVENUE'
    elif 'Project' in deal_type:
        return 'PROJECT REVENUE'
    return 'OTHER'

closed_won['Category'] = closed_won.apply(categorize_deal, axis=1)

# Identify which events each closed won deal attended
def get_event_attribution(company_norm):
    campaigns = company_campaigns.get(company_norm, set())
    events = []
    for c in campaigns:
        c_lower = c.lower() if c else ''
        if 'supper' in c_lower or 'dinner' in c_lower:
            events.append('AI Supper Club')
        elif 'lighthouse' in c_lower:
            events.append('Lighthouse')
        elif 'summit' in c_lower or 'augmented' in c_lower:
            events.append('Summit')
        elif 'iqpc' in c_lower or 'compliance' in c_lower or 'counsel' in c_lower:
            events.append('IQPC')
    return list(set(events))

closed_won['Events'] = closed_won['Account_Norm'].apply(get_event_attribution)

# Isolated clients (from screenshot - to be excluded from main totals but shown separately)
isolated_clients = ['coherent', 'cargill', 'intuit', 'meta']

print("\n" + "="*80)
print("EVENT-BY-EVENT ANALYSIS (with isolated clients separated)")
print("="*80)

# Process each event
events_data = {
    'AI Supper Club': {'commitment': 0, 'recurring': 0, 'project': 0, 'deals': [], 'isolated': []},
    'Lighthouse': {'commitment': 0, 'recurring': 0, 'project': 0, 'deals': [], 'isolated': []},
    'Summit': {'commitment': 0, 'recurring': 0, 'project': 0, 'deals': [], 'isolated': []},
    'IQPC': {'commitment': 0, 'recurring': 0, 'project': 0, 'deals': [], 'isolated': []},
}

for _, row in closed_won.iterrows():
    events = row['Events']
    if not events:
        continue
    
    account = row['Account Name']
    account_norm = row['Account_Norm']
    category = row['Category']
    
    # Get amount - try multiple columns
    amount = 0
    for col in ['Commitment ACV', 'Revenue', 'ACV', 'Amount']:
        if col in row and pd.notna(row[col]) and row[col] > 0:
            amount = row[col]
            break
    
    deal_name = row.get('Opportunity Name', account)
    
    # Check if this is an isolated client
    is_isolated = any(iso in account_norm for iso in isolated_clients)
    
    for event in events:
        if event in events_data:
            deal_info = {
                'account': account,
                'deal': deal_name,
                'amount': amount,
                'category': category
            }
            
            if is_isolated:
                events_data[event]['isolated'].append(deal_info)
            else:
                events_data[event]['deals'].append(deal_info)
                if category == 'COMMITMENT (LOI)':
                    events_data[event]['commitment'] += amount
                elif category == 'RECURRING REVENUE':
                    events_data[event]['recurring'] += amount
                elif category == 'PROJECT REVENUE':
                    events_data[event]['project'] += amount

# Print detailed analysis for each event
for event_name, data in events_data.items():
    print(f"\n{'─'*80}")
    print(f"📊 {event_name.upper()}")
    print(f"{'─'*80}")
    
    total_revenue = data['recurring'] + data['project']
    total_all = data['commitment'] + total_revenue
    
    print(f"\n  INCLUDED IN TOTALS:")
    print(f"  ├─ Commitment (LOI):     ${data['commitment']:,.0f}")
    print(f"  ├─ Recurring Revenue:    ${data['recurring']:,.0f}")
    print(f"  ├─ Project Revenue:      ${data['project']:,.0f}")
    print(f"  ├─ Total Actual Revenue: ${total_revenue:,.0f}")
    print(f"  └─ Grand Total:          ${total_all:,.0f}")
    
    if data['deals']:
        print(f"\n  DEALS (included):")
        for d in data['deals']:
            print(f"    • {d['account']}: ${d['amount']:,.0f} ({d['category']})")
    
    if data['isolated']:
        isolated_total = sum(d['amount'] for d in data['isolated'])
        print(f"\n  ⚠️  ISOLATED (excluded from totals, for reference):")
        for d in data['isolated']:
            print(f"    • {d['account']}: ${d['amount']:,.0f} ({d['category']})")
        print(f"    └─ Isolated Subtotal: ${isolated_total:,.0f}")

# Active pipeline by event
print("\n" + "="*80)
print("ACTIVE PIPELINE BY EVENT (current state)")
print("="*80)

active['Events'] = active['Account_Norm'].apply(get_event_attribution)

pipeline_by_event = {}
for event_name in ['AI Supper Club', 'Lighthouse', 'Summit', 'IQPC']:
    event_pipeline = active[active['Events'].apply(lambda x: event_name in x)]
    if len(event_pipeline) > 0:
        total_acv = event_pipeline['ACV'].sum() if 'ACV' in event_pipeline.columns else 0
        print(f"\n{event_name}: {len(event_pipeline)} active opportunities, ${total_acv:,.0f} pipeline ACV")
        pipeline_by_event[event_name] = []
        for _, row in event_pipeline.head(5).iterrows():
            stage = row.get('Stage', 'Unknown')
            acv = row.get('ACV', 0)
            print(f"  • {row['Account Name']}: {stage} - ${acv:,.0f}")
            pipeline_by_event[event_name].append({
                'account': row['Account Name'],
                'stage': stage,
                'acv': acv
            })

# Create final Excel output
print("\n" + "="*80)
print("GENERATING EXCEL OUTPUT...")
print("="*80)

wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2D5B8C", end_color="2D5B8C", fill_type="solid")
isolated_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
money_font = Font(name='Consolas')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: Executive Summary with answers to specific questions
ws1 = wb.active
ws1.title = "Executive Summary"

# Build summary data
summary_data = []
summary_data.append(["MARKETING EOY 2026 REVIEW - EXECUTIVE SUMMARY", "", "", ""])
summary_data.append(["", "", "", ""])

# AI Supper Club
sc_data = events_data['AI Supper Club']
sc_revenue = sc_data['recurring'] + sc_data['project']
summary_data.append(["AI SUPPER CLUB SERIES", "", "", ""])
summary_data.append(["Commitment (LOI) Attributed:", f"${sc_data['commitment']:,.0f}", "", ""])
summary_data.append(["Recurring Revenue Closed:", f"${sc_data['recurring']:,.0f}", "", ""])
summary_data.append(["Project Revenue Closed:", f"${sc_data['project']:,.0f}", "", ""])
summary_data.append(["Total Actual Revenue:", f"${sc_revenue:,.0f}", "", ""])
summary_data.append(["Grand Total (incl. LOI):", f"${sc_data['commitment'] + sc_revenue:,.0f}", "", ""])
summary_data.append(["", "", "", ""])

# Lighthouse
lh_data = events_data['Lighthouse']
lh_revenue = lh_data['recurring'] + lh_data['project']
summary_data.append(["LIGHTHOUSE EVENT", "", "", ""])
summary_data.append(["Commitment (LOI) Attributed:", f"${lh_data['commitment']:,.0f}", "", ""])
summary_data.append(["Recurring Revenue Closed:", f"${lh_data['recurring']:,.0f}", "", ""])
summary_data.append(["Project Revenue Closed:", f"${lh_data['project']:,.0f}", "", ""])
summary_data.append(["Total Actual Revenue:", f"${lh_revenue:,.0f}", "", ""])
summary_data.append(["Grand Total (incl. LOI):", f"${lh_data['commitment'] + lh_revenue:,.0f}", "", ""])
summary_data.append(["", "", "", ""])

# Summit
su_data = events_data['Summit']
su_revenue = su_data['recurring'] + su_data['project']
summary_data.append(["AUGMENTED INTELLIGENCE SUMMIT", "", "", ""])
summary_data.append(["Commitment (LOI) Attributed:", f"${su_data['commitment']:,.0f}", "", ""])
summary_data.append(["Recurring Revenue Closed:", f"${su_data['recurring']:,.0f}", "", ""])
summary_data.append(["Project Revenue Closed:", f"${su_data['project']:,.0f}", "", ""])
summary_data.append(["Total Actual Revenue:", f"${su_revenue:,.0f}", "", ""])
summary_data.append(["Grand Total (incl. LOI):", f"${su_data['commitment'] + su_revenue:,.0f}", "", ""])
summary_data.append(["", "", "", ""])

# IQPC
iq_data = events_data['IQPC']
iq_revenue = iq_data['recurring'] + iq_data['project']
summary_data.append(["IQPC CORPORATE COUNSEL & COMPLIANCE", "", "", ""])
summary_data.append(["Commitment (LOI) Attributed:", f"${iq_data['commitment']:,.0f}", "", ""])
summary_data.append(["Recurring Revenue Closed:", f"${iq_data['recurring']:,.0f}", "", ""])
summary_data.append(["Project Revenue Closed:", f"${iq_data['project']:,.0f}", "", ""])
summary_data.append(["Total Actual Revenue:", f"${iq_revenue:,.0f}", "", ""])
summary_data.append(["Grand Total (incl. LOI):", f"${iq_data['commitment'] + iq_revenue:,.0f}", "", ""])

for row in summary_data:
    ws1.append(row)

# Format header rows
ws1['A1'].font = Font(bold=True, size=14)
for row_num in [3, 10, 17, 24]:
    ws1.cell(row=row_num, column=1).font = Font(bold=True, size=12)

# Sheet 2: Detailed Deal Breakdown
ws2 = wb.create_sheet("Deal Breakdown")
ws2.append(["Event", "Category", "Account", "Deal Name", "Amount", "Included in Total"])

for event_name, data in events_data.items():
    for d in data['deals']:
        ws2.append([event_name, d['category'], d['account'], d['deal'], d['amount'], "YES"])
    for d in data['isolated']:
        row = ws2.max_row + 1
        ws2.append([event_name, d['category'], d['account'], d['deal'], d['amount'], "NO - ISOLATED"])
        # Highlight isolated rows
        for col in range(1, 7):
            ws2.cell(row=row, column=col).fill = isolated_fill

# Format header
for col in range(1, 7):
    ws2.cell(row=1, column=col).font = header_font
    ws2.cell(row=1, column=col).fill = header_fill

# Sheet 3: Active Pipeline
ws3 = wb.create_sheet("Active Pipeline")
ws3.append(["Event", "Account", "Stage", "ACV"])

for event_name, deals in pipeline_by_event.items():
    for d in deals:
        ws3.append([event_name, d['account'], d['stage'], d['acv']])

for col in range(1, 5):
    ws3.cell(row=1, column=col).font = header_font
    ws3.cell(row=1, column=col).fill = header_fill

# Sheet 4: Isolated Clients Reference
ws4 = wb.create_sheet("Isolated Clients (Reference)")
ws4.append(["Account", "Event", "Category", "Amount", "Notes"])
ws4.append(["Coherent", "Summit", "RECURRING", 1150000, "Coherent - Contracting"])
ws4.append(["Cargill", "Summit", "RECURRING", 521000, "Cargill - Contracting"])
ws4.append(["Cargill", "Summit", "PROJECT", 150000, "Cargill - Cortex Pilot"])
ws4.append(["Intuit", "Lighthouse", "RECURRING", 410000, "Intuit - Contracting, Marketing, Sigma"])
ws4.append(["Meta", "Summit", "RECURRING", 350000, "Meta - 2026 Extension"])

for col in range(1, 6):
    ws4.cell(row=1, column=col).font = header_font
    ws4.cell(row=1, column=col).fill = header_fill
    
for row in range(2, 7):
    for col in range(1, 6):
        ws4.cell(row=row, column=col).fill = isolated_fill

# Save
output_path = os.path.expanduser("~/Desktop/MARKETING_2026_EOY_REVIEW.xlsx")
wb.save(output_path)
print(f"\n✅ Saved to: {output_path}")

# Print final answers for copy/paste
print("\n" + "="*80)
print("FINAL ANSWERS FOR EOY REVIEW")
print("="*80)

print("""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

AI SUPPER CLUB SERIES
├─ Revenue Contribution:     ${:,.0f} (Recurring: ${:,.0f} | Project: ${:,.0f})
├─ Direct Close (ARR):       ${:,.0f}
└─ Total incl. Commitment:   ${:,.0f}

LIGHTHOUSE EVENT  
├─ Revenue Contribution:     ${:,.0f} (Recurring: ${:,.0f} | Project: ${:,.0f})
├─ Direct Close (ARR):       ${:,.0f}
└─ Total incl. Commitment:   ${:,.0f}
   Note: Intuit & Pure Storage converted within 12 months of attending

AUGMENTED INTELLIGENCE SUMMIT
├─ Recurring Revenue:        ${:,.0f}
├─ Project Revenue:          ${:,.0f}
├─ Commitment (LOI):         ${:,.0f}
└─ Total:                    ${:,.0f}

IQPC CORPORATE COUNSEL & COMPLIANCE (Bayer sourced)
├─ Revenue Contribution:     ${:,.0f} (Recurring: ${:,.0f} | Project: ${:,.0f})
├─ Direct Close (ARR):       ${:,.0f}
└─ Total incl. Commitment:   ${:,.0f}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚠️  ISOLATED CLIENTS (excluded from above, for reference):
   • Coherent: $1,150,000 (Recurring - Contracting)
   • Cargill: $671,000 ($521K Recurring + $150K Project)
   • Intuit: $410,000 (Recurring - Contracting, Marketing, Sigma)
   • Meta: $350,000 (Recurring - 2026 Extension)
   └─ Isolated Total: $2,581,000

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""".format(
    sc_revenue, sc_data['recurring'], sc_data['project'],
    sc_data['recurring'],
    sc_data['commitment'] + sc_revenue,
    
    lh_revenue, lh_data['recurring'], lh_data['project'],
    lh_data['recurring'],
    lh_data['commitment'] + lh_revenue,
    
    su_data['recurring'],
    su_data['project'],
    su_data['commitment'],
    su_data['commitment'] + su_revenue,
    
    iq_revenue, iq_data['recurring'], iq_data['project'],
    iq_data['recurring'],
    iq_data['commitment'] + iq_revenue
))
