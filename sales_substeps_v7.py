"""
Sales Sub-Steps Analysis v7
===========================
THREE-TAB FLOW:
1. SalesReview (human-editable) → 2. Inputs (formulas pull from SalesReview) → 3. Summary

SalesReview tab:
- Simple, sales-friendly view
- Account Owner field
- Only key fields to verify
- Verified flag for sign-off

Changes in SalesReview flow through to Summary automatically.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("="*80)
print("SALES SUB-STEPS ANALYSIS v7")
print("Sales-Friendly Review Template")
print("="*80)

# =============================================================================
# LOAD DATA (same as before)
# =============================================================================
print("\n[STEP 1] Loading data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'

xl_audit = pd.ExcelFile(audit_file)
xl_secondary = pd.ExcelFile(secondary_file)

df_accts = pd.read_excel(xl_audit, sheet_name='all accts')
df_meetings_audit = pd.read_excel(xl_audit, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel(xl_secondary, sheet_name='all meetings')

df_meetings_raw = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings_raw['Date'] = pd.to_datetime(df_meetings_raw['Date'], errors='coerce')
df_meetings_raw['Date_Only'] = df_meetings_raw['Date'].dt.date

df_meetings_raw['_key'] = (
    df_meetings_raw['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings_raw['Date_Only'].astype(str) + '|' +
    df_meetings_raw['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings_raw.drop_duplicates(subset='_key', keep='first').copy()
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Meetings loaded: {len(df_meetings)}")

# =============================================================================
# PROCESS DATA
# =============================================================================
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

df_meetings['_norm_acct'] = df_meetings['Company / Account'].apply(normalize_name)
account_counts = df_meetings.groupby('_norm_acct').size().reset_index(name='count')
accounts_with_data = account_counts[account_counts['count'] >= 3]['_norm_acct'].tolist()

validated_info = {}
for _, row in df_accts.iterrows():
    norm = normalize_name(row['Account Name'])
    validated_info[norm] = {
        'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row['First Meeting Date']) else None,
        'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row['First Deal Closed']) else None,
    }

def classify_substep(subject, meeting_number):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_number == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product', 'eudia ai']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['use case', 'use-case', 'requirements']):
        return 'UseCase'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['proposal', 'delivery plan']):
        return 'Proposal'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    if any(kw in s for kw in ['pilot', 'poc', 'kickoff']):
        return 'Pilot'
    return 'Followup'

print("\n[STEP 2] Building account data...")

account_data = []
all_meetings = []

for norm_acct in accounts_with_data:
    acct_meetings = df_meetings[df_meetings['_norm_acct'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) < 3:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    first_meeting = acct_meetings['Date'].min()
    
    meeting_dates = {}
    for idx, row in acct_meetings.iterrows():
        meeting_num = idx + 1
        substep = classify_substep(row['Subject'], meeting_num)
        
        all_meetings.append({
            'Account': orig_name,
            'MeetingNum': meeting_num,
            'Date': row['Date'],
            'Subject': row['Subject'],
            'Classification': substep,
        })
        
        if substep not in meeting_dates:
            meeting_dates[substep] = []
        meeting_dates[substep].append(row['Date'])
    
    record = {
        'Account': orig_name,
        'TotalMeetings': len(acct_meetings),
        'FirstMeetingDate': first_meeting.strftime('%m/%d/%Y') if pd.notna(first_meeting) else '',
    }
    
    if norm_acct in validated_info and validated_info[norm_acct]['first_close']:
        close_date = validated_info[norm_acct]['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        record['SalesCycleDays'] = (close_date - first_meeting).days
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    all_substeps = ['Intro', 'Followup', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Proposal', 'Compliance', 'Contracting', 'Pilot']
    for substep in all_substeps:
        if substep in meeting_dates and len(meeting_dates[substep]) > 0:
            first_date = min(meeting_dates[substep])
            record[f'DaysTo{substep}'] = (first_date - first_meeting).days
            record[f'{substep}Count'] = len(meeting_dates[substep])
        else:
            record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = 0
    
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    account_data.append(record)

accounts_df = pd.DataFrame(account_data)
meetings_df = pd.DataFrame(all_meetings)

print(f"  Total accounts: {len(accounts_df)}")

# =============================================================================
# CREATE WORKBOOK
# =============================================================================
print("\n[STEP 3] Creating workbook with SalesReview flow...")

wb = Workbook()

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
editable_fill = PatternFill(start_color="FFFFD6", end_color="FFFFD6", fill_type="solid")  # Light yellow = editable

# =============================================================================
# TAB 1: SalesReview (HUMAN-EDITABLE SOURCE)
# =============================================================================
ws_review = wb.active
ws_review.title = 'SalesReview'

# SalesReview columns - simplified for sales reps
review_cols = [
    'Account',           # A - Display only
    'Owner',             # B - EDITABLE - sales rep name
    'Include',           # C - EDITABLE - Y/N
    'Verified',          # D - EDITABLE - Y/N (has rep verified this data?)
    'TotalMeetings',     # E - EDITABLE (can override)
    'FirstMeetingDate',  # F - EDITABLE (can correct)
    'CloseDate',         # G - EDITABLE (can correct)
    'SalesCycleDays',    # H - EDITABLE (can override)
    'DaysToDemo',        # I - EDITABLE (can correct)
    'DemoCount',         # J - EDITABLE (can correct)
    'DaysToCAB',         # K - EDITABLE (can correct)
    'CABCount',          # L - EDITABLE (can correct)
    'Notes',             # M - EDITABLE - free text for rep comments
]

# Headers
for col_idx, col_name in enumerate(review_cols, 1):
    cell = ws_review.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Data rows
for row_idx, (_, row) in enumerate(accounts_df.iterrows(), 2):
    ws_review.cell(row=row_idx, column=1, value=row.get('Account', ''))
    ws_review.cell(row=row_idx, column=2, value='')  # Owner - blank for rep to fill
    ws_review.cell(row=row_idx, column=3, value='Y')  # Include
    ws_review.cell(row=row_idx, column=4, value='')  # Verified - blank until confirmed
    ws_review.cell(row=row_idx, column=5, value=row.get('TotalMeetings', ''))
    ws_review.cell(row=row_idx, column=6, value=row.get('FirstMeetingDate', ''))
    ws_review.cell(row=row_idx, column=7, value=row.get('CloseDate', ''))
    ws_review.cell(row=row_idx, column=8, value=row.get('SalesCycleDays', ''))
    ws_review.cell(row=row_idx, column=9, value=row.get('DaysToDemo', ''))
    ws_review.cell(row=row_idx, column=10, value=row.get('DemoCount', ''))
    ws_review.cell(row=row_idx, column=11, value=row.get('DaysToCAB', ''))
    ws_review.cell(row=row_idx, column=12, value=row.get('CABCount', ''))
    ws_review.cell(row=row_idx, column=13, value='')  # Notes
    
    # Highlight editable columns (B, C, D, M)
    for col in [2, 3, 4, 13]:
        ws_review.cell(row=row_idx, column=col).fill = editable_fill

num_accounts = len(accounts_df)
data_end_row = num_accounts + 1

# =============================================================================
# TAB 2: Inputs (FORMULAS PULL FROM SalesReview)
# =============================================================================
ws_inputs = wb.create_sheet('Inputs')

# Inputs has more columns - some reference SalesReview, others are for detailed data
input_cols = [
    ('Account', 'A'),           # =SalesReview!A{row}
    ('Owner', 'B'),             # =SalesReview!B{row}
    ('Include', 'C'),           # =SalesReview!C{row}
    ('Verified', 'D'),          # =SalesReview!D{row}
    ('TotalMeetings', 'E'),     # =SalesReview!E{row}
    ('FirstMeetingDate', 'F'),  # =SalesReview!F{row}
    ('CloseDate', 'G'),         # =SalesReview!G{row}
    ('SalesCycleDays', 'H'),    # =SalesReview!H{row}
    ('DaysToDemo', 'I'),        # =SalesReview!I{row}
    ('DemoCount', 'J'),         # =SalesReview!J{row}
    ('DaysToCAB', 'K'),         # =SalesReview!K{row}
    ('CABCount', 'L'),          # =SalesReview!L{row}
]

# Additional columns from full data (not in SalesReview - too detailed)
extra_cols = ['DaysFirstToSecond', 'DaysToScoping', 'ScopingCount', 'DaysToProposal', 'ProposalCount',
              'DaysToCompliance', 'ComplianceCount', 'DaysToContracting', 'ContractingCount']

# Headers
for col_idx, (col_name, _) in enumerate(input_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

for col_idx, col_name in enumerate(extra_cols, len(input_cols) + 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Data rows - formulas reference SalesReview
for row_idx in range(2, data_end_row + 1):
    # Columns that pull from SalesReview
    for col_idx, (col_name, sr_col) in enumerate(input_cols, 1):
        ws_inputs.cell(row=row_idx, column=col_idx, value=f'=SalesReview!{sr_col}{row_idx}')
    
    # Extra columns - static values from original data
    data_row_idx = row_idx - 2
    if data_row_idx < len(accounts_df):
        row_data = accounts_df.iloc[data_row_idx]
        for col_idx, col_name in enumerate(extra_cols, len(input_cols) + 1):
            ws_inputs.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

# =============================================================================
# TAB 3: Summary (FORMULAS REFERENCE Inputs)
# =============================================================================
ws_summary = wb.create_sheet('Summary')

headers = ['Category', 'Metric', 'Average', 'SampleN', 'Total', 'Notes']
for col_idx, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=1, column=col_idx, value=h)
    cell.font = header_font

row_num = 3

# DATA OVERVIEW
ws_summary.cell(row=row_num, column=1, value='DATA OVERVIEW')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Total accounts (Include=Y)')
ws_summary.cell(row=row_num, column=3, value='=COUNTIF(Inputs!C:C,"Y")')
ws_summary.cell(row=row_num, column=6, value='From SalesReview Include column')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Verified accounts')
ws_summary.cell(row=row_num, column=3, value='=COUNTIF(Inputs!D:D,"Y")')
ws_summary.cell(row=row_num, column=6, value='Accounts where Verified=Y')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Accounts with close date')
ws_summary.cell(row=row_num, column=3, value='=COUNTIFS(Inputs!C:C,"Y",Inputs!H:H,">0")')
row_num += 2

# SALES CYCLE
ws_summary.cell(row=row_num, column=1, value='SALES CYCLE (days)')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='First Meeting to Close')
ws_summary.cell(row=row_num, column=3, value='=AVERAGEIF(Inputs!C:C,"Y",Inputs!H:H)')
ws_summary.cell(row=row_num, column=4, value='=COUNTIFS(Inputs!C:C,"Y",Inputs!H:H,">0")')
row_num += 2

# TIME TO MILESTONE
ws_summary.cell(row=row_num, column=1, value='TIME TO MILESTONE (days)')
row_num += 1

milestones = [
    ('First to Demo', 'I', 'Key qualification step'),
    ('First to CAB', 'K', 'Champion identification'),
]

for label, col_let, note in milestones:
    ws_summary.cell(row=row_num, column=2, value=label)
    ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
    ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!C:C,"Y",Inputs!{col_let}:{col_let},">=0")')
    ws_summary.cell(row=row_num, column=6, value=note)
    row_num += 1

row_num += 1

# MEETING COUNTS
ws_summary.cell(row=row_num, column=1, value='MEETING COUNTS')
row_num += 1

counts = [
    ('Demo meetings', 'J'),
    ('CAB discussions', 'L'),
]

for label, col_let in counts:
    ws_summary.cell(row=row_num, column=2, value=label)
    ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
    ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!C:C,"Y",Inputs!{col_let}:{col_let},">0")')
    ws_summary.cell(row=row_num, column=5, value=f'=SUMIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
    row_num += 1

row_num += 1

# HOW IT WORKS
ws_summary.cell(row=row_num, column=1, value='HOW THIS WORKS')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='1. Sales reps edit SalesReview tab (yellow cells)')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='2. Inputs tab pulls from SalesReview via formulas')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='3. Summary calculates from Inputs')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='4. Change SalesReview → Summary updates automatically')

# =============================================================================
# TAB 4: Meetings (for reference)
# =============================================================================
ws_meetings = wb.create_sheet('Meetings')

meeting_cols = ['Account', 'MeetingNum', 'Date', 'Subject', 'Classification']
for col_idx, col_name in enumerate(meeting_cols, 1):
    cell = ws_meetings.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

for row_idx, (_, row) in enumerate(meetings_df.iterrows(), 2):
    for col_idx, col_name in enumerate(meeting_cols, 1):
        val = row.get(col_name, '')
        if col_name == 'Date' and pd.notna(val):
            val = val.strftime('%m/%d/%Y')
        ws_meetings.cell(row=row_idx, column=col_idx, value=val)

# =============================================================================
# TAB 5: Instructions
# =============================================================================
ws_help = wb.create_sheet('Instructions')

instructions = [
    ['SALES REVIEW WORKFLOW'],
    [''],
    ['DATA FLOW:'],
    ['SalesReview (edit here) → Inputs (formulas) → Summary (auto-calculates)'],
    [''],
    ['FOR SALES REPS:'],
    ['1. Open SalesReview tab'],
    ['2. Find your accounts'],
    ['3. Fill in Owner column with your name'],
    ['4. Review the data for accuracy'],
    ['5. If correct, set Verified = Y'],
    ['6. If wrong, edit the value directly'],
    ['7. Add any notes in Notes column'],
    [''],
    ['EDITABLE COLUMNS (yellow):'],
    ['- Owner: Your name'],
    ['- Include: Y to include in analysis, N to exclude'],
    ['- Verified: Y once you confirm data is accurate'],
    ['- Notes: Free text for comments/corrections'],
    [''],
    ['TO CORRECT A VALUE:'],
    ['- Edit directly in SalesReview tab'],
    ['- Example: Wrong DaysToDemo? Just type the correct number'],
    ['- Inputs and Summary will update automatically'],
    [''],
    ['TO EXCLUDE AN ACCOUNT:'],
    ['- Set Include = N in SalesReview'],
    ['- Account will be excluded from all Summary calculations'],
    [''],
    ['FOR MANAGERS:'],
    ['- Summary tab shows aggregated metrics'],
    ['- Verified count shows how many accounts are confirmed'],
    ['- Filter SalesReview by Owner to see by rep'],
]

for row_idx, row_data in enumerate(instructions, 1):
    ws_help.cell(row=row_idx, column=1, value=row_data[0] if row_data else '')

# Set column widths
ws_review.column_dimensions['A'].width = 25
ws_review.column_dimensions['B'].width = 15
ws_review.column_dimensions['M'].width = 30

# Save
output_path = '/Users/keiganpesenti/Desktop/sales_substeps_v7.xlsx'
wb.save(output_path)

print(f"\n  Workbook saved: {output_path}")
print("\n  TABS:")
print("    SalesReview - EDIT HERE (yellow cells) - sales reps verify data")
print("    Inputs      - Formulas pull from SalesReview")
print("    Summary     - Auto-calculates from Inputs")
print("    Meetings    - Raw meeting data for reference")
print("    Instructions - How to use")
print("\n" + "="*80)
print("FLOW: SalesReview → Inputs → Summary")
print("Edit SalesReview, Summary updates automatically!")
print("="*80)

