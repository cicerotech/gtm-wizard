"""
Sales Sub-Steps Analysis v6
===========================
MAXIMUM COMPATIBILITY - Only uses:
- AVERAGEIF (Excel 2007+)
- COUNTIF/COUNTIFS (Excel 2007+)
- SUMIF (all versions)

NO MINIFS/MAXIFS - these only work in Excel 2019+
Min/Max calculated as static values in Python
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

print("="*80)
print("SALES SUB-STEPS ANALYSIS v6")
print("Maximum Excel Compatibility")
print("="*80)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[STEP 1] Loading data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'

xl_audit = pd.ExcelFile(audit_file)
xl_secondary = pd.ExcelFile(secondary_file)

df_accts = pd.read_excel(xl_audit, sheet_name='all accts')
df_meetings_audit = pd.read_excel(xl_audit, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel(xl_secondary, sheet_name='all meetings')

# Combine and clean meetings
df_meetings_raw = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings_raw['Date'] = pd.to_datetime(df_meetings_raw['Date'], errors='coerce')
df_meetings_raw['Date_Only'] = df_meetings_raw['Date'].dt.date

# Deduplicate
df_meetings_raw['_key'] = (
    df_meetings_raw['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings_raw['Date_Only'].astype(str) + '|' +
    df_meetings_raw['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings_raw.drop_duplicates(subset='_key', keep='first').copy()
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Exclude test accounts
test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Meetings loaded: {len(df_meetings)}")

# =============================================================================
# NORMALIZE ACCOUNT NAMES
# =============================================================================
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

df_meetings['_norm_acct'] = df_meetings['Company / Account'].apply(normalize_name)
account_counts = df_meetings.groupby('_norm_acct').size().reset_index(name='count')
accounts_with_data = account_counts[account_counts['count'] >= 3]['_norm_acct'].tolist()

# Get validated close info
validated_info = {}
for _, row in df_accts.iterrows():
    norm = normalize_name(row['Account Name'])
    validated_info[norm] = {
        'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row['First Meeting Date']) else None,
        'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row['First Deal Closed']) else None,
    }

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
def classify_substep(subject, meeting_number):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_number == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product', 'eudia ai']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['use case', 'use-case', 'requirements']):
        return 'UseCase'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['proposal', 'delivery plan']):
        return 'Proposal'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    if any(kw in s for kw in ['pilot', 'poc', 'kickoff']):
        return 'Pilot'
    return 'Followup'

# =============================================================================
# BUILD ACCOUNT-LEVEL DATA
# =============================================================================
print("\n[STEP 2] Building account-level metrics...")

account_data = []
all_meetings = []

for norm_acct in accounts_with_data:
    acct_meetings = df_meetings[df_meetings['_norm_acct'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) < 3:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    first_meeting = acct_meetings['Date'].min()
    
    # Classify meetings and track dates
    meeting_dates = {}
    for idx, row in acct_meetings.iterrows():
        meeting_num = idx + 1
        substep = classify_substep(row['Subject'], meeting_num)
        
        all_meetings.append({
            'Account': orig_name,
            'MeetingNum': meeting_num,
            'Date': row['Date'],
            'Subject': row['Subject'],
            'Classification': substep,
            'IncludeFlag': 'Y',
        })
        
        if substep not in meeting_dates:
            meeting_dates[substep] = []
        meeting_dates[substep].append(row['Date'])
    
    # Build record
    record = {
        'Account': orig_name,
        'Include': 'Y',
        'TotalMeetings': len(acct_meetings),
        'FirstMeetingDate': first_meeting.strftime('%m/%d/%Y') if pd.notna(first_meeting) else '',
    }
    
    # Close date
    if norm_acct in validated_info and validated_info[norm_acct]['first_close']:
        close_date = validated_info[norm_acct]['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        record['SalesCycleDays'] = (close_date - first_meeting).days
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    # Days to each substep
    all_substeps = ['Intro', 'Followup', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Proposal', 'Compliance', 'Contracting', 'Pilot']
    for substep in all_substeps:
        if substep in meeting_dates and len(meeting_dates[substep]) > 0:
            first_date = min(meeting_dates[substep])
            record[f'DaysTo{substep}'] = (first_date - first_meeting).days
            record[f'{substep}Count'] = len(meeting_dates[substep])
        else:
            record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = 0
    
    # First to second meeting
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    account_data.append(record)

accounts_df = pd.DataFrame(account_data)
meetings_df = pd.DataFrame(all_meetings)

print(f"  Total accounts: {len(accounts_df)}")
print(f"  Total meetings: {len(meetings_df)}")

# =============================================================================
# PRE-CALCULATE MIN/MAX VALUES (since MINIFS/MAXIFS not universally supported)
# =============================================================================
def calc_stats(df, col, include_col='Include'):
    """Calculate stats for accounts where Include='Y' and value is valid"""
    valid = df[(df[include_col] == 'Y') & (df[col].apply(lambda x: isinstance(x, (int, float)) and x >= 0))]
    values = valid[col].dropna()
    if len(values) == 0:
        return {'avg': '', 'count': 0, 'min': '', 'max': ''}
    return {
        'avg': round(values.mean(), 1),
        'count': len(values),
        'min': int(values.min()),
        'max': int(values.max())
    }

# =============================================================================
# CREATE EXCEL WORKBOOK
# =============================================================================
print("\n[STEP 3] Creating workbook...")

wb = Workbook()

# === TAB 1: Inputs ===
ws_inputs = wb.active
ws_inputs.title = 'Inputs'

input_cols = ['Account', 'Include', 'TotalMeetings', 'FirstMeetingDate', 'CloseDate', 'SalesCycleDays',
              'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 'DaysToCAB', 'CABCount',
              'DaysToUseCase', 'UseCaseCount', 'DaysToScoping', 'ScopingCount',
              'DaysToProposal', 'ProposalCount', 'DaysToCompliance', 'ComplianceCount',
              'DaysToContracting', 'ContractingCount', 'DaysToIntro', 'IntroCount',
              'DaysToFollowup', 'FollowupCount', 'DaysToPilot', 'PilotCount']

input_cols = [c for c in input_cols if c in accounts_df.columns]

for col_idx, col_name in enumerate(input_cols, 1):
    ws_inputs.cell(row=1, column=col_idx, value=col_name)

for row_idx, (_, row) in enumerate(accounts_df.iterrows(), 2):
    for col_idx, col_name in enumerate(input_cols, 1):
        val = row.get(col_name, '')
        ws_inputs.cell(row=row_idx, column=col_idx, value=val)

num_accounts = len(accounts_df)
data_end_row = num_accounts + 1

# === TAB 2: Summary ===
ws_summary = wb.create_sheet('Summary')

# Headers
headers = ['Category', 'Metric', 'Average', 'SampleN', 'Min', 'Max', 'Total', 'Notes']
for col_idx, h in enumerate(headers, 1):
    ws_summary.cell(row=1, column=col_idx, value=h)

row_num = 3

# DATA OVERVIEW
ws_summary.cell(row=row_num, column=1, value='DATA OVERVIEW')
row_num += 1

ws_summary.cell(row=row_num, column=2, value='Total accounts (Include=Y)')
ws_summary.cell(row=row_num, column=3, value='=COUNTIF(Inputs!B:B,"Y")')
ws_summary.cell(row=row_num, column=8, value='Change Include to N to exclude')
row_num += 1

ws_summary.cell(row=row_num, column=2, value='Accounts with close date')
ws_summary.cell(row=row_num, column=3, value='=COUNTIFS(Inputs!B:B,"Y",Inputs!F:F,">0")')
ws_summary.cell(row=row_num, column=8, value='Closed deals only')
row_num += 2

# SALES CYCLE
ws_summary.cell(row=row_num, column=1, value='SALES CYCLE (days)')
row_num += 1

stats = calc_stats(accounts_df, 'SalesCycleDays')
ws_summary.cell(row=row_num, column=2, value='First Meeting to Close')
ws_summary.cell(row=row_num, column=3, value='=AVERAGEIF(Inputs!B:B,"Y",Inputs!F:F)')
ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!B:B,"Y",Inputs!F:F,">0")')
ws_summary.cell(row=row_num, column=5, value=stats['min'])  # Static - calculated in Python
ws_summary.cell(row=row_num, column=6, value=stats['max'])  # Static - calculated in Python
ws_summary.cell(row=row_num, column=8, value='Min/Max are static values')
row_num += 2

# TIME TO MILESTONE
ws_summary.cell(row=row_num, column=1, value='TIME TO MILESTONE (days from first meeting)')
row_num += 1

milestones = [
    ('DaysFirstToSecond', 'G', 'First to Second Meeting', 'Engagement velocity'),
    ('DaysToDemo', 'H', 'First to Demo', 'Key qualification step'),
    ('DaysToCAB', 'J', 'First to CAB Discussion', 'Champion identification'),
    ('DaysToUseCase', 'L', 'First to Use Case', 'Requirements'),
    ('DaysToScoping', 'N', 'First to Scoping', 'Pricing discussions'),
    ('DaysToProposal', 'P', 'First to Proposal', 'Deal progression'),
    ('DaysToContracting', 'T', 'First to Contracting', 'Near close'),
    ('DaysToCompliance', 'R', 'First to Compliance', 'Security review'),
]

for col_name, col_let, label, note in milestones:
    if col_name in accounts_df.columns:
        stats = calc_stats(accounts_df, col_name)
        ws_summary.cell(row=row_num, column=2, value=label)
        ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!B:B,"Y",Inputs!{col_let}:{col_let})')
        ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">=0")')
        ws_summary.cell(row=row_num, column=5, value=stats['min'])
        ws_summary.cell(row=row_num, column=6, value=stats['max'])
        ws_summary.cell(row=row_num, column=8, value=note)
        row_num += 1

row_num += 1

# MEETING COUNTS
ws_summary.cell(row=row_num, column=1, value='MEETING COUNTS (per account)')
row_num += 1

counts = [
    ('DemoCount', 'I', 'Demo meetings'),
    ('CABCount', 'K', 'CAB discussions'),
    ('UseCaseCount', 'M', 'Use case meetings'),
    ('ScopingCount', 'O', 'Scoping meetings'),
    ('ProposalCount', 'Q', 'Proposal discussions'),
    ('ComplianceCount', 'S', 'Compliance meetings'),
    ('ContractingCount', 'U', 'Contracting meetings'),
]

for col_name, col_let, label in counts:
    if col_name in accounts_df.columns:
        ws_summary.cell(row=row_num, column=2, value=label)
        ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!B:B,"Y",Inputs!{col_let}:{col_let})')
        ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">0")')
        ws_summary.cell(row=row_num, column=7, value=f'=SUMIF(Inputs!B:B,"Y",Inputs!{col_let}:{col_let})')
        row_num += 1

row_num += 1

# NOTES
ws_summary.cell(row=row_num, column=1, value='NOTES')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Average, SampleN, Total = Excel formulas (auto-update)')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Min, Max = Static values (update manually if needed)')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='To exclude account: change Include from Y to N in Inputs tab')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='To override value: edit the cell directly in Inputs tab')

# === TAB 3: Meetings ===
ws_meetings = wb.create_sheet('Meetings')

meeting_cols = ['Account', 'MeetingNum', 'Date', 'Subject', 'Classification', 'IncludeFlag']
for col_idx, col_name in enumerate(meeting_cols, 1):
    ws_meetings.cell(row=1, column=col_idx, value=col_name)

for row_idx, (_, row) in enumerate(meetings_df.iterrows(), 2):
    for col_idx, col_name in enumerate(meeting_cols, 1):
        val = row.get(col_name, '')
        if col_name == 'Date' and pd.notna(val):
            val = val.strftime('%m/%d/%Y')
        ws_meetings.cell(row=row_idx, column=col_idx, value=val)

# === TAB 4: HowToUse ===
ws_help = wb.create_sheet('HowToUse')
help_text = [
    'HOW TO USE THIS WORKBOOK',
    '',
    'FORMULAS USED (Excel 2007+ compatible):',
    '- AVERAGEIF',
    '- COUNTIF, COUNTIFS', 
    '- SUMIF',
    '',
    'Min/Max are STATIC values (not formulas) because',
    'MINIFS/MAXIFS only work in Excel 2019+',
    '',
    'TO EXCLUDE AN ACCOUNT:',
    '1. Go to Inputs tab',
    '2. Find the account row',
    '3. Change column B (Include) from Y to N',
    '4. Summary Average/Count/Total will update',
    '5. Min/Max may need manual update if outlier removed',
    '',
    'TO OVERRIDE A VALUE:',
    '1. Go to Inputs tab',
    '2. Find the account and column',
    '3. Change the value',
    '4. Summary formulas recalculate',
]

for row_idx, text in enumerate(help_text, 1):
    ws_help.cell(row=row_idx, column=1, value=text)

# Save
output_path = '/Users/keiganpesenti/Desktop/sales_substeps_v6.xlsx'
wb.save(output_path)

print(f"\n  Workbook saved: {output_path}")
print("\n  FORMULAS: AVERAGEIF, COUNTIF, COUNTIFS, SUMIF only")
print("  MIN/MAX: Static values (no MINIFS/MAXIFS - not compatible with all Excel)")
print("\n  This should work in Excel 2007 and later with NO #NAME? errors")

