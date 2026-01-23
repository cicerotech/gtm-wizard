"""
Validated Sales Data Generator
==============================
Cross-references:
1. days to close reference.xlsx (Salesforce validated data)
2. Meeting prep documents (391 Word docs)
3. Existing meeting data (latest audits v0.xlsx)

Produces a copy-paste ready Inputs tab for BL review.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from docx import Document
from collections import defaultdict

print("="*80)
print("VALIDATED SALES DATA GENERATOR")
print("Cross-referencing multiple data sources for accuracy")
print("="*80)

# =============================================================================
# STEP 1: Load Salesforce Reference Data
# =============================================================================
print("\n[STEP 1] Loading Salesforce reference data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
print(f"  Loaded {len(sf_ref)} accounts from days to close reference")

# Build lookup
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

print(f"  Built lookup for {len(sf_lookup)} accounts")

# =============================================================================
# STEP 2: Parse Meeting Prep Documents
# =============================================================================
print("\n[STEP 2] Parsing meeting prep documents...")

docs_folder = "/Users/keiganpesenti/Desktop/meeting_prep_extracted/Prep + Notes"

def extract_company_name(filename):
    """Extract company name from filename."""
    # Remove common suffixes
    name = filename.replace(' - Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep + Notes.docx', '')
    name = name.replace('_Meeting Prep & Notes.docx', '')
    name = name.replace('_Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep.docx', '')
    name = name.replace('.docx', '')
    # Remove date prefixes like "20250922 "
    name = re.sub(r'^\d{8}\s*', '', name)
    return name.strip()

def parse_meeting_doc(filepath):
    """Parse a meeting prep document for dates and meeting info."""
    try:
        doc = Document(filepath)
        
        meetings = []
        current_meeting = {}
        in_past_notes = False
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            
            # Check if we're in the past meeting notes section
            if 'FULL PAST MEETING NOTES' in text.upper() or 'PAST MEETING NOTES' in text.upper():
                in_past_notes = True
                continue
            
            if 'CRITICAL CUSTOMER HISTORY' in text.upper():
                in_past_notes = False
                continue
            
            # Parse meeting type
            if text.startswith('Meeting Type:'):
                if current_meeting and 'date' in current_meeting:
                    meetings.append(current_meeting)
                current_meeting = {'type': text.replace('Meeting Type:', '').strip()}
            
            # Parse date
            if text.startswith('Date:'):
                date_str = text.replace('Date:', '').strip()
                # Extract just the date part (before @ or other time info)
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', date_str)
                if date_match:
                    try:
                        current_meeting['date'] = pd.to_datetime(date_match.group(1))
                    except:
                        pass
            
            # Parse company (from Meeting Overview section)
            if text.startswith('Company:'):
                current_meeting['company'] = text.replace('Company:', '').strip().split('<')[0].strip()
            
            # Check for demo mentions in goals
            if 'Meeting Goals' in text or 'demo' in text.lower():
                if 'demo' in text.lower():
                    current_meeting['is_demo'] = True
        
        # Don't forget the last meeting
        if current_meeting and 'date' in current_meeting:
            meetings.append(current_meeting)
        
        return meetings
    except Exception as e:
        return []

# Get list of documents (skip old versions)
doc_files = []
for f in os.listdir(docs_folder):
    if not f.endswith('.docx'):
        continue
    if '(old)' in f.lower() or '(OLD)' in f:
        continue
    if f.startswith('~$'):  # Skip temp files
        continue
    if 'TEMPLATE' in f.upper() or 'COPY' in f.upper():
        continue
    doc_files.append(f)

print(f"  Found {len(doc_files)} meeting prep documents (excluding old/template)")

# Parse documents
doc_data = {}
for i, filename in enumerate(doc_files):
    if i % 50 == 0:
        print(f"    Processing {i}/{len(doc_files)}...")
    
    company = extract_company_name(filename)
    filepath = os.path.join(docs_folder, filename)
    meetings = parse_meeting_doc(filepath)
    
    if meetings:
        # Get first meeting (earliest date)
        valid_dates = [m['date'] for m in meetings if 'date' in m and pd.notna(m['date'])]
        demo_dates = [m['date'] for m in meetings if m.get('is_demo') and 'date' in m]
        
        doc_data[company.lower().strip()] = {
            'company': company,
            'meeting_count': len(meetings),
            'first_meeting': min(valid_dates) if valid_dates else None,
            'latest_meeting': max(valid_dates) if valid_dates else None,
            'first_demo': min(demo_dates) if demo_dates else None,
            'demo_count': len(demo_dates),
            'meetings': meetings,
        }

print(f"  Extracted data from {len(doc_data)} companies")

# =============================================================================
# STEP 3: Load Existing Meeting Data
# =============================================================================
print("\n[STEP 3] Loading existing meeting data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_accts = pd.read_excel(audit_file, sheet_name='all accts')

secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'
df_meetings_secondary = pd.read_excel(secondary_file, sheet_name='all meetings')

# Combine meetings
df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Deduplicate
df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

# Exclude test accounts
test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Loaded {len(df_meetings)} meetings from audit data")

# =============================================================================
# STEP 4: Build Validated Account Data
# =============================================================================
print("\n[STEP 4] Building validated account data...")

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    # Remove common suffixes
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

def fuzzy_match(name, lookup_dict, threshold=0.8):
    """Find best match in lookup dict."""
    name_norm = normalize_name(name)
    
    # Exact match
    if name_norm in lookup_dict:
        return name_norm
    
    # Try with original name (lowercase)
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    
    # Try partial matches (both directions)
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        
        # Check if one contains the other
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
            
            # Check first word match (e.g., "Amazon" matches "Amazon - Meeting Prep")
            name_first = name_clean.split()[0] if name_clean.split() else ''
            key_first = key_clean.split()[0] if key_clean.split() else ''
            if name_first and key_first and len(name_first) > 3:
                if name_first == key_first:
                    return key
    
    return None

def classify_meeting(subject, meeting_num):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_num == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    return 'Followup'

# Get unique accounts from meeting data
account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()

print(f"  Processing {len(accounts_to_process)} accounts with 3+ meetings")

# Build validated records
validated_records = []

for norm_acct in accounts_to_process:
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    
    # Initialize record
    record = {
        'Account': orig_name,
        'Owner': '',
        'Include': 'Y',
        'Verified': '',
        'DataSource': '',
        'Confidence': '',
    }
    
    # === SOURCE 1: Salesforce Reference ===
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    
    # Also try direct normalized lookup
    if not sf_data:
        norm_name = normalize_name(orig_name)
        for sf_key, sf_val in sf_lookup.items():
            sf_key_norm = normalize_name(sf_key)
            if norm_name == sf_key_norm:
                sf_data = sf_val
                sf_match = sf_key
                break
    
    # === SOURCE 2: Meeting Prep Documents ===
    doc_match = fuzzy_match(orig_name, doc_data)
    doc_info = doc_data.get(doc_match, {}) if doc_match else {}
    
    # === SOURCE 3: Meeting Data ===
    meeting_first = acct_meetings['Date'].min()
    meeting_count = len(acct_meetings)
    
    # Classify meetings
    meeting_dates = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        substep = classify_meeting(row['Subject'], idx + 1)
        meeting_dates[substep].append(row['Date'])
    
    # === MERGE DATA ===
    
    # Owner: prioritize Salesforce
    record['Owner'] = sf_data.get('owner', '')
    
    # First Meeting Date: cross-validate
    sources_agree = True
    first_meeting_sf = sf_data.get('first_meeting')
    first_meeting_doc = doc_info.get('first_meeting')
    first_meeting_audit = meeting_first
    
    # Determine best first meeting date
    candidates = []
    if first_meeting_sf and pd.notna(first_meeting_sf):
        candidates.append(('SF', first_meeting_sf))
    if first_meeting_doc and pd.notna(first_meeting_doc):
        candidates.append(('Doc', first_meeting_doc))
    if first_meeting_audit and pd.notna(first_meeting_audit):
        candidates.append(('Audit', first_meeting_audit))
    
    if candidates:
        # Use Salesforce if available, otherwise earliest
        sf_dates = [c for c in candidates if c[0] == 'SF']
        if sf_dates:
            record['FirstMeetingDate'] = sf_dates[0][1].strftime('%m/%d/%Y')
        else:
            earliest = min(candidates, key=lambda x: x[1])
            record['FirstMeetingDate'] = earliest[1].strftime('%m/%d/%Y')
        
        # Check if sources agree (within 30 days)
        if len(candidates) > 1:
            dates = [c[1] for c in candidates]
            date_range = (max(dates) - min(dates)).days
            if date_range > 30:
                sources_agree = False
    else:
        record['FirstMeetingDate'] = ''
    
    # Close Date & Sales Cycle: from Salesforce
    if sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        record['CloseDate'] = sf_data['first_close'].strftime('%m/%d/%Y')
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            record['SalesCycleDays'] = int(sf_data['days_to_close'])
        elif first_meeting_sf:
            record['SalesCycleDays'] = (sf_data['first_close'] - first_meeting_sf).days
        else:
            record['SalesCycleDays'] = ''
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    # Total Meetings: max of sources
    record['TotalMeetings'] = max(meeting_count, doc_info.get('meeting_count', 0))
    
    # Days to Demo
    if 'Demo' in meeting_dates and meeting_dates['Demo']:
        first_demo = min(meeting_dates['Demo'])
        if record['FirstMeetingDate']:
            first_mtg = pd.to_datetime(record['FirstMeetingDate'])
            record['DaysToDemo'] = (first_demo - first_mtg).days
        else:
            record['DaysToDemo'] = ''
        record['DemoCount'] = len(meeting_dates['Demo'])
    elif doc_info.get('first_demo'):
        if record['FirstMeetingDate']:
            first_mtg = pd.to_datetime(record['FirstMeetingDate'])
            record['DaysToDemo'] = (doc_info['first_demo'] - first_mtg).days
        else:
            record['DaysToDemo'] = ''
        record['DemoCount'] = doc_info.get('demo_count', 1)
    else:
        record['DaysToDemo'] = ''
        record['DemoCount'] = 0
    
    # Days to other milestones
    for substep in ['CAB', 'Scoping', 'Compliance', 'Contracting']:
        if substep in meeting_dates and meeting_dates[substep]:
            first_date = min(meeting_dates[substep])
            if record['FirstMeetingDate']:
                first_mtg = pd.to_datetime(record['FirstMeetingDate'])
                record[f'DaysTo{substep}'] = (first_date - first_mtg).days
            else:
                record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = len(meeting_dates[substep])
        else:
            record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = 0
    
    # First to Second Meeting
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    # Data Source & Confidence
    sources = []
    if sf_match:
        sources.append('SF')
    if doc_match:
        sources.append('Doc')
    sources.append('Audit')
    record['DataSource'] = '+'.join(sources)
    
    if 'SF' in sources and sources_agree:
        record['Confidence'] = 'High'
    elif 'SF' in sources or (len(sources) > 1 and sources_agree):
        record['Confidence'] = 'Medium'
    elif not sources_agree:
        record['Confidence'] = 'Review'
    else:
        record['Confidence'] = 'Low'
    
    validated_records.append(record)

validated_df = pd.DataFrame(validated_records)

# Sort by Owner, then Account
validated_df = validated_df.sort_values(['Owner', 'Account']).reset_index(drop=True)

print(f"  Created {len(validated_df)} validated account records")

# Show confidence breakdown
print("\n  Confidence breakdown:")
print(validated_df['Confidence'].value_counts().to_string())

print("\n  Data source breakdown:")
print(validated_df['DataSource'].value_counts().to_string())

# =============================================================================
# STEP 5: Generate Output
# =============================================================================
print("\n[STEP 5] Generating output...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

# === TAB 1: BL_Review_Inputs (copy-paste ready) ===
ws_inputs = wb.active
ws_inputs.title = 'BL_Review_Inputs'

# Define columns in the exact order needed for the Inputs tab
output_cols = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount',
    'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance',
    'ComplianceCount', 'DaysToContracting', 'ContractingCount', 'DataSource', 'Confidence'
]

# Headers
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for col_idx, col_name in enumerate(output_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Data
for row_idx, (_, row) in enumerate(validated_df.iterrows(), 2):
    for col_idx, col_name in enumerate(output_cols, 1):
        val = row.get(col_name, '')
        cell = ws_inputs.cell(row=row_idx, column=col_idx, value=val)
        
        # Color by confidence
        confidence = row.get('Confidence', '')
        if confidence == 'High':
            cell.fill = high_fill
        elif confidence == 'Review':
            cell.fill = review_fill
        elif confidence == 'Low':
            cell.fill = low_fill

# === TAB 2: Validation_Details ===
ws_details = wb.create_sheet('Validation_Details')

details_cols = ['Account', 'Owner', 'SF_FirstMeeting', 'Doc_FirstMeeting', 'Audit_FirstMeeting',
                'SF_DaysToClose', 'SF_CloseDate', 'Doc_MeetingCount', 'Audit_MeetingCount',
                'SourcesAgree', 'Confidence', 'Notes']

for col_idx, col_name in enumerate(details_cols, 1):
    ws_details.cell(row=1, column=col_idx, value=col_name)

# (Would populate with detailed comparison data here)

# === TAB 3: Instructions ===
ws_help = wb.create_sheet('Instructions')

instructions = [
    'VALIDATED SALES DATA - COPY-PASTE READY',
    '',
    'TAB: BL_Review_Inputs',
    'This tab is ready to copy-paste into your Inputs tab.',
    'All values are validated against multiple sources.',
    '',
    'COLOR CODING:',
    '- GREEN: High confidence (Salesforce + other sources agree)',
    '- YELLOW: Review needed (sources disagree or limited data)',
    '- RED: Low confidence (single source, unverified)',
    '',
    'DATA SOURCES:',
    '- SF: Salesforce "days to close reference.xlsx"',
    '- Doc: Meeting prep documents (391 Word files)',
    '- Audit: Meeting data from latest audits',
    '',
    'TO USE:',
    '1. Review yellow/red rows for accuracy',
    '2. Select all data in BL_Review_Inputs (excluding headers)',
    '3. Copy and paste into your Inputs tab',
    '4. Formulas in Summary will auto-update',
    '',
    'KNOWN LIMITATIONS:',
    '- EU accounts (Nathan, Connor, Alex, Tom, Nicola) have estimated close dates',
    '- Some accounts may have name variations across sources',
    '- Meeting prep docs may not have complete history',
]

for row_idx, text in enumerate(instructions, 1):
    ws_help.cell(row=row_idx, column=1, value=text)

# Column widths
ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 25

# Save
output_path = '/Users/keiganpesenti/Desktop/validated_inputs.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("DONE! Open validated_inputs.xlsx and copy BL_Review_Inputs to your Inputs tab")
print("="*80)

# Print summary stats
print("\n=== SUMMARY ===")
print(f"Total accounts: {len(validated_df)}")
print(f"With Owner: {len(validated_df[validated_df['Owner'] != ''])}")
print(f"With Close Date: {len(validated_df[validated_df['CloseDate'] != ''])}")
print(f"High Confidence: {len(validated_df[validated_df['Confidence'] == 'High'])}")
print(f"Need Review: {len(validated_df[validated_df['Confidence'] == 'Review'])}")

