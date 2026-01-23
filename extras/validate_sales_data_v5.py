"""
Validated Sales Data Generator v5
=================================
- Source_Meetings tab showing raw classified meetings
- Every metric traceable to source data
- Additional account exclusions
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from docx import Document
from collections import defaultdict

print("="*80)
print("VALIDATED SALES DATA GENERATOR v5")
print("With Source_Meetings tab for full traceability")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    # New exclusions
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# STEP 1: Load Reference Data
# =============================================================================
print("\n[STEP 1] Loading reference data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }
print(f"  SF reference: {len(sf_lookup)} accounts")

# =============================================================================
# STEP 2: Load Meeting Audit Data
# =============================================================================
print("\n[STEP 2] Loading meeting audit data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# STEP 3: Classify All Meetings (Source Data)
# =============================================================================
print("\n[STEP 3] Classifying all meetings for source data...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

# Build source meetings table
source_meetings = []

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

for norm_acct in accounts_to_process:
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    first_date = acct_meetings['Date'].min()
    
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_date)
        classification = classify_meeting(row['Subject'], is_first)
        
        source_meetings.append({
            'Account': orig_name,
            'Date': row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',
            'Subject': str(row['Subject'])[:80] if pd.notna(row['Subject']) else '',
            'Classification': classification,
            'IsFirstMeeting': 'Y' if is_first else '',
            'DaysFromFirst': (row['Date'] - first_date).days if pd.notna(row['Date']) else '',
        })

source_df = pd.DataFrame(source_meetings)
print(f"  Created source data for {len(source_df)} meetings across {len(accounts_to_process)} accounts")

# =============================================================================
# STEP 4: Build Summary Data
# =============================================================================
print("\n[STEP 4] Building summary data...")

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
            name_first = name_clean.split()[0] if name_clean.split() else ''
            key_first = key_clean.split()[0] if key_clean.split() else ''
            if name_first and key_first and len(name_first) > 3 and name_first == key_first:
                return key
    return None

validated_records = []

for norm_acct in accounts_to_process:
    acct_source = source_df[source_df['Account'].str.lower().str.strip() == norm_acct]
    if len(acct_source) == 0:
        continue
    
    orig_name = acct_source['Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    # Match to SF
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    record = {
        'Account': orig_name,
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Include': 'Y',
        'Verified': '',
    }
    
    # First Meeting Date
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    first_meeting_audit = acct_meetings['Date'].min() if len(acct_meetings) > 0 else None
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        record['FirstMeetingDate'] = first_meeting_sf.strftime('%m/%d/%Y')
    elif first_meeting_audit and pd.notna(first_meeting_audit):
        record['FirstMeetingDate'] = first_meeting_audit.strftime('%m/%d/%Y')
    else:
        record['FirstMeetingDate'] = ''
    
    first_mtg = pd.to_datetime(record['FirstMeetingDate']) if record['FirstMeetingDate'] else None
    
    # Total Meetings
    record['TotalMeetings'] = len(acct_source)
    
    # Days First to Second
    acct_meetings_sorted = acct_meetings.sort_values('Date')
    if len(acct_meetings_sorted) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings_sorted.iloc[1]['Date'] - acct_meetings_sorted.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    # Get classified meetings from source
    demos = acct_source[acct_source['Classification'] == 'Demo']
    cabs = acct_source[acct_source['Classification'] == 'CAB']
    scopings = acct_source[acct_source['Classification'] == 'Scoping']
    compliances = acct_source[acct_source['Classification'] == 'Compliance']
    contractings = acct_source[acct_source['Classification'] == 'Contracting']
    
    # Demo metrics
    if len(demos) > 0 and first_mtg:
        first_demo_days = demos['DaysFromFirst'].min()
        record['DaysToDemo'] = first_demo_days if pd.notna(first_demo_days) else ''
        record['DemoCount'] = len(demos)
    else:
        record['DaysToDemo'] = ''
        record['DemoCount'] = 0
    
    # CAB metrics
    if len(cabs) > 0 and first_mtg:
        first_cab_days = cabs['DaysFromFirst'].min()
        record['DaysToCAB'] = first_cab_days if pd.notna(first_cab_days) else ''
        record['CABCount'] = len(cabs)
    else:
        record['DaysToCAB'] = ''
        record['CABCount'] = 0
    
    # Scoping metrics
    if len(scopings) > 0 and first_mtg:
        first_scoping_days = scopings['DaysFromFirst'].min()
        record['DaysToScoping'] = first_scoping_days if pd.notna(first_scoping_days) else ''
        record['ScopingCount'] = len(scopings)
    else:
        record['DaysToScoping'] = ''
        record['ScopingCount'] = 0
    
    # Compliance metrics
    if len(compliances) > 0 and first_mtg:
        first_compliance_days = compliances['DaysFromFirst'].min()
        record['DaysToCompliance'] = first_compliance_days if pd.notna(first_compliance_days) else ''
        record['ComplianceCount'] = len(compliances)
    else:
        record['DaysToCompliance'] = ''
        record['ComplianceCount'] = 0
    
    # Contracting metrics
    if len(contractings) > 0 and first_mtg:
        first_contracting_days = contractings['DaysFromFirst'].min()
        record['DaysToContracting'] = first_contracting_days if pd.notna(first_contracting_days) else ''
        record['ContractingCount'] = len(contractings)
    else:
        record['DaysToContracting'] = ''
        record['ContractingCount'] = 0
    
    # Data Source
    record['DataSource'] = 'SF+Audit' if sf_data else 'Audit'
    
    # Confidence
    if has_no_close(orig_name):
        record['Confidence'] = 'Medium' if sf_data else 'Standard'
    elif sf_data and sf_data.get('first_close'):
        record['Confidence'] = 'High'
    elif sf_data:
        record['Confidence'] = 'Medium'
    else:
        record['Confidence'] = 'Standard'
    
    # Close Date
    if has_no_close(orig_name):
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            record['SalesCycleDays'] = int(sf_data['days_to_close'])
        elif first_mtg and close_date > first_mtg:
            record['SalesCycleDays'] = (close_date - first_mtg).days
        else:
            record['SalesCycleDays'] = ''
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    validated_records.append(record)

validated_df = pd.DataFrame(validated_records)

# Sort by Owner, then Account
validated_df['_has_owner'] = validated_df['Owner'].apply(lambda x: 0 if pd.notna(x) and x != '' else 1)
validated_df = validated_df.sort_values(['_has_owner', 'Owner', 'Account']).reset_index(drop=True)
validated_df = validated_df.drop(columns=['_has_owner'])

print(f"  Created {len(validated_df)} validated records")

# =============================================================================
# STEP 5: Generate Output with Source Tab
# =============================================================================
print("\n[STEP 5] Generating output with source data tab...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

# === TAB 1: BL_Review_Inputs ===
ws_inputs = wb.active
ws_inputs.title = 'BL_Review_Inputs'

input_cols = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 'DaysToCAB', 'CABCount',
    'DaysToScoping', 'ScopingCount', 'DaysToCompliance', 'ComplianceCount',
    'DaysToContracting', 'ContractingCount', 'DataSource', 'Confidence',
    'CloseDate', 'SalesCycleDays'
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for col_idx, col_name in enumerate(input_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, (_, row) in enumerate(validated_df.iterrows(), 2):
    for col_idx, col_name in enumerate(input_cols, 1):
        val = row.get(col_name, '')
        cell = ws_inputs.cell(row=row_idx, column=col_idx, value=val)
        
        confidence = row.get('Confidence', '')
        if confidence == 'High':
            cell.fill = high_fill
        elif confidence == 'Medium':
            cell.fill = medium_fill

ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 25

# === TAB 2: Source_Meetings (Raw classified meetings) ===
ws_source = wb.create_sheet('Source_Meetings')

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'IsFirstMeeting', 'DaysFromFirst']

for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Color-code by classification
class_fills = {
    'Demo': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    'CAB': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'Scoping': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    'Compliance': PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
    'Contracting': PatternFill(start_color="D5A6BD", end_color="D5A6BD", fill_type="solid"),
}

for row_idx, (_, row) in enumerate(source_df.iterrows(), 2):
    for col_idx, col_name in enumerate(source_cols, 1):
        val = row.get(col_name, '')
        cell = ws_source.cell(row=row_idx, column=col_idx, value=val)
        
        classification = row.get('Classification', '')
        if classification in class_fills:
            cell.fill = class_fills[classification]

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 60
ws_source.column_dimensions['D'].width = 15

# === TAB 3: Classification_Legend ===
ws_legend = wb.create_sheet('Classification_Legend')

legend_data = [
    ['Classification', 'Keywords Used', 'Color'],
    ['Intro', 'First meeting OR "intro", "introduction" in subject', 'White'],
    ['Demo', '"demo", "sigma", "cortex", "platform", "walkthrough", "product overview"', 'Blue'],
    ['CAB', '"cab", "customer advisory", "advisory board"', 'Green'],
    ['Scoping', '"scoping", "scope", "pricing", "proposal"', 'Yellow'],
    ['Compliance', '"infosec", "security", "compliance"', 'Orange'],
    ['Contracting', '"contract", "redline", "msa", "negotiation", "legal"', 'Purple'],
    ['Followup', 'Default - no keywords matched', 'White'],
    ['', '', ''],
    ['HOW TO USE THIS FILE:', '', ''],
    ['1. BL_Review_Inputs tab contains the summary metrics for each account', '', ''],
    ['2. Source_Meetings tab shows EVERY meeting used to calculate those metrics', '', ''],
    ['3. To verify a metric, filter Source_Meetings by Account + Classification', '', ''],
    ['', '', ''],
    ['EXAMPLE: To verify Demo metrics for "Amazon":', '', ''],
    ['  - Filter Source_Meetings: Account = "Amazon", Classification = "Demo"', '', ''],
    ['  - DemoCount = number of rows', '', ''],
    ['  - DaysToDemo = smallest DaysFromFirst value', '', ''],
]

for row_idx, row_data in enumerate(legend_data, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_legend.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

ws_legend.column_dimensions['A'].width = 20
ws_legend.column_dimensions['B'].width = 70

# Save
output_path = '/Users/keiganpesenti/Desktop/validated_inputs_v5.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total accounts: {len(validated_df)}")
print(f"Total source meetings: {len(source_df)}")
print(f"With Owner: {len(validated_df[validated_df['Owner'].notna() & (validated_df['Owner'] != '')])}")
print(f"With Close Date: {len(validated_df[validated_df['CloseDate'].notna() & (validated_df['CloseDate'] != '')])}")

# Classification breakdown
print("\n=== MEETING CLASSIFICATIONS ===")
class_counts = source_df['Classification'].value_counts()
for cls, count in class_counts.items():
    print(f"  {cls}: {count}")

# Bank of America check
boa = validated_df[validated_df['Account'].str.contains('Bank of America', case=False, na=False)]
if len(boa) > 0:
    print(f"\nBank of America: Owner={boa.iloc[0]['Owner']}, CloseDate='{boa.iloc[0]['CloseDate']}'")

