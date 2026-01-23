"""
Simplified BL Review Structure
==============================
BL Review = Simple inputs for sales to verify
Inputs_formula = Formulas pulling from BL Review + calculations
Summary = Formulas pulling from Inputs_formula
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from collections import defaultdict

print("="*80)
print("CREATING SIMPLIFIED BL REVIEW STRUCTURE")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[1] Loading data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
print("\n[2] Processing accounts...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
    return None

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

# Build data for each account
bl_review_data = []
source_data = []

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) == 0:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    # SF match
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    # First meeting date
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min()
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        first_mtg = first_meeting_sf
    else:
        first_mtg = first_meeting_audit
    
    # Classify meetings
    meeting_classes = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_mtg) if first_mtg else (idx == 0)
        cls = classify_meeting(row['Subject'], is_first)
        days_from_first = (row['Date'] - first_mtg).days if first_mtg and pd.notna(row['Date']) else 0
        meeting_classes[cls].append({
            'date': row['Date'],
            'days': days_from_first,
            'subject': row['Subject']
        })
        
        # Add to source data
        source_data.append({
            'Account': orig_name,
            'Date': row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',
            'Subject': str(row['Subject'])[:60] if pd.notna(row['Subject']) else '',
            'Classification': cls,
            'DaysFromFirst': days_from_first,
        })
    
    # Calculate metrics
    demos = meeting_classes.get('Demo', [])
    cabs = meeting_classes.get('CAB', [])
    scopings = meeting_classes.get('Scoping', [])
    
    demo_count = len(demos)
    days_to_demo = min([d['days'] for d in demos]) if demos else ''
    
    cab_count = len(cabs)
    days_to_cab = min([d['days'] for d in cabs]) if cabs else ''
    
    scoping_count = len(scopings)
    days_to_scoping = min([d['days'] for d in scopings]) if scopings else ''
    
    # Close date
    if has_no_close(orig_name):
        close_date = ''
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close'].strftime('%m/%d/%Y')
    else:
        close_date = ''
    
    bl_review_data.append({
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Account': orig_name,
        'Include': 'Y',
        'Verified': '',
        'FirstMeetingDate': first_mtg.strftime('%m/%d/%Y') if first_mtg else '',
        'CloseDate': close_date,
        'DemoCount': demo_count,
        'DaysToDemo_Override': '',  # BL can override if calculated is wrong
        'CABCount': cab_count,
        'ScopingCount': scoping_count,
        'Notes': '',
        # Hidden/calculated fields for reference
        '_DaysToDemo_Calc': days_to_demo,
        '_DaysToCAB_Calc': days_to_cab,
        '_DaysToScoping_Calc': days_to_scoping,
        '_TotalMeetings': len(acct_meetings),
    })

print(f"  Processed {len(bl_review_data)} accounts")

# =============================================================================
# GENERATE EXCEL
# =============================================================================
print("\n[3] Generating Excel workbook...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

# =============================================================================
# TAB 1: BL_Review (Simplified for sales to verify)
# =============================================================================
ws_bl = wb.active
ws_bl.title = 'BL_Review'

# Simple columns for BL to verify
bl_columns = [
    ('A', 'Owner', 18, 'For filtering by rep'),
    ('B', 'Account', 28, 'Account name'),
    ('C', 'Include', 8, 'Y=include, N=exclude'),
    ('D', 'Verified', 8, 'Mark Y when confirmed'),
    ('E', 'FirstMeetingDate', 14, 'VERIFY: First meeting date'),
    ('F', 'CloseDate', 12, 'VERIFY: Deal close date'),
    ('G', 'DemoCount', 10, 'VERIFY: # of demos'),
    ('H', 'DaysToDemo', 11, 'Override if wrong'),
    ('I', 'CABCount', 9, 'VERIFY: # CAB mtgs'),
    ('J', 'ScopingCount', 11, 'VERIFY: # scoping mtgs'),
    ('K', 'Notes', 30, 'Add notes here'),
]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
verify_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for fields to verify
override_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green for optional override

# Headers
for col_idx, (col_letter, col_name, width, tooltip) in enumerate(bl_columns, 1):
    cell = ws_bl.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    ws_bl.column_dimensions[col_letter].width = width

# Data rows
for row_idx, record in enumerate(bl_review_data, 2):
    ws_bl.cell(row=row_idx, column=1, value=record['Owner'])
    ws_bl.cell(row=row_idx, column=2, value=record['Account'])
    ws_bl.cell(row=row_idx, column=3, value=record['Include'])
    ws_bl.cell(row=row_idx, column=4, value=record['Verified'])
    
    # Fields to VERIFY (yellow background)
    cell_e = ws_bl.cell(row=row_idx, column=5, value=record['FirstMeetingDate'])
    cell_e.fill = verify_fill
    
    cell_f = ws_bl.cell(row=row_idx, column=6, value=record['CloseDate'])
    cell_f.fill = verify_fill
    
    cell_g = ws_bl.cell(row=row_idx, column=7, value=record['DemoCount'])
    cell_g.fill = verify_fill
    
    # Optional override (green background)
    cell_h = ws_bl.cell(row=row_idx, column=8, value=record['_DaysToDemo_Calc'] if record['_DaysToDemo_Calc'] != '' else '')
    cell_h.fill = override_fill
    
    cell_i = ws_bl.cell(row=row_idx, column=9, value=record['CABCount'])
    cell_i.fill = verify_fill
    
    cell_j = ws_bl.cell(row=row_idx, column=10, value=record['ScopingCount'])
    cell_j.fill = verify_fill
    
    ws_bl.cell(row=row_idx, column=11, value=record['Notes'])

# =============================================================================
# TAB 2: Inputs_formula (Formulas pulling from BL_Review)
# =============================================================================
ws_inputs = wb.create_sheet('Inputs_formula')

# Full columns for Summary formulas
input_columns = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount',
    'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance',
    'ComplianceCount', 'DaysToContract', 'ContractingCount', 'Notes'
]

calc_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue for calculated

for col_idx, col_name in enumerate(input_columns, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Data rows with formulas
for row_idx, record in enumerate(bl_review_data, 2):
    # A: Owner (from BL_Review)
    ws_inputs.cell(row=row_idx, column=1, value=f"=BL_Review!A{row_idx}")
    
    # B: Account (from BL_Review)
    ws_inputs.cell(row=row_idx, column=2, value=f"=BL_Review!B{row_idx}")
    
    # C: Include (from BL_Review)
    ws_inputs.cell(row=row_idx, column=3, value=f"=BL_Review!C{row_idx}")
    
    # D: Verified (from BL_Review)
    ws_inputs.cell(row=row_idx, column=4, value=f"=BL_Review!D{row_idx}")
    
    # E: TotalMeetings (static - from source data)
    ws_inputs.cell(row=row_idx, column=5, value=record['_TotalMeetings'])
    
    # F: FirstMeetingDate (from BL_Review)
    ws_inputs.cell(row=row_idx, column=6, value=f"=BL_Review!E{row_idx}")
    
    # G: CloseDate (from BL_Review)
    ws_inputs.cell(row=row_idx, column=7, value=f"=BL_Review!F{row_idx}")
    
    # H: SalesCycleDays (CALCULATED from dates)
    ws_inputs.cell(row=row_idx, column=8, 
                   value=f'=IF(AND(G{row_idx}<>"",F{row_idx}<>""),G{row_idx}-F{row_idx},0)').fill = calc_fill
    
    # I: DaysFirstToSecond (static from source)
    if len([s for s in source_data if s['Account'] == record['Account']]) >= 2:
        acct_source = [s for s in source_data if s['Account'] == record['Account']]
        if len(acct_source) >= 2:
            ws_inputs.cell(row=row_idx, column=9, value=acct_source[1]['DaysFromFirst'])
        else:
            ws_inputs.cell(row=row_idx, column=9, value=0)
    else:
        ws_inputs.cell(row=row_idx, column=9, value=0)
    
    # J: DaysToDemo (from BL_Review override column H)
    ws_inputs.cell(row=row_idx, column=10, value=f"=BL_Review!H{row_idx}")
    
    # K: DemoCount (from BL_Review)
    ws_inputs.cell(row=row_idx, column=11, value=f"=BL_Review!G{row_idx}")
    
    # L: DaysToCAB (calculated from source)
    ws_inputs.cell(row=row_idx, column=12, value=record['_DaysToCAB_Calc'] if record['_DaysToCAB_Calc'] != '' else 0)
    
    # M: CABCount (from BL_Review)
    ws_inputs.cell(row=row_idx, column=13, value=f"=BL_Review!I{row_idx}")
    
    # N: DaysToScoping (calculated from source)
    ws_inputs.cell(row=row_idx, column=14, value=record['_DaysToScoping_Calc'] if record['_DaysToScoping_Calc'] != '' else 0)
    
    # O: ScopingCount (from BL_Review)
    ws_inputs.cell(row=row_idx, column=15, value=f"=BL_Review!J{row_idx}")
    
    # P-S: Compliance/Contracting (static from source for now)
    ws_inputs.cell(row=row_idx, column=16, value=0)  # DaysToCompliance
    ws_inputs.cell(row=row_idx, column=17, value=0)  # ComplianceCount
    ws_inputs.cell(row=row_idx, column=18, value=0)  # DaysToContract
    ws_inputs.cell(row=row_idx, column=19, value=0)  # ContractingCount
    
    # T: Notes (from BL_Review)
    ws_inputs.cell(row=row_idx, column=20, value=f"=BL_Review!K{row_idx}")

ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 28

# =============================================================================
# TAB 3: Source_Meetings (Reference)
# =============================================================================
ws_source = wb.create_sheet('Source_Meetings')

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'DaysFromFirst']
for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, row_data in enumerate(source_data, 2):
    for col_idx, col_name in enumerate(source_cols, 1):
        ws_source.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 50

# =============================================================================
# TAB 4: Instructions
# =============================================================================
ws_inst = wb.create_sheet('Instructions')

instructions = [
    ['SIMPLIFIED BL REVIEW WORKFLOW', ''],
    ['', ''],
    ['HOW IT WORKS:', ''],
    ['1. BL_Review tab', 'Sales verifies key fields (yellow = verify, green = override if wrong)'],
    ['2. Inputs_formula tab', 'Auto-pulls from BL_Review + calculates derived fields'],
    ['3. Summary tab', 'Your existing Summary formulas pull from Inputs_formula'],
    ['', ''],
    ['BL_REVIEW COLUMNS:', ''],
    ['Owner', 'For filtering - which rep owns this account'],
    ['Account', 'Account name'],
    ['Include', 'Y = include in summary, N = exclude'],
    ['Verified', 'BL marks Y when they have confirmed the data'],
    ['FirstMeetingDate (YELLOW)', 'VERIFY: Is this the correct first meeting date?'],
    ['CloseDate (YELLOW)', 'VERIFY: Is this the correct close date?'],
    ['DemoCount (YELLOW)', 'VERIFY: How many demos did we do?'],
    ['DaysToDemo (GREEN)', 'Override: Days from first meeting to first demo (edit if wrong)'],
    ['CABCount (YELLOW)', 'VERIFY: How many CAB meetings?'],
    ['ScopingCount (YELLOW)', 'VERIFY: How many scoping meetings?'],
    ['Notes', 'Add any notes or corrections'],
    ['', ''],
    ['CALCULATED FIELDS (in Inputs_formula):', ''],
    ['SalesCycleDays', '= CloseDate - FirstMeetingDate (auto-calculated)'],
    ['DaysFirstToSecond', 'From Source_Meetings data'],
    ['DaysToDemo', 'From BL_Review override column'],
    ['', ''],
    ['TO COPY TO YOUR WORKBOOK:', ''],
    ['1.', 'Copy BL_Review tab to your BL Review tab'],
    ['2.', 'Copy Inputs_formula tab to your Inputs_formula tab'],
    ['3.', 'Summary formulas will auto-update'],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_inst.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)

ws_inst.column_dimensions['A'].width = 25
ws_inst.column_dimensions['B'].width = 60

# Save
output_path = '/Users/keiganpesenti/Desktop/SIMPLIFIED_BL_REVIEW.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("STRUCTURE")
print("="*80)
print("""
BL_Review (Simple - Sales verifies):
  A: Owner           - For filtering
  B: Account         - Account name
  C: Include         - Y/N
  D: Verified        - Mark Y when confirmed
  E: FirstMeetingDate (YELLOW) - VERIFY
  F: CloseDate       (YELLOW) - VERIFY
  G: DemoCount       (YELLOW) - VERIFY
  H: DaysToDemo      (GREEN)  - Override if calculated is wrong
  I: CABCount        (YELLOW) - VERIFY
  J: ScopingCount    (YELLOW) - VERIFY
  K: Notes

Inputs_formula (Formulas pull from BL_Review):
  - Owner, Account, Include, Verified → =BL_Review!...
  - SalesCycleDays → =CloseDate-FirstMeetingDate (calculated)
  - DemoCount, CABCount, etc. → =BL_Review!...
  
Summary → Pulls from Inputs_formula (your existing formulas)
""")

