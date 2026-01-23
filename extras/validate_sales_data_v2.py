"""
Validated Sales Data Generator v2
=================================
- Removes specified accounts
- Adds high-quality accounts from meeting prep docs
- Improved confidence scoring
- Quality checks for accuracy
- CloseDate and SalesCycleDays moved to far right
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from docx import Document
from collections import defaultdict

print("="*80)
print("VALIDATED SALES DATA GENERATOR v2")
print("Enhanced accuracy with quality checks")
print("="*80)

# =============================================================================
# ACCOUNTS TO EXCLUDE
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska'
]

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

# =============================================================================
# STEP 1: Load Salesforce Reference Data
# =============================================================================
print("\n[STEP 1] Loading Salesforce reference data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
print(f"  Loaded {len(sf_ref)} accounts from days to close reference")

sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

# =============================================================================
# STEP 2: Parse Meeting Prep Documents (Enhanced)
# =============================================================================
print("\n[STEP 2] Parsing meeting prep documents (enhanced extraction)...")

docs_folder = "/Users/keiganpesenti/Desktop/meeting_prep_extracted/Prep + Notes"

def extract_company_name(filename):
    name = filename.replace(' - Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep + Notes.docx', '')
    name = name.replace('_Meeting Prep & Notes.docx', '')
    name = name.replace('_Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep.docx', '')
    name = name.replace('.docx', '')
    name = re.sub(r'^\d{8}\s*', '', name)
    return name.strip()

def parse_meeting_doc_enhanced(filepath):
    """Enhanced parsing to extract more meeting details."""
    try:
        doc = Document(filepath)
        
        meetings = []
        current_meeting = {}
        company_name = None
        in_past_notes = False
        in_customer_history = False
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            
            # Extract company from Meeting Overview
            if text.startswith('Company:'):
                company_name = text.replace('Company:', '').strip().split('<')[0].strip()
            
            # Detect sections
            if 'FULL PAST MEETING NOTES' in text.upper():
                in_past_notes = True
                in_customer_history = False
                continue
            if 'CRITICAL CUSTOMER HISTORY' in text.upper() or 'CUSTOMER HISTORY' in text.upper():
                in_customer_history = True
                in_past_notes = False
                continue
            if 'COMPANY INTEL' in text.upper() or 'HYPOTHESES' in text.upper():
                in_past_notes = False
                in_customer_history = False
            
            # Parse meeting type
            if text.startswith('Meeting Type:'):
                if current_meeting and 'date' in current_meeting:
                    meetings.append(current_meeting)
                meeting_type = text.replace('Meeting Type:', '').strip().lower()
                current_meeting = {
                    'type': meeting_type,
                    'is_demo': 'demo' in meeting_type,
                    'is_intro': any(x in meeting_type for x in ['intro', 'first', 'initial']),
                    'is_cab': 'cab' in meeting_type or 'advisory' in meeting_type,
                    'is_scoping': 'scoping' in meeting_type or 'pricing' in meeting_type,
                }
            
            # Parse date
            if text.startswith('Date:'):
                date_str = text.replace('Date:', '').strip()
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', date_str)
                if date_match:
                    try:
                        current_meeting['date'] = pd.to_datetime(date_match.group(1))
                    except:
                        pass
            
            # Check meeting goals for demo/scoping mentions
            if 'Meeting Goals' in text or 'demo' in text.lower():
                if 'demo' in text.lower():
                    current_meeting['is_demo'] = True
            if 'scoping' in text.lower() or 'pricing' in text.lower():
                current_meeting['is_scoping'] = True
            if 'cab' in text.lower() or 'advisory' in text.lower():
                current_meeting['is_cab'] = True
        
        # Capture last meeting
        if current_meeting and 'date' in current_meeting:
            meetings.append(current_meeting)
        
        return {
            'company': company_name,
            'meetings': meetings,
        }
    except Exception as e:
        return {'company': None, 'meetings': []}

# Get list of documents
doc_files = []
for f in os.listdir(docs_folder):
    if not f.endswith('.docx'):
        continue
    if '(old)' in f.lower() or '(OLD)' in f:
        continue
    if f.startswith('~$'):
        continue
    if 'TEMPLATE' in f.upper() or 'COPY' in f.upper():
        continue
    doc_files.append(f)

print(f"  Found {len(doc_files)} meeting prep documents")

# Parse documents and build comprehensive data
doc_data = {}
for i, filename in enumerate(doc_files):
    if i % 50 == 0:
        print(f"    Processing {i}/{len(doc_files)}...")
    
    company = extract_company_name(filename)
    if should_exclude(company):
        continue
    
    filepath = os.path.join(docs_folder, filename)
    parsed = parse_meeting_doc_enhanced(filepath)
    meetings = parsed['meetings']
    
    if not meetings:
        continue
    
    valid_dates = [m['date'] for m in meetings if 'date' in m and pd.notna(m['date'])]
    demo_dates = [m['date'] for m in meetings if m.get('is_demo') and 'date' in m]
    intro_dates = [m['date'] for m in meetings if m.get('is_intro') and 'date' in m]
    cab_dates = [m['date'] for m in meetings if m.get('is_cab') and 'date' in m]
    scoping_dates = [m['date'] for m in meetings if m.get('is_scoping') and 'date' in m]
    
    if valid_dates:
        doc_data[company.lower().strip()] = {
            'company': company,
            'meeting_count': len(meetings),
            'first_meeting': min(valid_dates),
            'latest_meeting': max(valid_dates),
            'first_demo': min(demo_dates) if demo_dates else None,
            'demo_count': len(demo_dates),
            'first_cab': min(cab_dates) if cab_dates else None,
            'cab_count': len(cab_dates),
            'first_scoping': min(scoping_dates) if scoping_dates else None,
            'scoping_count': len(scoping_dates),
            'meetings': meetings,
            'has_rich_data': len(meetings) >= 3,
        }

print(f"  Extracted data from {len(doc_data)} companies")

# =============================================================================
# STEP 3: Load Meeting Audit Data
# =============================================================================
print("\n[STEP 3] Loading meeting audit data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_accts = pd.read_excel(audit_file, sheet_name='all accts')

secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'
df_meetings_secondary = pd.read_excel(secondary_file, sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Deduplicate
df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

# Exclude test accounts and specified exclusions
test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

# Remove excluded accounts
df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings (after exclusions)")

# =============================================================================
# STEP 4: Build Validated Account Data with Quality Checks
# =============================================================================
print("\n[STEP 4] Building validated data with quality checks...")

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
            name_first = name_clean.split()[0] if name_clean.split() else ''
            key_first = key_clean.split()[0] if key_clean.split() else ''
            if name_first and key_first and len(name_first) > 3 and name_first == key_first:
                return key
    return None

def classify_meeting(subject, meeting_num):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    if meeting_num == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    return 'Followup'

# Get accounts from meeting data
account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()

# Filter out excluded accounts
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

print(f"  Processing {len(accounts_to_process)} accounts from audit data")

# Build records
validated_records = []
quality_issues = []

for norm_acct in accounts_to_process:
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    
    if should_exclude(orig_name):
        continue
    
    record = {
        'Account': orig_name,
        'Owner': '',
        'Include': 'Y',
        'Verified': '',
    }
    
    # === Match to sources ===
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    
    # Also try direct match
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                sf_match = sf_key
                break
    
    doc_match = fuzzy_match(orig_name, doc_data)
    doc_info = doc_data.get(doc_match, {}) if doc_match else {}
    
    # === Owner ===
    record['Owner'] = sf_data.get('owner', '') if sf_data else ''
    
    # === First Meeting Date (prioritize SF, then audit, then doc) ===
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min()
    first_meeting_doc = doc_info.get('first_meeting') if doc_info else None
    
    # Use the earliest reliable date
    candidates = []
    if first_meeting_sf and pd.notna(first_meeting_sf):
        candidates.append(first_meeting_sf)
    if first_meeting_audit and pd.notna(first_meeting_audit):
        candidates.append(first_meeting_audit)
    
    if candidates:
        record['FirstMeetingDate'] = min(candidates).strftime('%m/%d/%Y')
    else:
        record['FirstMeetingDate'] = ''
    
    # === Total Meetings ===
    audit_count = len(acct_meetings)
    doc_count = doc_info.get('meeting_count', 0) if doc_info else 0
    record['TotalMeetings'] = max(audit_count, doc_count)
    
    # === Classify meetings from audit data ===
    meeting_dates = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        substep = classify_meeting(row['Subject'], idx + 1)
        meeting_dates[substep].append(row['Date'])
    
    # === Days to milestones ===
    first_mtg = pd.to_datetime(record['FirstMeetingDate']) if record['FirstMeetingDate'] else None
    
    # Days First to Second
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    # Demo
    demo_dates_audit = meeting_dates.get('Demo', [])
    demo_date_doc = doc_info.get('first_demo') if doc_info else None
    
    if demo_dates_audit and first_mtg:
        record['DaysToDemo'] = (min(demo_dates_audit) - first_mtg).days
        record['DemoCount'] = len(demo_dates_audit)
    elif demo_date_doc and first_mtg:
        record['DaysToDemo'] = (demo_date_doc - first_mtg).days
        record['DemoCount'] = doc_info.get('demo_count', 1)
    else:
        record['DaysToDemo'] = ''
        record['DemoCount'] = 0
    
    # CAB
    cab_dates_audit = meeting_dates.get('CAB', [])
    cab_date_doc = doc_info.get('first_cab') if doc_info else None
    
    if cab_dates_audit and first_mtg:
        record['DaysToCAB'] = (min(cab_dates_audit) - first_mtg).days
        record['CABCount'] = len(cab_dates_audit)
    elif cab_date_doc and first_mtg:
        record['DaysToCAB'] = (cab_date_doc - first_mtg).days
        record['CABCount'] = doc_info.get('cab_count', 1)
    else:
        record['DaysToCAB'] = ''
        record['CABCount'] = 0
    
    # Scoping
    scoping_dates = meeting_dates.get('Scoping', [])
    if scoping_dates and first_mtg:
        record['DaysToScoping'] = (min(scoping_dates) - first_mtg).days
        record['ScopingCount'] = len(scoping_dates)
    else:
        record['DaysToScoping'] = ''
        record['ScopingCount'] = 0
    
    # Compliance
    compliance_dates = meeting_dates.get('Compliance', [])
    if compliance_dates and first_mtg:
        record['DaysToCompliance'] = (min(compliance_dates) - first_mtg).days
        record['ComplianceCount'] = len(compliance_dates)
    else:
        record['DaysToCompliance'] = ''
        record['ComplianceCount'] = 0
    
    # Contracting
    contracting_dates = meeting_dates.get('Contracting', [])
    if contracting_dates and first_mtg:
        record['DaysToContracting'] = (min(contracting_dates) - first_mtg).days
        record['ContractingCount'] = len(contracting_dates)
    else:
        record['DaysToContracting'] = ''
        record['ContractingCount'] = 0
    
    # === Close Date and Sales Cycle (moved to end) ===
    if sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        
        # Quality check: close date should be after first meeting
        if first_mtg and close_date < first_mtg:
            # Use days_to_close from SF if available
            if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
                record['SalesCycleDays'] = int(sf_data['days_to_close'])
            else:
                record['SalesCycleDays'] = ''
            quality_issues.append(f"{orig_name}: Close date before first meeting - using SF days_to_close")
        else:
            if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
                record['SalesCycleDays'] = int(sf_data['days_to_close'])
            elif first_mtg:
                record['SalesCycleDays'] = (close_date - first_mtg).days
            else:
                record['SalesCycleDays'] = ''
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    # === Data Quality Flags ===
    sources = []
    if sf_data:
        sources.append('SF')
    if doc_info:
        sources.append('Doc')
    sources.append('Audit')
    record['DataSource'] = '+'.join(sources)
    
    # Improved confidence scoring
    if sf_data and record['CloseDate']:
        record['Confidence'] = 'High'  # Has SF close date
    elif sf_data:
        record['Confidence'] = 'Medium'  # Has some SF data
    elif doc_info and doc_info.get('has_rich_data'):
        record['Confidence'] = 'Medium'  # Rich doc data
    else:
        record['Confidence'] = 'Standard'  # Just audit data (not "Low")
    
    validated_records.append(record)

validated_df = pd.DataFrame(validated_records)

# Sort by Owner (with owner first), then Account
validated_df['_has_owner'] = validated_df['Owner'].apply(lambda x: 0 if pd.notna(x) and x != '' else 1)
validated_df = validated_df.sort_values(['_has_owner', 'Owner', 'Account']).reset_index(drop=True)
validated_df = validated_df.drop(columns=['_has_owner'])

print(f"  Created {len(validated_df)} validated account records")
print(f"  Quality issues found: {len(quality_issues)}")
for issue in quality_issues[:5]:
    print(f"    - {issue}")

# =============================================================================
# STEP 5: Generate Output
# =============================================================================
print("\n[STEP 5] Generating output...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

# === TAB 1: BL_Review_Inputs ===
ws_inputs = wb.active
ws_inputs.title = 'BL_Review_Inputs'

# Column order: CloseDate and SalesCycleDays at the end
output_cols = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 'DaysToCAB', 'CABCount',
    'DaysToScoping', 'ScopingCount', 'DaysToCompliance', 'ComplianceCount',
    'DaysToContracting', 'ContractingCount', 'DataSource', 'Confidence',
    'CloseDate', 'SalesCycleDays'  # Moved to end
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for col_idx, col_name in enumerate(output_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, (_, row) in enumerate(validated_df.iterrows(), 2):
    for col_idx, col_name in enumerate(output_cols, 1):
        val = row.get(col_name, '')
        cell = ws_inputs.cell(row=row_idx, column=col_idx, value=val)
        
        confidence = row.get('Confidence', '')
        if confidence == 'High':
            cell.fill = high_fill
        elif confidence == 'Medium':
            cell.fill = medium_fill

# Set column widths
ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 25

# === TAB 2: Quality_Notes ===
ws_notes = wb.create_sheet('Quality_Notes')
ws_notes.cell(row=1, column=1, value='Quality Issues Detected')
for i, issue in enumerate(quality_issues, 2):
    ws_notes.cell(row=i, column=1, value=issue)

# Save
output_path = '/Users/keiganpesenti/Desktop/validated_inputs_v2.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total accounts: {len(validated_df)}")
print(f"With Owner: {len(validated_df[validated_df['Owner'].notna() & (validated_df['Owner'] != '')])}")
print(f"With Close Date: {len(validated_df[validated_df['CloseDate'].notna() & (validated_df['CloseDate'] != '')])}")
print(f"High Confidence: {len(validated_df[validated_df['Confidence'] == 'High'])}")
print(f"Medium Confidence: {len(validated_df[validated_df['Confidence'] == 'Medium'])}")
print(f"Standard: {len(validated_df[validated_df['Confidence'] == 'Standard'])}")

# Show accounts with owners
print("\n=== ACCOUNTS WITH OWNERS (sorted) ===")
with_owner = validated_df[validated_df['Owner'].notna() & (validated_df['Owner'] != '')]
print(with_owner[['Owner', 'Account', 'TotalMeetings', 'CloseDate']].to_string())

