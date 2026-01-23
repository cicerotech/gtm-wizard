"""
Clean BL Review Structure
=========================
- Dates stored as Excel dates (not text) so formulas work
- Simplified fields for BL to verify
- CAB as Y/N not count
- Remove confusing Scoping count
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from collections import defaultdict

print("="*80)
print("CREATING CLEAN BL REVIEW STRUCTURE")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[1] Loading data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
print("\n[2] Processing accounts...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
    return None

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

# Build data
bl_review_data = []
source_data = []

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) == 0:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    # SF match
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    # First meeting date
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min()
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        first_mtg = first_meeting_sf
    else:
        first_mtg = first_meeting_audit
    
    # Classify meetings
    meeting_classes = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_mtg) if first_mtg else (idx == 0)
        cls = classify_meeting(row['Subject'], is_first)
        days_from_first = (row['Date'] - first_mtg).days if first_mtg and pd.notna(row['Date']) else 0
        meeting_classes[cls].append({
            'date': row['Date'],
            'days': days_from_first,
        })
        
        source_data.append({
            'Account': orig_name,
            'Date': row['Date'],
            'Subject': str(row['Subject'])[:60] if pd.notna(row['Subject']) else '',
            'Classification': cls,
            'DaysFromFirst': days_from_first,
        })
    
    # Calculate metrics
    demos = meeting_classes.get('Demo', [])
    cabs = meeting_classes.get('CAB', [])
    compliances = meeting_classes.get('Compliance', [])
    contractings = meeting_classes.get('Contracting', [])
    
    demo_count = len(demos)
    days_to_demo = min([d['days'] for d in demos]) if demos else 0
    
    had_cab = 'Y' if len(cabs) > 0 else ''
    had_compliance = 'Y' if len(compliances) > 0 else ''
    had_contracting = 'Y' if len(contractings) > 0 else ''
    
    # Days first to second
    if len(acct_meetings) >= 2:
        days_first_to_second = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        days_first_to_second = 0
    
    # Close date
    if has_no_close(orig_name):
        close_date = None
        sales_cycle = None
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            sales_cycle = int(sf_data['days_to_close'])
        elif first_mtg:
            sales_cycle = (close_date - first_mtg).days
        else:
            sales_cycle = None
    else:
        close_date = None
        sales_cycle = None
    
    bl_review_data.append({
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Account': orig_name,
        'Include': 'Y',
        'Verified': '',
        'FirstMeetingDate': first_mtg,  # Keep as datetime
        'CloseDate': close_date,  # Keep as datetime
        'TotalMeetings': len(acct_meetings),
        'DemoCount': demo_count,
        'HadCAB': had_cab,
        'Notes': '',
        # For Inputs_formula (calculated)
        '_SalesCycleDays': sales_cycle,
        '_DaysFirstToSecond': days_first_to_second,
        '_DaysToDemo': days_to_demo,
        '_CABCount': len(cabs),
        '_ComplianceCount': len(compliances),
        '_ContractingCount': len(contractings),
    })

print(f"  Processed {len(bl_review_data)} accounts")

# =============================================================================
# GENERATE EXCEL
# =============================================================================
print("\n[3] Generating Excel workbook...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

# =============================================================================
# TAB 1: BL_Review (Super Simple)
# =============================================================================
ws_bl = wb.active
ws_bl.title = 'BL_Review'

# SIMPLE columns - only what BLs need to verify
bl_columns = [
    ('A', 'Owner', 18),
    ('B', 'Account', 28),
    ('C', 'Include', 8),
    ('D', 'Verified', 8),
    ('E', 'FirstMeetingDate', 15),
    ('F', 'CloseDate', 12),
    ('G', 'TotalMeetings', 12),
    ('H', 'DemoCount', 10),
    ('I', 'HadCAB', 8),
    ('J', 'Notes', 30),
]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
verify_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for col_idx, (col_letter, col_name, width) in enumerate(bl_columns, 1):
    cell = ws_bl.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    ws_bl.column_dimensions[col_letter].width = width

for row_idx, record in enumerate(bl_review_data, 2):
    ws_bl.cell(row=row_idx, column=1, value=record['Owner'])
    ws_bl.cell(row=row_idx, column=2, value=record['Account'])
    ws_bl.cell(row=row_idx, column=3, value=record['Include'])
    ws_bl.cell(row=row_idx, column=4, value=record['Verified'])
    
    # FirstMeetingDate - as actual date
    cell_e = ws_bl.cell(row=row_idx, column=5)
    if record['FirstMeetingDate']:
        cell_e.value = record['FirstMeetingDate']
        cell_e.number_format = 'MM/DD/YYYY'
    cell_e.fill = verify_fill
    
    # CloseDate - as actual date
    cell_f = ws_bl.cell(row=row_idx, column=6)
    if record['CloseDate']:
        cell_f.value = record['CloseDate']
        cell_f.number_format = 'MM/DD/YYYY'
    cell_f.fill = verify_fill
    
    # TotalMeetings
    cell_g = ws_bl.cell(row=row_idx, column=7, value=record['TotalMeetings'])
    cell_g.fill = verify_fill
    
    # DemoCount
    cell_h = ws_bl.cell(row=row_idx, column=8, value=record['DemoCount'])
    cell_h.fill = verify_fill
    
    # HadCAB (Y/N)
    cell_i = ws_bl.cell(row=row_idx, column=9, value=record['HadCAB'])
    cell_i.fill = verify_fill
    
    ws_bl.cell(row=row_idx, column=10, value=record['Notes'])

# =============================================================================
# TAB 2: Inputs_formula (Formulas + calculated fields)
# =============================================================================
ws_inputs = wb.create_sheet('Inputs_formula')

input_columns = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount',
    'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance',
    'ComplianceCount', 'DaysToContract', 'ContractingCount', 'Notes'
]

for col_idx, col_name in enumerate(input_columns, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, record in enumerate(bl_review_data, 2):
    # A-D: From BL_Review
    ws_inputs.cell(row=row_idx, column=1, value=f"=BL_Review!A{row_idx}")
    ws_inputs.cell(row=row_idx, column=2, value=f"=BL_Review!B{row_idx}")
    ws_inputs.cell(row=row_idx, column=3, value=f"=BL_Review!C{row_idx}")
    ws_inputs.cell(row=row_idx, column=4, value=f"=BL_Review!D{row_idx}")
    
    # E: TotalMeetings
    ws_inputs.cell(row=row_idx, column=5, value=f"=BL_Review!G{row_idx}")
    
    # F: FirstMeetingDate
    ws_inputs.cell(row=row_idx, column=6, value=f"=BL_Review!E{row_idx}")
    
    # G: CloseDate
    ws_inputs.cell(row=row_idx, column=7, value=f"=BL_Review!F{row_idx}")
    
    # H: SalesCycleDays = CloseDate - FirstMeetingDate (FIXED formula)
    ws_inputs.cell(row=row_idx, column=8, 
                   value=f'=IF(AND(G{row_idx}<>"",F{row_idx}<>""),G{row_idx}-F{row_idx},0)')
    
    # I: DaysFirstToSecond (static from source)
    ws_inputs.cell(row=row_idx, column=9, value=record['_DaysFirstToSecond'])
    
    # J: DaysToDemo (static from source)
    ws_inputs.cell(row=row_idx, column=10, value=record['_DaysToDemo'])
    
    # K: DemoCount
    ws_inputs.cell(row=row_idx, column=11, value=f"=BL_Review!H{row_idx}")
    
    # L: DaysToCAB (static - 0 if no CAB)
    ws_inputs.cell(row=row_idx, column=12, value=0)
    
    # M: CABCount (from HadCAB Y/N)
    ws_inputs.cell(row=row_idx, column=13, value=f'=IF(BL_Review!I{row_idx}="Y",1,0)')
    
    # N-O: Scoping (static)
    ws_inputs.cell(row=row_idx, column=14, value=0)
    ws_inputs.cell(row=row_idx, column=15, value=0)
    
    # P-Q: Compliance (static)
    ws_inputs.cell(row=row_idx, column=16, value=0)
    ws_inputs.cell(row=row_idx, column=17, value=record['_ComplianceCount'])
    
    # R-S: Contracting (static)
    ws_inputs.cell(row=row_idx, column=18, value=0)
    ws_inputs.cell(row=row_idx, column=19, value=record['_ContractingCount'])
    
    # T: Notes
    ws_inputs.cell(row=row_idx, column=20, value=f"=BL_Review!J{row_idx}")

ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 28
ws_inputs.column_dimensions['F'].width = 14
ws_inputs.column_dimensions['G'].width = 12

# =============================================================================
# TAB 3: Source_Meetings
# =============================================================================
ws_source = wb.create_sheet('Source_Meetings')

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'DaysFromFirst']
for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, row_data in enumerate(source_data, 2):
    ws_source.cell(row=row_idx, column=1, value=row_data['Account'])
    cell_date = ws_source.cell(row=row_idx, column=2)
    if row_data['Date']:
        cell_date.value = row_data['Date']
        cell_date.number_format = 'MM/DD/YYYY'
    ws_source.cell(row=row_idx, column=3, value=row_data['Subject'])
    ws_source.cell(row=row_idx, column=4, value=row_data['Classification'])
    ws_source.cell(row=row_idx, column=5, value=row_data['DaysFromFirst'])

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 50

# =============================================================================
# TAB 4: Instructions
# =============================================================================
ws_inst = wb.create_sheet('Instructions')

instructions = [
    ['BL REVIEW - SIMPLIFIED', ''],
    ['', ''],
    ['BL_Review TAB (What Sales Verifies):', ''],
    ['Column', 'What to verify'],
    ['Owner', 'For filtering - which rep owns this account'],
    ['Account', 'Account name'],
    ['Include', 'Y = include, change to N to exclude from summary'],
    ['Verified', 'Mark Y once you have confirmed the data is correct'],
    ['FirstMeetingDate', 'VERIFY: Is this the correct date of our first meeting?'],
    ['CloseDate', 'VERIFY: Is this the correct deal close date?'],
    ['TotalMeetings', 'VERIFY: Total number of meetings we had'],
    ['DemoCount', 'VERIFY: How many demo/product meetings did we have?'],
    ['HadCAB', 'VERIFY: Did we have a CAB (Customer Advisory Board) meeting? Y or blank'],
    ['Notes', 'Add any corrections or notes here'],
    ['', ''],
    ['WHAT GETS CALCULATED AUTOMATICALLY:', ''],
    ['SalesCycleDays', '= CloseDate - FirstMeetingDate'],
    ['CABCount', '= 1 if HadCAB is Y, else 0'],
    ['', ''],
    ['HOW IT FLOWS:', ''],
    ['1.', 'You edit BL_Review tab'],
    ['2.', 'Inputs_formula tab auto-updates (pulls from BL_Review + calculates)'],
    ['3.', 'Summary tab auto-updates (pulls from Inputs_formula)'],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_inst.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        if row_idx == 4:
            cell.font = Font(bold=True)

ws_inst.column_dimensions['A'].width = 20
ws_inst.column_dimensions['B'].width = 60

# Save
output_path = '/Users/keiganpesenti/Desktop/CLEAN_BL_REVIEW.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("BL_REVIEW COLUMNS (SIMPLIFIED)")
print("="*80)
print("""
  A: Owner           - For filtering
  B: Account         - Account name
  C: Include         - Y/N
  D: Verified        - Mark Y when confirmed
  E: FirstMeetingDate (YELLOW) - VERIFY: correct first meeting date?
  F: CloseDate       (YELLOW) - VERIFY: correct close date?
  G: TotalMeetings   (YELLOW) - VERIFY: total meetings count
  H: DemoCount       (YELLOW) - VERIFY: how many demos?
  I: HadCAB          (YELLOW) - VERIFY: Y if had CAB meeting
  J: Notes           - Add corrections

DATES ARE STORED AS EXCEL DATES (not text)
- SalesCycleDays formula will work correctly: =CloseDate-FirstMeetingDate
""")

# Verification
print("\n=== SAMPLE DATA ===")
for record in bl_review_data[:5]:
    print(f"  {record['Account']}: FirstMtg={record['FirstMeetingDate']}, Close={record['CloseDate']}, Cycle={record['_SalesCycleDays']}")

