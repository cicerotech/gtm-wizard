"""
Validated Sales Data Generator v4
=================================
- Enriched from meeting prep docs
- Cross-references opportunities data
- Better confidence scoring
- Comprehensive audit notes
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from docx import Document
from collections import defaultdict

print("="*80)
print("VALIDATED SALES DATA GENERATOR v4")
print("Enriched data with opportunity cross-reference")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# STEP 1: Load All Reference Data
# =============================================================================
print("\n[STEP 1] Loading all reference data...")

# SF Reference
sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }
print(f"  SF reference: {len(sf_lookup)} accounts")

# Opportunity History
audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_opps = pd.read_excel(audit_file, sheet_name='all opps')
print(f"  Opportunities: {len(df_opps)} records")

# Build opp lookup by account
opp_lookup = defaultdict(list)
for _, row in df_opps.iterrows():
    acct = str(row.get('Account Name', '')).lower().strip() if pd.notna(row.get('Account Name')) else ''
    if acct:
        opp_lookup[acct].append({
            'stage': row.get('Stage'),
            'amount': row.get('Amount'),
            'close_date': pd.to_datetime(row.get('Close Date')) if pd.notna(row.get('Close Date')) else None,
        })

# =============================================================================
# STEP 2: Parse Meeting Prep Docs with Rich Detail
# =============================================================================
print("\n[STEP 2] Parsing meeting prep documents...")

docs_folder = "/Users/keiganpesenti/Desktop/meeting_prep_extracted/Prep + Notes"

def extract_company_name(filename):
    name = filename.replace(' - Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep + Notes.docx', '')
    name = name.replace('_Meeting Prep & Notes.docx', '')
    name = name.replace('_Meeting Prep + Notes.docx', '')
    name = name.replace(' Meeting Prep.docx', '')
    name = name.replace('.docx', '')
    name = re.sub(r'^\d{8}\s*', '', name)
    # Remove any remaining file artifacts
    name = re.sub(r'-[A-F0-9]{20,}$', '', name)  # Remove hash suffixes
    name = re.sub(r'Meeting Prep.*$', '', name, flags=re.IGNORECASE)
    return name.strip()

def parse_meeting_doc(filepath):
    """Parse document for meetings and customer history."""
    try:
        doc = Document(filepath)
        
        meetings = []
        current_meeting = {}
        company_name = None
        customer_history_notes = []
        in_customer_history = False
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            
            if text.startswith('Company:'):
                company_name = text.replace('Company:', '').strip().split('<')[0].strip()
            
            # Track customer history section
            if 'CUSTOMER HISTORY' in text.upper() or 'CRITICAL CUSTOMER HISTORY' in text.upper():
                in_customer_history = True
                continue
            
            if in_customer_history and text and len(text) > 10:
                # Look for dates in customer history
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})', text)
                if date_match:
                    customer_history_notes.append(text[:100])
            
            if 'COMPANY INTEL' in text.upper() or 'HYPOTHESES' in text.upper():
                in_customer_history = False
            
            if text.startswith('Meeting Type:'):
                if current_meeting and 'date' in current_meeting:
                    meetings.append(current_meeting)
                meeting_type = text.replace('Meeting Type:', '').strip()
                current_meeting = {
                    'type': meeting_type,
                    'type_lower': meeting_type.lower(),
                    'is_demo': any(x in meeting_type.lower() for x in ['demo', 'walkthrough', 'platform', 'sigma', 'cortex']),
                    'is_intro': any(x in meeting_type.lower() for x in ['intro', 'first', 'initial']),
                    'is_cab': any(x in meeting_type.lower() for x in ['cab', 'advisory']),
                    'is_scoping': any(x in meeting_type.lower() for x in ['scoping', 'pricing', 'proposal']),
                    'is_compliance': any(x in meeting_type.lower() for x in ['infosec', 'security', 'compliance']),
                    'is_contracting': any(x in meeting_type.lower() for x in ['contract', 'redline', 'msa', 'legal']),
                }
            
            if text.startswith('Date:'):
                date_str = text.replace('Date:', '').strip()
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', date_str)
                if date_match:
                    try:
                        current_meeting['date'] = pd.to_datetime(date_match.group(1))
                        current_meeting['date_str'] = date_match.group(1)
                    except:
                        pass
        
        if current_meeting and 'date' in current_meeting:
            meetings.append(current_meeting)
        
        return {
            'company': company_name, 
            'meetings': meetings,
            'customer_history': customer_history_notes,
            'has_rich_history': len(customer_history_notes) >= 2 or len(meetings) >= 3,
        }
    except:
        return {'company': None, 'meetings': [], 'customer_history': [], 'has_rich_history': False}

# Get documents
doc_files = [f for f in os.listdir(docs_folder) 
             if f.endswith('.docx') 
             and '(old)' not in f.lower() 
             and not f.startswith('~$')
             and 'TEMPLATE' not in f.upper()
             and 'COPY' not in f.upper()]

doc_data = {}
for i, filename in enumerate(doc_files):
    if i % 100 == 0:
        print(f"    Processing {i}/{len(doc_files)}...")
    
    company = extract_company_name(filename)
    if should_exclude(company):
        continue
    
    filepath = os.path.join(docs_folder, filename)
    parsed = parse_meeting_doc(filepath)
    meetings = parsed['meetings']
    
    key = company.lower().strip()
    if meetings or parsed['has_rich_history']:
        doc_data[key] = {
            'company': company,
            'meetings': meetings,
            'meeting_count': len(meetings),
            'customer_history': parsed['customer_history'],
            'has_rich_history': parsed['has_rich_history'],
        }

print(f"  Extracted data from {len(doc_data)} companies")

# =============================================================================
# STEP 3: Load Meeting Audit Data
# =============================================================================
print("\n[STEP 3] Loading meeting audit data...")

df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# STEP 4: Build Validated Data
# =============================================================================
print("\n[STEP 4] Building validated data...")

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
            name_first = name_clean.split()[0] if name_clean.split() else ''
            key_first = key_clean.split()[0] if key_clean.split() else ''
            if name_first and key_first and len(name_first) > 3 and name_first == key_first:
                return key
    return None

def classify_meeting_subject(subject, meeting_num):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_num == 1:
        return 'Intro', 'First meeting in sequence'
    if 'intro' in s or 'introduction' in s:
        return 'Intro', f'Subject: "{subject[:40]}"'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product']):
        return 'Demo', f'Subject: "{subject[:40]}"'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB', f'Subject: "{subject[:40]}"'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping', f'Subject: "{subject[:40]}"'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance', f'Subject: "{subject[:40]}"'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting', f'Subject: "{subject[:40]}"'
    return 'Followup', 'General meeting'

# Get accounts from meeting data
account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = set(account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist())

# Also add accounts from meeting prep docs with rich history
for key, info in doc_data.items():
    if info['has_rich_history'] and not should_exclude(info['company']):
        accounts_to_process.add(key)

accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]
print(f"  Processing {len(accounts_to_process)} accounts")

validated_records = []

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    
    if len(acct_meetings) > 0:
        acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
        orig_name = acct_meetings['Company / Account'].iloc[0]
    else:
        # Account from doc only
        doc_info = doc_data.get(norm_acct, {})
        if not doc_info:
            continue
        orig_name = doc_info.get('company', norm_acct)
    
    if should_exclude(orig_name):
        continue
    
    notes_parts = []
    demo_notes = []
    cab_notes = []
    scoping_notes = []
    compliance_notes = []
    contracting_notes = []
    
    # Match to sources
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    doc_match = fuzzy_match(orig_name, doc_data)
    doc_info = doc_data.get(doc_match, {}) if doc_match else {}
    
    opp_match = fuzzy_match(orig_name, opp_lookup)
    opp_data = opp_lookup.get(opp_match, []) if opp_match else []
    
    record = {
        'Account': orig_name,
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Include': 'Y',
        'Verified': '',
    }
    
    # === First Meeting Date ===
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min() if len(acct_meetings) > 0 else None
    first_meeting_doc = min([m['date'] for m in doc_info.get('meetings', []) if 'date' in m], default=None) if doc_info else None
    
    # Use earliest reliable
    candidates = [d for d in [first_meeting_sf, first_meeting_audit, first_meeting_doc] if d and pd.notna(d)]
    if candidates:
        record['FirstMeetingDate'] = min(candidates).strftime('%m/%d/%Y')
        if first_meeting_sf and pd.notna(first_meeting_sf) and min(candidates) == first_meeting_sf:
            notes_parts.append("1st Mtg: SF")
        elif first_meeting_audit and pd.notna(first_meeting_audit) and min(candidates) == first_meeting_audit:
            notes_parts.append("1st Mtg: Audit")
        else:
            notes_parts.append("1st Mtg: Doc")
    else:
        record['FirstMeetingDate'] = ''
    
    # === Total Meetings ===
    audit_count = len(acct_meetings)
    doc_count = len(doc_info.get('meetings', [])) if doc_info else 0
    record['TotalMeetings'] = max(audit_count, doc_count, 1)
    
    # === Classify meetings ===
    first_mtg = pd.to_datetime(record['FirstMeetingDate']) if record['FirstMeetingDate'] else None
    
    meeting_classifications = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        substep, reason = classify_meeting_subject(row['Subject'], idx + 1)
        meeting_classifications[substep].append({
            'date': row['Date'],
            'date_str': row['Date'].strftime('%m/%d/%y'),
            'reason': reason,
        })
    
    # Also add from doc data
    for m in doc_info.get('meetings', []) if doc_info else []:
        if 'date' not in m:
            continue
        date = m['date']
        if m.get('is_demo'):
            meeting_classifications['Demo'].append({'date': date, 'date_str': date.strftime('%m/%d/%y'), 'reason': f"Doc: {m.get('type', 'Demo')}"})
        if m.get('is_cab'):
            meeting_classifications['CAB'].append({'date': date, 'date_str': date.strftime('%m/%d/%y'), 'reason': f"Doc: {m.get('type', 'CAB')}"})
        if m.get('is_scoping'):
            meeting_classifications['Scoping'].append({'date': date, 'date_str': date.strftime('%m/%d/%y'), 'reason': f"Doc: {m.get('type', 'Scoping')}"})
    
    # === Days First to Second ===
    if len(acct_meetings) >= 2:
        days = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
        record['DaysFirstToSecond'] = days
    else:
        record['DaysFirstToSecond'] = ''
    
    # === Demo metrics ===
    demo_meetings = meeting_classifications.get('Demo', [])
    if demo_meetings and first_mtg:
        # Dedupe by date
        unique_demos = {}
        for m in demo_meetings:
            d = m['date'].strftime('%Y-%m-%d')
            if d not in unique_demos:
                unique_demos[d] = m
        demo_meetings = list(unique_demos.values())
        
        first_demo = min(m['date'] for m in demo_meetings)
        record['DaysToDemo'] = (first_demo - first_mtg).days
        record['DemoCount'] = len(demo_meetings)
        demo_notes = [f"{m['date_str']}: {m['reason'][:30]}" for m in demo_meetings[:2]]
    else:
        record['DaysToDemo'] = ''
        record['DemoCount'] = 0
    
    # === CAB metrics ===
    cab_meetings = meeting_classifications.get('CAB', [])
    if cab_meetings and first_mtg:
        unique_cabs = {}
        for m in cab_meetings:
            d = m['date'].strftime('%Y-%m-%d')
            if d not in unique_cabs:
                unique_cabs[d] = m
        cab_meetings = list(unique_cabs.values())
        
        first_cab = min(m['date'] for m in cab_meetings)
        record['DaysToCAB'] = (first_cab - first_mtg).days
        record['CABCount'] = len(cab_meetings)
        cab_notes = [f"{m['date_str']}: {m['reason'][:30]}" for m in cab_meetings[:2]]
    else:
        record['DaysToCAB'] = ''
        record['CABCount'] = 0
    
    # === Scoping metrics ===
    scoping_meetings = meeting_classifications.get('Scoping', [])
    if scoping_meetings and first_mtg:
        first_scoping = min(m['date'] for m in scoping_meetings)
        record['DaysToScoping'] = (first_scoping - first_mtg).days
        record['ScopingCount'] = len(scoping_meetings)
        scoping_notes = [f"{m['date_str']}: {m['reason'][:30]}" for m in scoping_meetings[:2]]
    else:
        record['DaysToScoping'] = ''
        record['ScopingCount'] = 0
    
    # === Compliance metrics ===
    compliance_meetings = meeting_classifications.get('Compliance', [])
    if compliance_meetings and first_mtg:
        first_compliance = min(m['date'] for m in compliance_meetings)
        record['DaysToCompliance'] = (first_compliance - first_mtg).days
        record['ComplianceCount'] = len(compliance_meetings)
        compliance_notes = [f"{m['date_str']}" for m in compliance_meetings[:2]]
    else:
        record['DaysToCompliance'] = ''
        record['ComplianceCount'] = 0
    
    # === Contracting metrics ===
    contracting_meetings = meeting_classifications.get('Contracting', [])
    if contracting_meetings and first_mtg:
        first_contracting = min(m['date'] for m in contracting_meetings)
        record['DaysToContracting'] = (first_contracting - first_mtg).days
        record['ContractingCount'] = len(contracting_meetings)
        contracting_notes = [f"{m['date_str']}" for m in contracting_meetings[:2]]
    else:
        record['DaysToContracting'] = ''
        record['ContractingCount'] = 0
    
    # === Data Source ===
    sources = []
    if sf_data:
        sources.append('SF')
    if doc_info:
        sources.append('Doc')
    if len(acct_meetings) > 0:
        sources.append('Audit')
    if opp_data:
        sources.append('Opp')
    record['DataSource'] = '+'.join(sources)
    
    # === Confidence ===
    if has_no_close(orig_name):
        record['Confidence'] = 'Medium' if sf_data else 'Standard'
    elif sf_data and sf_data.get('first_close'):
        record['Confidence'] = 'High'
    elif sf_data:
        record['Confidence'] = 'Medium'
    elif doc_info and doc_info.get('has_rich_history'):
        record['Confidence'] = 'Medium'
    else:
        record['Confidence'] = 'Standard'
    
    # === Close Date ===
    if has_no_close(orig_name):
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
        notes_parts.append("NOT CLOSED")
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            record['SalesCycleDays'] = int(sf_data['days_to_close'])
            notes_parts.append(f"Cycle: SF {int(sf_data['days_to_close'])}d")
        elif first_mtg and close_date > first_mtg:
            record['SalesCycleDays'] = (close_date - first_mtg).days
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    # === Build Audit Notes ===
    all_notes = notes_parts.copy()
    if demo_notes:
        all_notes.append(f"Demo: {'; '.join(demo_notes)}")
    if cab_notes:
        all_notes.append(f"CAB: {'; '.join(cab_notes)}")
    if scoping_notes:
        all_notes.append(f"Scoping: {'; '.join(scoping_notes)}")
    if compliance_notes:
        all_notes.append(f"Compliance: {'; '.join(compliance_notes)}")
    if contracting_notes:
        all_notes.append(f"Contract: {'; '.join(contracting_notes)}")
    
    record['AuditNotes'] = ' | '.join(all_notes) if all_notes else 'Audit data only'
    
    validated_records.append(record)

validated_df = pd.DataFrame(validated_records)

# Sort by Owner (with owner first), then Account
validated_df['_has_owner'] = validated_df['Owner'].apply(lambda x: 0 if pd.notna(x) and x != '' else 1)
validated_df = validated_df.sort_values(['_has_owner', 'Owner', 'Account']).reset_index(drop=True)
validated_df = validated_df.drop(columns=['_has_owner'])

print(f"  Created {len(validated_df)} validated records")

# =============================================================================
# STEP 5: Generate Output
# =============================================================================
print("\n[STEP 5] Generating output...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws_inputs = wb.active
ws_inputs.title = 'BL_Review_Inputs'

output_cols = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 'DaysToCAB', 'CABCount',
    'DaysToScoping', 'ScopingCount', 'DaysToCompliance', 'ComplianceCount',
    'DaysToContracting', 'ContractingCount', 'DataSource', 'Confidence',
    'CloseDate', 'SalesCycleDays', 'AuditNotes'
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for col_idx, col_name in enumerate(output_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, (_, row) in enumerate(validated_df.iterrows(), 2):
    for col_idx, col_name in enumerate(output_cols, 1):
        val = row.get(col_name, '')
        cell = ws_inputs.cell(row=row_idx, column=col_idx, value=val)
        
        confidence = row.get('Confidence', '')
        if col_idx < 20:
            if confidence == 'High':
                cell.fill = high_fill
            elif confidence == 'Medium':
                cell.fill = medium_fill

ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 25
ws_inputs.column_dimensions['V'].width = 80

output_path = '/Users/keiganpesenti/Desktop/validated_inputs_v4.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total accounts: {len(validated_df)}")
print(f"With Owner: {len(validated_df[validated_df['Owner'].notna() & (validated_df['Owner'] != '')])}")
print(f"With Close Date: {len(validated_df[validated_df['CloseDate'].notna() & (validated_df['CloseDate'] != '')])}")
print(f"High Confidence: {len(validated_df[validated_df['Confidence'] == 'High'])}")
print(f"Medium Confidence: {len(validated_df[validated_df['Confidence'] == 'Medium'])}")
print(f"Standard: {len(validated_df[validated_df['Confidence'] == 'Standard'])}")

# Bank of America verification
boa = validated_df[validated_df['Account'].str.contains('Bank of America', case=False, na=False)]
if len(boa) > 0:
    print(f"\nBank of America verification:")
    print(f"  Owner: {boa.iloc[0]['Owner']}")
    print(f"  CloseDate: '{boa.iloc[0]['CloseDate']}'")
    print(f"  Confidence: {boa.iloc[0]['Confidence']}")
    print(f"  Notes: {boa.iloc[0]['AuditNotes']}")

# Show sample notes
print("\n=== SAMPLE AUDIT NOTES ===")
for _, row in validated_df.head(8).iterrows():
    print(f"  {row['Account']}: {row['AuditNotes'][:70]}...")

