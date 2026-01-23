"""
Sales Sub-Steps Analysis v4
===========================
Formula-driven workbook:
- Account_Inputs tab = editable source data
- Summary tab = Excel formulas referencing Account_Inputs
- User can override values, summary updates automatically
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

print("="*80)
print("SALES SUB-STEPS ANALYSIS v4")
print("Formula-Driven Workbook with Editable Inputs")
print("="*80)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[STEP 1] Loading data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'

xl_audit = pd.ExcelFile(audit_file)
xl_secondary = pd.ExcelFile(secondary_file)

df_accts = pd.read_excel(xl_audit, sheet_name='all accts')
df_meetings_audit = pd.read_excel(xl_audit, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel(xl_secondary, sheet_name='all meetings')

# Combine and clean meetings
df_meetings_raw = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings_raw['Date'] = pd.to_datetime(df_meetings_raw['Date'], errors='coerce')
df_meetings_raw['Date_Only'] = df_meetings_raw['Date'].dt.date

# Deduplicate
df_meetings_raw['_key'] = (
    df_meetings_raw['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings_raw['Date_Only'].astype(str) + '|' +
    df_meetings_raw['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings_raw.drop_duplicates(subset='_key', keep='first').copy()
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Exclude test accounts
test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Meetings loaded: {len(df_meetings)}")

# =============================================================================
# NORMALIZE ACCOUNT NAMES
# =============================================================================
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

df_meetings['_norm_acct'] = df_meetings['Company / Account'].apply(normalize_name)
account_counts = df_meetings.groupby('_norm_acct').size().reset_index(name='count')
accounts_with_data = account_counts[account_counts['count'] >= 3]['_norm_acct'].tolist()

# Get validated close info
validated_info = {}
for _, row in df_accts.iterrows():
    norm = normalize_name(row['Account Name'])
    validated_info[norm] = {
        'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row['First Meeting Date']) else None,
        'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row['First Deal Closed']) else None,
    }

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
def classify_substep(subject, meeting_number):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_number == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product', 'eudia ai']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['use case', 'use-case', 'requirements']):
        return 'UseCase'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['proposal', 'delivery plan']):
        return 'Proposal'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    if any(kw in s for kw in ['pilot', 'poc', 'kickoff']):
        return 'Pilot'
    return 'Follow-up'

# =============================================================================
# BUILD ACCOUNT-LEVEL DATA
# =============================================================================
print("\n[STEP 2] Building account-level metrics...")

account_data = []
all_meetings = []

for norm_acct in accounts_with_data:
    acct_meetings = df_meetings[df_meetings['_norm_acct'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) < 3:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    first_meeting = acct_meetings['Date'].min()
    
    # Classify meetings and track dates
    meeting_dates = {}
    for idx, row in acct_meetings.iterrows():
        meeting_num = idx + 1
        substep = classify_substep(row['Subject'], meeting_num)
        
        all_meetings.append({
            'Account': orig_name,
            'Meeting_Number': meeting_num,
            'Date': row['Date'],
            'Subject': row['Subject'],
            'Classification': substep,
            'Include_Flag': 'Y',  # User can change to 'N' to exclude
        })
        
        if substep not in meeting_dates:
            meeting_dates[substep] = []
        meeting_dates[substep].append(row['Date'])
    
    # Build record
    record = {
        'Account': orig_name,
        'Include': 'Y',  # Master include flag - user can set to 'N'
        'Total_Meetings': len(acct_meetings),
        'First_Meeting_Date': first_meeting.strftime('%m/%d/%Y') if pd.notna(first_meeting) else '',
    }
    
    # Close date
    if norm_acct in validated_info and validated_info[norm_acct]['first_close']:
        close_date = validated_info[norm_acct]['first_close']
        record['Close_Date'] = close_date.strftime('%m/%d/%Y')
        record['Sales_Cycle_Days'] = (close_date - first_meeting).days
    else:
        record['Close_Date'] = ''
        record['Sales_Cycle_Days'] = ''
    
    # Days to each substep
    all_substeps = ['Intro', 'Follow-up', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Proposal', 'Compliance', 'Contracting', 'Pilot']
    for substep in all_substeps:
        if substep in meeting_dates and len(meeting_dates[substep]) > 0:
            first_date = min(meeting_dates[substep])
            record[f'Days_to_{substep}'] = (first_date - first_meeting).days
            record[f'{substep}_Count'] = len(meeting_dates[substep])
        else:
            record[f'Days_to_{substep}'] = ''
            record[f'{substep}_Count'] = 0
    
    # First to second meeting
    if len(acct_meetings) >= 2:
        record['Days_First_to_Second'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['Days_First_to_Second'] = ''
    
    account_data.append(record)

accounts_df = pd.DataFrame(account_data)
meetings_df = pd.DataFrame(all_meetings)

print(f"  Total accounts: {len(accounts_df)}")
print(f"  Total meetings: {len(meetings_df)}")

# =============================================================================
# CREATE EXCEL WORKBOOK WITH FORMULAS
# =============================================================================
print("\n[STEP 3] Creating formula-driven workbook...")

wb = Workbook()

# === TAB 1: Account_Inputs (editable source data) ===
ws_inputs = wb.active
ws_inputs.title = '1_Account_Inputs'

# Column order for Account_Inputs
input_cols = ['Account', 'Include', 'Total_Meetings', 'First_Meeting_Date', 'Close_Date', 'Sales_Cycle_Days',
              'Days_First_to_Second', 'Days_to_Demo', 'Demo_Count', 'Days_to_CAB', 'CAB_Count',
              'Days_to_UseCase', 'UseCase_Count', 'Days_to_Scoping', 'Scoping_Count',
              'Days_to_Proposal', 'Proposal_Count', 'Days_to_Compliance', 'Compliance_Count',
              'Days_to_Contracting', 'Contracting_Count', 'Days_to_Intro', 'Intro_Count',
              'Days_to_Follow-up', 'Follow-up_Count', 'Days_to_Pilot', 'Pilot_Count']

# Filter to columns that exist
input_cols = [c for c in input_cols if c in accounts_df.columns]

# Write headers
for col_idx, col_name in enumerate(input_cols, 1):
    ws_inputs.cell(row=1, column=col_idx, value=col_name)

# Write data
for row_idx, (_, row) in enumerate(accounts_df.iterrows(), 2):
    for col_idx, col_name in enumerate(input_cols, 1):
        val = row.get(col_name, '')
        ws_inputs.cell(row=row_idx, column=col_idx, value=val)

num_accounts = len(accounts_df)
data_end_row = num_accounts + 1

# Get column letters for formulas
col_letters = {col: get_column_letter(idx+1) for idx, col in enumerate(input_cols)}

# === TAB 2: Summary (formula-driven) ===
ws_summary = wb.create_sheet('2_Summary')

# Build summary structure with formulas
summary_structure = [
    # Row 1: Headers
    ['Category', 'Metric', 'Average', 'Sample_n', 'Median', 'Min', 'Max', 'Notes'],
    ['', '', '', '', '', '', '', ''],
    
    # Data Overview
    ['DATA OVERVIEW', '', '', '', '', '', '', ''],
    ['', 'Total accounts', f'=COUNTIF(\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row},"Y")', '', '', '', '', 'Accounts with Include=Y'],
    ['', 'Total with close date', f'=COUNTIF(\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},">0")', '', '', '', '', 'Closed deals'],
    ['', '', '', '', '', '', '', ''],
    
    # Sales Cycle
    ['SALES CYCLE (days)', '', '', '', '', '', '', ''],
    ['', 'First Meeting → Close', 
     f'=AVERAGEIF(\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row},"Y",\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row})',
     f'=COUNTIFS(\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row},"Y",\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},">0")',
     f'=MEDIAN(IF(\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row}="Y",IF(\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row}>0,\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row})))',
     f'=MINIFS(\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row},"Y",\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},">0")',
     f'=MAXIFS(\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},\'1_Account_Inputs\'!{col_letters["Include"]}2:{col_letters["Include"]}{data_end_row},"Y",\'1_Account_Inputs\'!{col_letters["Sales_Cycle_Days"]}2:{col_letters["Sales_Cycle_Days"]}{data_end_row},">0")',
     'Use median for forecasting'],
    ['', '', '', '', '', '', '', ''],
    
    # Time to Milestones
    ['TIME TO MILESTONE (days from first meeting)', '', '', '', '', '', '', ''],
]

# Add milestone rows with formulas
milestones = [
    ('Days_First_to_Second', 'First → Second Meeting', 'Engagement velocity'),
    ('Days_to_Demo', 'First → Demo', 'Key qualification indicator'),
    ('Days_to_CAB', 'First → CAB Discussion', 'Champion identification'),
    ('Days_to_UseCase', 'First → Use Case ID', 'Requirements gathering'),
    ('Days_to_Scoping', 'First → Scoping/Pricing', 'Moving toward proposal'),
    ('Days_to_Proposal', 'First → Proposal', 'Deal progression'),
    ('Days_to_Contracting', 'First → Contracting', 'Near close'),
    ('Days_to_Compliance', 'First → Compliance', 'Security review'),
]

for col_name, label, note in milestones:
    if col_name in col_letters:
        col_let = col_letters[col_name]
        include_let = col_letters['Include']
        summary_structure.append([
            '', label,
            f'=AVERAGEIFS(\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},">=0")',
            f'=COUNTIFS(\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},">=0")',
            f'=MEDIAN(IF(\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row}="Y",IF(\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row}>=0,\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row})))',
            f'=MINIFS(\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},">=0")',
            f'=MAXIFS(\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},">=0")',
            note
        ])

summary_structure.append(['', '', '', '', '', '', '', ''])
summary_structure.append(['MEETING COUNTS (per account)', '', '', '', '', '', '', ''])

# Add meeting count rows
meeting_counts = [
    ('Demo_Count', 'Demo meetings'),
    ('CAB_Count', 'CAB discussions'),
    ('UseCase_Count', 'Use case meetings'),
    ('Scoping_Count', 'Scoping meetings'),
    ('Proposal_Count', 'Proposal discussions'),
    ('Compliance_Count', 'Compliance meetings'),
    ('Contracting_Count', 'Contracting meetings'),
]

for col_name, label in meeting_counts:
    if col_name in col_letters:
        col_let = col_letters[col_name]
        include_let = col_letters['Include']
        summary_structure.append([
            '', label,
            f'=AVERAGEIF(\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row})',
            f'=COUNTIFS(\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},">0")',
            '',
            '',
            f'=MAXIF(\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row},\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y")',
            f'Total: =SUMIF(\'1_Account_Inputs\'!{include_let}2:{include_let}{data_end_row},"Y",\'1_Account_Inputs\'!{col_let}2:{col_let}{data_end_row})'
        ])

# Write summary
for row_idx, row_data in enumerate(summary_structure, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws_summary.cell(row=row_idx, column=col_idx, value=value)

# === TAB 3: Meeting_Source (for tracing back) ===
ws_meetings = wb.create_sheet('3_Meeting_Source')

meeting_cols = ['Account', 'Meeting_Number', 'Date', 'Subject', 'Classification', 'Include_Flag']
for col_idx, col_name in enumerate(meeting_cols, 1):
    ws_meetings.cell(row=1, column=col_idx, value=col_name)

for row_idx, (_, row) in enumerate(meetings_df.iterrows(), 2):
    for col_idx, col_name in enumerate(meeting_cols, 1):
        val = row.get(col_name, '')
        if col_name == 'Date' and pd.notna(val):
            val = val.strftime('%m/%d/%Y')
        ws_meetings.cell(row=row_idx, column=col_idx, value=val)

# === TAB 4: Instructions ===
ws_instructions = wb.create_sheet('4_Instructions')
instructions = [
    ['HOW TO USE THIS WORKBOOK'],
    [''],
    ['This workbook is FORMULA-DRIVEN. Changes to Tab 1 automatically update Tab 2.'],
    [''],
    ['TO EXCLUDE AN ACCOUNT FROM CALCULATIONS:'],
    ['1. Go to Tab 1 (Account_Inputs)'],
    ['2. Find the account row'],
    ['3. Change the "Include" column from "Y" to "N"'],
    ['4. Go to Tab 2 (Summary) - values will update automatically'],
    [''],
    ['TO OVERRIDE A SPECIFIC VALUE:'],
    ['1. Go to Tab 1 (Account_Inputs)'],
    ['2. Find the account and column (e.g., Days_to_Demo)'],
    ['3. Change the value'],
    ['4. Tab 2 will recalculate averages/medians automatically'],
    [''],
    ['TO TRACE A MEETING CLASSIFICATION:'],
    ['1. Go to Tab 3 (Meeting_Source)'],
    ['2. Filter by Account'],
    ['3. See each meeting, its date, subject, and classification'],
    ['4. If classification is wrong, note it and adjust Tab 1 values accordingly'],
    [''],
    ['CLASSIFICATION RULES:'],
    ['- Intro: First meeting with account, OR subject contains "intro"'],
    ['- Demo: Subject contains "demo", "sigma", "cortex", "platform", "walkthrough", "product"'],
    ['- CAB: Subject contains "cab", "customer advisory", "advisory board"'],
    ['- UseCase: Subject contains "use case", "requirements"'],
    ['- Scoping: Subject contains "scoping", "scope", "pricing"'],
    ['- Proposal: Subject contains "proposal", "delivery plan"'],
    ['- Compliance: Subject contains "infosec", "security", "compliance"'],
    ['- Contracting: Subject contains "contract", "redline", "msa", "negotiation"'],
    ['- Follow-up: Default for meetings 2+ without specific keywords'],
    [''],
    ['NOTE: Some MEDIAN formulas use array syntax and may need Ctrl+Shift+Enter in older Excel.'],
]

for row_idx, row_data in enumerate(instructions, 1):
    ws_instructions.cell(row=row_idx, column=1, value=row_data[0] if row_data else '')

# Save
output_path = '/Users/keiganpesenti/Desktop/sales_substeps_v4.xlsx'
wb.save(output_path)

print(f"\n  Workbook saved: {output_path}")
print("\n  Tabs:")
print("    1_Account_Inputs - EDITABLE source data (change Include to N to exclude)")
print("    2_Summary - FORMULA-DRIVEN summary (auto-updates from Tab 1)")
print("    3_Meeting_Source - Raw meeting classifications for tracing")
print("    4_Instructions - How to use and override")

print("\n" + "="*80)
print("KEY: To exclude an account, change Include='Y' to Include='N' in Tab 1")
print("Summary will automatically recalculate!")
print("="*80)

