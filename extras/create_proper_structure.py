"""
Proper Structure - Fixed
========================
- BL_Review: Simple fields for sales to verify (dates formatted properly)
- Source_Meetings: Raw meeting data with classifications
- Inputs_formula: BLENDED - pulls from BOTH BL_Review AND Source_Meetings
- Dates stored as Excel serial numbers with proper formatting
- SalesCycleDays formula handles empty CloseDate correctly
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
from collections import defaultdict

print("="*80)
print("CREATING PROPER BLENDED STRUCTURE")
print("="*80)

# =============================================================================
# CONFIGURATION
# =============================================================================
EXCLUDE_ACCOUNTS = [
    'army applications lab', 'bortstein legal group', 'borstein legal group',
    'box.com', 'box', 'dla piper', 'dte energy', 'defense commissary agency',
    'dentsu', 'docusign', 'dow jones', 'ec coorporid', 'ec corpid', 'ec corporate id',
    'floodgate', 'gainsight', 'general catalyst', 'gov dod', 'green oaks capital',
    'innovative driven', 'insight enterprise', 'jb hi-fi', 'john hopkins medicine',
    'johns hopkins', 'jewel labs', 'msc industrial', 'state of alaska',
    'cambridge university press', 'guardian life', 'mckinsey', 'sonos'
]

NO_CLOSE_DATE = ['bank of america']

def should_exclude(name):
    name_lower = str(name).lower().strip()
    for exc in EXCLUDE_ACCOUNTS:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def has_no_close(name):
    name_lower = str(name).lower().strip()
    for exc in NO_CLOSE_DATE:
        if exc in name_lower or name_lower in exc:
            return True
    return False

def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[1] Loading data...")

sf_ref = pd.read_excel('/Users/keiganpesenti/Desktop/days to close reference.xlsx')
sf_lookup = {}
for _, row in sf_ref.iterrows():
    if pd.notna(row['Account Name']):
        name = str(row['Account Name']).lower().strip()
        sf_lookup[name] = {
            'owner': row.get('Account Owner', ''),
            'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row.get('First Meeting Date')) else None,
            'days_to_close': row.get('Days to Close', ''),
            'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row.get('First Deal Closed')) else None,
        }

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

df_meetings = df_meetings[~df_meetings['_acct_lower'].apply(should_exclude)]

print(f"  Loaded {len(df_meetings)} meetings")

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
print("\n[2] Processing accounts...")

def classify_meeting(subject, is_first_meeting=False):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if is_first_meeting:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    return 'Followup'

def fuzzy_match(name, lookup_dict):
    name_norm = normalize_name(name)
    if name_norm in lookup_dict:
        return name_norm
    name_lower = str(name).lower().strip() if name else ''
    if name_lower in lookup_dict:
        return name_lower
    for key in lookup_dict:
        key_clean = key.lower().strip()
        name_clean = name_norm.lower().strip()
        if name_clean and key_clean:
            if name_clean in key_clean or key_clean in name_clean:
                return key
    return None

account_counts = df_meetings.groupby('_acct_lower').size().reset_index(name='count')
accounts_to_process = account_counts[account_counts['count'] >= 3]['_acct_lower'].tolist()
accounts_to_process = [a for a in accounts_to_process if not should_exclude(a)]

# Build data
bl_review_data = []
source_data = []

for norm_acct in sorted(accounts_to_process):
    acct_meetings = df_meetings[df_meetings['_acct_lower'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) == 0:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    if should_exclude(orig_name):
        continue
    
    # SF match
    sf_match = fuzzy_match(orig_name, sf_lookup)
    sf_data = sf_lookup.get(sf_match, {}) if sf_match else {}
    if not sf_data:
        for sf_key, sf_val in sf_lookup.items():
            if normalize_name(sf_key) == normalize_name(orig_name):
                sf_data = sf_val
                break
    
    # First meeting date
    first_meeting_sf = sf_data.get('first_meeting') if sf_data else None
    first_meeting_audit = acct_meetings['Date'].min()
    
    if first_meeting_sf and pd.notna(first_meeting_sf):
        first_mtg = first_meeting_sf
    else:
        first_mtg = first_meeting_audit
    
    # Classify meetings and build source data
    meeting_classes = defaultdict(list)
    for idx, row in acct_meetings.iterrows():
        is_first = (row['Date'] == first_mtg) if first_mtg else (idx == 0)
        cls = classify_meeting(row['Subject'], is_first)
        days_from_first = (row['Date'] - first_mtg).days if first_mtg and pd.notna(row['Date']) else 0
        meeting_classes[cls].append(days_from_first)
        
        source_data.append({
            'Account': orig_name,
            'Date': row['Date'],
            'Subject': str(row['Subject'])[:60] if pd.notna(row['Subject']) else '',
            'Classification': cls,
            'DaysFromFirst': days_from_first,
        })
    
    # Get counts
    demo_count = len(meeting_classes.get('Demo', []))
    cab_count = len(meeting_classes.get('CAB', []))
    scoping_count = len(meeting_classes.get('Scoping', []))
    compliance_count = len(meeting_classes.get('Compliance', []))
    contracting_count = len(meeting_classes.get('Contracting', []))
    
    # Days to first occurrence
    demos = meeting_classes.get('Demo', [])
    cabs = meeting_classes.get('CAB', [])
    scopings = meeting_classes.get('Scoping', [])
    compliances = meeting_classes.get('Compliance', [])
    contractings = meeting_classes.get('Contracting', [])
    
    days_to_demo = min(demos) if demos else 0
    days_to_cab = min(cabs) if cabs else 0
    days_to_scoping = min(scopings) if scopings else 0
    days_to_compliance = min(compliances) if compliances else 0
    days_to_contracting = min(contractings) if contractings else 0
    
    # Days first to second
    if len(acct_meetings) >= 2:
        days_first_to_second = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        days_first_to_second = 0
    
    # Close date and sales cycle
    if has_no_close(orig_name):
        close_date = None
        sales_cycle = 0
    elif sf_data and sf_data.get('first_close') and pd.notna(sf_data['first_close']):
        close_date = sf_data['first_close']
        if sf_data.get('days_to_close') and pd.notna(sf_data['days_to_close']):
            sales_cycle = int(sf_data['days_to_close'])
        elif first_mtg:
            sales_cycle = (close_date - first_mtg).days
        else:
            sales_cycle = 0
    else:
        close_date = None
        sales_cycle = 0
    
    bl_review_data.append({
        'Owner': sf_data.get('owner', '') if sf_data else '',
        'Account': orig_name,
        'Include': 'Y',
        'Verified': '',
        'FirstMeetingDate': first_mtg,
        'CloseDate': close_date,
        'TotalMeetings': len(acct_meetings),
        'DemoCount': demo_count,
        'HadCAB': 'Y' if cab_count > 0 else '',
        'Notes': '',
        # Pre-calculated for Inputs_formula (static values)
        '_SalesCycleDays': sales_cycle if close_date else 0,
        '_DaysFirstToSecond': days_first_to_second,
        '_DaysToDemo': days_to_demo,
        '_DaysToCAB': days_to_cab,
        '_CABCount': cab_count,
        '_DaysToScoping': days_to_scoping,
        '_ScopingCount': scoping_count,
        '_DaysToCompliance': days_to_compliance,
        '_ComplianceCount': compliance_count,
        '_DaysToContracting': days_to_contracting,
        '_ContractingCount': contracting_count,
    })

# Sort by Owner, then Account
bl_review_data.sort(key=lambda x: (0 if x['Owner'] else 1, x['Owner'] or '', x['Account']))

print(f"  Processed {len(bl_review_data)} accounts")
print(f"  Source meetings: {len(source_data)}")

# =============================================================================
# GENERATE EXCEL
# =============================================================================
print("\n[3] Generating Excel workbook...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

# =============================================================================
# TAB 1: Source_Meetings (FIRST - for COUNTIFS references)
# =============================================================================
ws_source = wb.active
ws_source.title = 'Source_Meetings'

source_cols = ['Account', 'Date', 'Subject', 'Classification', 'DaysFromFirst']
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col_idx, col_name in enumerate(source_cols, 1):
    cell = ws_source.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, row_data in enumerate(source_data, 2):
    ws_source.cell(row=row_idx, column=1, value=row_data['Account'])
    cell_date = ws_source.cell(row=row_idx, column=2)
    if row_data['Date']:
        cell_date.value = row_data['Date']
        cell_date.number_format = 'MM/DD/YYYY'
    ws_source.cell(row=row_idx, column=3, value=row_data['Subject'])
    ws_source.cell(row=row_idx, column=4, value=row_data['Classification'])
    ws_source.cell(row=row_idx, column=5, value=row_data['DaysFromFirst'])

ws_source.column_dimensions['A'].width = 25
ws_source.column_dimensions['B'].width = 12
ws_source.column_dimensions['C'].width = 50

# =============================================================================
# TAB 2: BL_Review (Simple for sales)
# =============================================================================
ws_bl = wb.create_sheet('BL_Review')

verify_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

bl_columns = ['Owner', 'Account', 'Include', 'Verified', 'FirstMeetingDate', 'CloseDate', 
              'TotalMeetings', 'DemoCount', 'HadCAB', 'Notes']

for col_idx, col_name in enumerate(bl_columns, 1):
    cell = ws_bl.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

for row_idx, record in enumerate(bl_review_data, 2):
    ws_bl.cell(row=row_idx, column=1, value=record['Owner'])
    ws_bl.cell(row=row_idx, column=2, value=record['Account'])
    ws_bl.cell(row=row_idx, column=3, value=record['Include'])
    ws_bl.cell(row=row_idx, column=4, value=record['Verified'])
    
    # FirstMeetingDate - properly formatted
    cell_e = ws_bl.cell(row=row_idx, column=5)
    if record['FirstMeetingDate']:
        cell_e.value = record['FirstMeetingDate']
        cell_e.number_format = 'MM/DD/YYYY'
    cell_e.fill = verify_fill
    
    # CloseDate - properly formatted
    cell_f = ws_bl.cell(row=row_idx, column=6)
    if record['CloseDate']:
        cell_f.value = record['CloseDate']
        cell_f.number_format = 'MM/DD/YYYY'
    cell_f.fill = verify_fill
    
    cell_g = ws_bl.cell(row=row_idx, column=7, value=record['TotalMeetings'])
    cell_g.fill = verify_fill
    
    cell_h = ws_bl.cell(row=row_idx, column=8, value=record['DemoCount'])
    cell_h.fill = verify_fill
    
    cell_i = ws_bl.cell(row=row_idx, column=9, value=record['HadCAB'])
    cell_i.fill = verify_fill
    
    ws_bl.cell(row=row_idx, column=10, value=record['Notes'])

ws_bl.column_dimensions['A'].width = 18
ws_bl.column_dimensions['B'].width = 28
ws_bl.column_dimensions['E'].width = 14
ws_bl.column_dimensions['F'].width = 12

# =============================================================================
# TAB 3: Inputs_formula (BLENDED - from BL_Review + Source_Meetings)
# =============================================================================
ws_inputs = wb.create_sheet('Inputs_formula')

input_columns = [
    'Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate',
    'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount',
    'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance',
    'ComplianceCount', 'DaysToContract', 'ContractingCount', 'Notes'
]

formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
source_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

for col_idx, col_name in enumerate(input_columns, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

num_rows = len(bl_review_data)

for row_idx, record in enumerate(bl_review_data, 2):
    account_name = record['Account']
    
    # A: Owner (from BL_Review)
    ws_inputs.cell(row=row_idx, column=1, value=f"=BL_Review!A{row_idx}")
    
    # B: Account (from BL_Review)
    ws_inputs.cell(row=row_idx, column=2, value=f"=BL_Review!B{row_idx}")
    
    # C: Include (from BL_Review)
    ws_inputs.cell(row=row_idx, column=3, value=f"=BL_Review!C{row_idx}")
    
    # D: Verified (from BL_Review)
    ws_inputs.cell(row=row_idx, column=4, value=f"=BL_Review!D{row_idx}")
    
    # E: TotalMeetings (from Source_Meetings via COUNTIF)
    cell_e = ws_inputs.cell(row=row_idx, column=5, 
                            value=f'=COUNTIF(Source_Meetings!A:A,B{row_idx})')
    cell_e.fill = source_fill
    
    # F: FirstMeetingDate (from BL_Review)
    ws_inputs.cell(row=row_idx, column=6, value=f"=BL_Review!E{row_idx}")
    
    # G: CloseDate (from BL_Review)
    ws_inputs.cell(row=row_idx, column=7, value=f"=BL_Review!F{row_idx}")
    
    # H: SalesCycleDays (CALCULATED - only if CloseDate exists and is > FirstMeetingDate)
    cell_h = ws_inputs.cell(row=row_idx, column=8, 
                            value=f'=IF(AND(G{row_idx}<>"",G{row_idx}>0,F{row_idx}<>"",F{row_idx}>0),G{row_idx}-F{row_idx},0)')
    cell_h.fill = formula_fill
    
    # I: DaysFirstToSecond (static - pre-calculated)
    ws_inputs.cell(row=row_idx, column=9, value=record['_DaysFirstToSecond'])
    
    # J: DaysToDemo (from Source_Meetings via MINIFS)
    cell_j = ws_inputs.cell(row=row_idx, column=10, 
                            value=f'=IFERROR(MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Demo",Source_Meetings!E:E,">0"),0)')
    cell_j.fill = source_fill
    
    # K: DemoCount (from Source_Meetings via COUNTIFS)
    cell_k = ws_inputs.cell(row=row_idx, column=11, 
                            value=f'=COUNTIFS(Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Demo")')
    cell_k.fill = source_fill
    
    # L: DaysToCAB (from Source_Meetings)
    cell_l = ws_inputs.cell(row=row_idx, column=12, 
                            value=f'=IFERROR(MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"CAB",Source_Meetings!E:E,">0"),0)')
    cell_l.fill = source_fill
    
    # M: CABCount (from Source_Meetings)
    cell_m = ws_inputs.cell(row=row_idx, column=13, 
                            value=f'=COUNTIFS(Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"CAB")')
    cell_m.fill = source_fill
    
    # N: DaysToScoping (from Source_Meetings)
    cell_n = ws_inputs.cell(row=row_idx, column=14, 
                            value=f'=IFERROR(MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Scoping",Source_Meetings!E:E,">0"),0)')
    cell_n.fill = source_fill
    
    # O: ScopingCount (from Source_Meetings)
    cell_o = ws_inputs.cell(row=row_idx, column=15, 
                            value=f'=COUNTIFS(Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Scoping")')
    cell_o.fill = source_fill
    
    # P: DaysToCompliance (from Source_Meetings)
    cell_p = ws_inputs.cell(row=row_idx, column=16, 
                            value=f'=IFERROR(MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Compliance",Source_Meetings!E:E,">0"),0)')
    cell_p.fill = source_fill
    
    # Q: ComplianceCount (from Source_Meetings)
    cell_q = ws_inputs.cell(row=row_idx, column=17, 
                            value=f'=COUNTIFS(Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Compliance")')
    cell_q.fill = source_fill
    
    # R: DaysToContract (from Source_Meetings)
    cell_r = ws_inputs.cell(row=row_idx, column=18, 
                            value=f'=IFERROR(MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Contracting",Source_Meetings!E:E,">0"),0)')
    cell_r.fill = source_fill
    
    # S: ContractingCount (from Source_Meetings)
    cell_s = ws_inputs.cell(row=row_idx, column=19, 
                            value=f'=COUNTIFS(Source_Meetings!A:A,B{row_idx},Source_Meetings!D:D,"Contracting")')
    cell_s.fill = source_fill
    
    # T: Notes (from BL_Review)
    ws_inputs.cell(row=row_idx, column=20, value=f"=BL_Review!J{row_idx}")

ws_inputs.column_dimensions['A'].width = 18
ws_inputs.column_dimensions['B'].width = 28

# =============================================================================
# TAB 4: Instructions
# =============================================================================
ws_inst = wb.create_sheet('Instructions')

instructions = [
    ['BLENDED INPUTS STRUCTURE', ''],
    ['', ''],
    ['DATA FLOW:', ''],
    ['Source_Meetings', 'Raw classified meetings - formulas count from here'],
    ['BL_Review', 'Sales verifies key fields (yellow cells)'],
    ['Inputs_formula', 'BLENDED - pulls from both BL_Review AND Source_Meetings'],
    ['Summary', 'Your formulas pull from Inputs_formula'],
    ['', ''],
    ['INPUTS_FORMULA COLUMN SOURCES:', ''],
    ['Column', 'Source', 'Color'],
    ['Owner, Account, Include, Verified', 'BL_Review', 'White'],
    ['TotalMeetings', 'Source_Meetings (COUNTIF)', 'Blue'],
    ['FirstMeetingDate, CloseDate', 'BL_Review', 'White'],
    ['SalesCycleDays', 'CALCULATED (CloseDate-FirstMeetingDate)', 'Green'],
    ['DaysToDemo, DemoCount', 'Source_Meetings (MINIFS/COUNTIFS)', 'Blue'],
    ['DaysToCAB, CABCount', 'Source_Meetings (MINIFS/COUNTIFS)', 'Blue'],
    ['etc.', '', ''],
    ['', ''],
    ['COLOR LEGEND:', ''],
    ['Yellow (BL_Review)', 'Fields for sales to verify', ''],
    ['Blue (Inputs_formula)', 'Pulled from Source_Meetings via formulas', ''],
    ['Green (Inputs_formula)', 'Calculated from other fields', ''],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_inst.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)

ws_inst.column_dimensions['A'].width = 35
ws_inst.column_dimensions['B'].width = 45

# Save
output_path = '/Users/keiganpesenti/Desktop/PROPER_BLENDED.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("STRUCTURE")
print("="*80)
print("""
Source_Meetings → Raw meetings with classifications
                  (TotalMeetings, DemoCount, CABCount, etc. come from here via COUNTIFS)

BL_Review → Simple fields for sales to verify
            - FirstMeetingDate (formatted as MM/DD/YYYY)
            - CloseDate (formatted as MM/DD/YYYY)
            - TotalMeetings, DemoCount, HadCAB

Inputs_formula → BLENDED from both sources:
  - Owner, Account, Include, Verified, FirstMeetingDate, CloseDate, Notes → from BL_Review
  - TotalMeetings → =COUNTIF(Source_Meetings!A:A,B2)
  - DemoCount → =COUNTIFS(Source_Meetings!A:A,B2,Source_Meetings!D:D,"Demo")
  - DaysToDemo → =MINIFS(Source_Meetings!E:E,Source_Meetings!A:A,B2,Source_Meetings!D:D,"Demo")
  - SalesCycleDays → =IF(G2>0 AND F2>0,G2-F2,0) (handles empty/zero CloseDate)

Summary → Your existing formulas pull from Inputs_formula
""")

