"""
Sales Sub-Steps Analysis v5
===========================
Formula-driven with COMPATIBLE Excel formulas:
- No array formulas (no Ctrl+Shift+Enter needed)
- Simple tab names for easy copying
- Uses standard Excel functions only
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

print("="*80)
print("SALES SUB-STEPS ANALYSIS v5")
print("Compatible Excel Formulas")
print("="*80)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[STEP 1] Loading data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'

xl_audit = pd.ExcelFile(audit_file)
xl_secondary = pd.ExcelFile(secondary_file)

df_accts = pd.read_excel(xl_audit, sheet_name='all accts')
df_meetings_audit = pd.read_excel(xl_audit, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel(xl_secondary, sheet_name='all meetings')

# Combine and clean meetings
df_meetings_raw = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings_raw['Date'] = pd.to_datetime(df_meetings_raw['Date'], errors='coerce')
df_meetings_raw['Date_Only'] = df_meetings_raw['Date'].dt.date

# Deduplicate
df_meetings_raw['_key'] = (
    df_meetings_raw['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings_raw['Date_Only'].astype(str) + '|' +
    df_meetings_raw['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings_raw.drop_duplicates(subset='_key', keep='first').copy()
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Exclude test accounts
test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Meetings loaded: {len(df_meetings)}")

# =============================================================================
# NORMALIZE ACCOUNT NAMES
# =============================================================================
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

df_meetings['_norm_acct'] = df_meetings['Company / Account'].apply(normalize_name)
account_counts = df_meetings.groupby('_norm_acct').size().reset_index(name='count')
accounts_with_data = account_counts[account_counts['count'] >= 3]['_norm_acct'].tolist()

# Get validated close info
validated_info = {}
for _, row in df_accts.iterrows():
    norm = normalize_name(row['Account Name'])
    validated_info[norm] = {
        'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row['First Meeting Date']) else None,
        'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row['First Deal Closed']) else None,
    }

# =============================================================================
# CLASSIFY MEETINGS
# =============================================================================
def classify_substep(subject, meeting_number):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_number == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product', 'eudia ai']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['use case', 'use-case', 'requirements']):
        return 'UseCase'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['proposal', 'delivery plan']):
        return 'Proposal'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    if any(kw in s for kw in ['pilot', 'poc', 'kickoff']):
        return 'Pilot'
    return 'Followup'

# =============================================================================
# BUILD ACCOUNT-LEVEL DATA
# =============================================================================
print("\n[STEP 2] Building account-level metrics...")

account_data = []
all_meetings = []

for norm_acct in accounts_with_data:
    acct_meetings = df_meetings[df_meetings['_norm_acct'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) < 3:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    first_meeting = acct_meetings['Date'].min()
    
    # Classify meetings and track dates
    meeting_dates = {}
    for idx, row in acct_meetings.iterrows():
        meeting_num = idx + 1
        substep = classify_substep(row['Subject'], meeting_num)
        
        all_meetings.append({
            'Account': orig_name,
            'MeetingNum': meeting_num,
            'Date': row['Date'],
            'Subject': row['Subject'],
            'Classification': substep,
            'IncludeFlag': 'Y',
        })
        
        if substep not in meeting_dates:
            meeting_dates[substep] = []
        meeting_dates[substep].append(row['Date'])
    
    # Build record
    record = {
        'Account': orig_name,
        'Include': 'Y',
        'TotalMeetings': len(acct_meetings),
        'FirstMeetingDate': first_meeting.strftime('%m/%d/%Y') if pd.notna(first_meeting) else '',
    }
    
    # Close date
    if norm_acct in validated_info and validated_info[norm_acct]['first_close']:
        close_date = validated_info[norm_acct]['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        record['SalesCycleDays'] = (close_date - first_meeting).days
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    # Days to each substep
    all_substeps = ['Intro', 'Followup', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Proposal', 'Compliance', 'Contracting', 'Pilot']
    for substep in all_substeps:
        if substep in meeting_dates and len(meeting_dates[substep]) > 0:
            first_date = min(meeting_dates[substep])
            record[f'DaysTo{substep}'] = (first_date - first_meeting).days
            record[f'{substep}Count'] = len(meeting_dates[substep])
        else:
            record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = 0
    
    # First to second meeting
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    account_data.append(record)

accounts_df = pd.DataFrame(account_data)
meetings_df = pd.DataFrame(all_meetings)

print(f"  Total accounts: {len(accounts_df)}")
print(f"  Total meetings: {len(meetings_df)}")

# =============================================================================
# CREATE EXCEL WORKBOOK WITH COMPATIBLE FORMULAS
# =============================================================================
print("\n[STEP 3] Creating formula-driven workbook...")

wb = Workbook()

# === TAB 1: Inputs (simple name, no prefix) ===
ws_inputs = wb.active
ws_inputs.title = 'Inputs'

# Column order
input_cols = ['Account', 'Include', 'TotalMeetings', 'FirstMeetingDate', 'CloseDate', 'SalesCycleDays',
              'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 'DaysToCAB', 'CABCount',
              'DaysToUseCase', 'UseCaseCount', 'DaysToScoping', 'ScopingCount',
              'DaysToProposal', 'ProposalCount', 'DaysToCompliance', 'ComplianceCount',
              'DaysToContracting', 'ContractingCount', 'DaysToIntro', 'IntroCount',
              'DaysToFollowup', 'FollowupCount', 'DaysToPilot', 'PilotCount']

input_cols = [c for c in input_cols if c in accounts_df.columns]

# Write headers
for col_idx, col_name in enumerate(input_cols, 1):
    ws_inputs.cell(row=1, column=col_idx, value=col_name)

# Write data
for row_idx, (_, row) in enumerate(accounts_df.iterrows(), 2):
    for col_idx, col_name in enumerate(input_cols, 1):
        val = row.get(col_name, '')
        ws_inputs.cell(row=row_idx, column=col_idx, value=val)

num_accounts = len(accounts_df)
data_end_row = num_accounts + 1

# Get column letters
col_letters = {col: get_column_letter(idx+1) for idx, col in enumerate(input_cols)}

# === TAB 2: Summary (simple name, compatible formulas) ===
ws_summary = wb.create_sheet('Summary')

# Use only standard Excel functions (AVERAGEIF, AVERAGEIFS, COUNTIF, COUNTIFS, MINIFS, MAXIFS, SUMIF)
# NO array formulas - skip MEDIAN to avoid #NAME? errors

summary_data = [
    ['Category', 'Metric', 'Average', 'SampleN', 'Min', 'Max', 'Total', 'Notes'],
    [],
    ['DATA OVERVIEW', '', '', '', '', '', '', ''],
    ['', 'Total accounts (Include=Y)', f'=COUNTIF(Inputs!B:B,"Y")', '', '', '', '', 'Change Include to N to exclude'],
    ['', 'Accounts with close date', f'=COUNTIFS(Inputs!B:B,"Y",Inputs!F:F,">0")', '', '', '', '', 'Closed deals only'],
    [],
    ['SALES CYCLE (days)', '', '', '', '', '', '', ''],
    ['', 'First Meeting to Close',
     f'=AVERAGEIFS(Inputs!F:F,Inputs!B:B,"Y",Inputs!F:F,">0")',
     f'=COUNTIFS(Inputs!B:B,"Y",Inputs!F:F,">0")',
     f'=MINIFS(Inputs!F:F,Inputs!B:B,"Y",Inputs!F:F,">0")',
     f'=MAXIFS(Inputs!F:F,Inputs!B:B,"Y",Inputs!F:F,">0")',
     '',
     'Sales cycle for closed deals'],
    [],
    ['TIME TO MILESTONE (days)', '', '', '', '', '', '', ''],
]

# Milestone formulas - using column letters from the Inputs tab
milestones = [
    ('DaysFirstToSecond', 'G', 'First to Second Meeting', 'Engagement velocity'),
    ('DaysToDemo', 'H', 'First to Demo', 'Key qualification step'),
    ('DaysToCAB', 'J', 'First to CAB Discussion', 'Champion identification'),
    ('DaysToUseCase', 'L', 'First to Use Case', 'Requirements'),
    ('DaysToScoping', 'N', 'First to Scoping', 'Pricing discussions'),
    ('DaysToProposal', 'P', 'First to Proposal', 'Deal progression'),
    ('DaysToContracting', 'T', 'First to Contracting', 'Near close'),
    ('DaysToCompliance', 'R', 'First to Compliance', 'Security review'),
]

for col_name, col_let, label, note in milestones:
    summary_data.append([
        '', label,
        f'=AVERAGEIFS(Inputs!{col_let}:{col_let},Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">=0")',
        f'=COUNTIFS(Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">=0")',
        f'=MINIFS(Inputs!{col_let}:{col_let},Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">=0")',
        f'=MAXIFS(Inputs!{col_let}:{col_let},Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">=0")',
        '',
        note
    ])

summary_data.append([])
summary_data.append(['MEETING COUNTS', '', '', '', '', '', '', ''])

# Meeting count formulas
counts = [
    ('DemoCount', 'I', 'Demo meetings'),
    ('CABCount', 'K', 'CAB discussions'),
    ('UseCaseCount', 'M', 'Use case meetings'),
    ('ScopingCount', 'O', 'Scoping meetings'),
    ('ProposalCount', 'Q', 'Proposal discussions'),
    ('ComplianceCount', 'S', 'Compliance meetings'),
    ('ContractingCount', 'U', 'Contracting meetings'),
]

for col_name, col_let, label in counts:
    summary_data.append([
        '', label,
        f'=AVERAGEIF(Inputs!B:B,"Y",Inputs!{col_let}:{col_let})',
        f'=COUNTIFS(Inputs!B:B,"Y",Inputs!{col_let}:{col_let},">0")',
        '',
        f'=MAXIFS(Inputs!{col_let}:{col_let},Inputs!B:B,"Y")',
        f'=SUMIF(Inputs!B:B,"Y",Inputs!{col_let}:{col_let})',
        ''
    ])

summary_data.append([])
summary_data.append(['NOTES', '', '', '', '', '', '', ''])
summary_data.append(['', 'To exclude account: change Include from Y to N in Inputs tab', '', '', '', '', '', ''])
summary_data.append(['', 'To override value: edit the cell directly in Inputs tab', '', '', '', '', '', ''])
summary_data.append(['', 'Formulas auto-update when Inputs change', '', '', '', '', '', ''])

# Write summary
for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        if value is not None:
            ws_summary.cell(row=row_idx, column=col_idx, value=value)

# === TAB 3: Meetings (for tracing) ===
ws_meetings = wb.create_sheet('Meetings')

meeting_cols = ['Account', 'MeetingNum', 'Date', 'Subject', 'Classification', 'IncludeFlag']
for col_idx, col_name in enumerate(meeting_cols, 1):
    ws_meetings.cell(row=1, column=col_idx, value=col_name)

for row_idx, (_, row) in enumerate(meetings_df.iterrows(), 2):
    for col_idx, col_name in enumerate(meeting_cols, 1):
        val = row.get(col_name, '')
        if col_name == 'Date' and pd.notna(val):
            val = val.strftime('%m/%d/%Y')
        ws_meetings.cell(row=row_idx, column=col_idx, value=val)

# === TAB 4: How To Use ===
ws_help = wb.create_sheet('HowToUse')
help_text = [
    ['HOW TO USE THIS WORKBOOK'],
    [''],
    ['This workbook uses Excel formulas. Summary tab auto-updates from Inputs.'],
    [''],
    ['TAB: Inputs'],
    ['- Each row = one account'],
    ['- Column B (Include) = Y or N'],
    ['- Set to N to exclude from all calculations'],
    ['- Edit any value to override'],
    [''],
    ['TAB: Summary'],
    ['- All cells contain formulas referencing Inputs'],
    ['- No manual entry needed here'],
    ['- Values update automatically'],
    [''],
    ['TAB: Meetings'],
    ['- Raw meeting data with classifications'],
    ['- Use to trace where values came from'],
    ['- Filter by Account to see that account\'s meetings'],
    [''],
    ['EXAMPLE: Exclude Cargill from averages'],
    ['1. Go to Inputs tab'],
    ['2. Find Cargill row'],
    ['3. Change column B from Y to N'],
    ['4. Summary recalculates without Cargill'],
    [''],
    ['EXAMPLE: Fix wrong Days to Demo'],
    ['1. Go to Inputs tab'],
    ['2. Find the account'],
    ['3. Change DaysToDemo column to correct value'],
    ['4. Summary average updates'],
    [''],
    ['FORMULAS USED (all standard Excel):'],
    ['- AVERAGEIF, AVERAGEIFS'],
    ['- COUNTIF, COUNTIFS'],
    ['- MINIFS, MAXIFS'],
    ['- SUMIF'],
]

for row_idx, row_data in enumerate(help_text, 1):
    ws_help.cell(row=row_idx, column=1, value=row_data[0] if row_data else '')

# Save
output_path = '/Users/keiganpesenti/Desktop/sales_substeps_v5.xlsx'
wb.save(output_path)

print(f"\n  Workbook saved: {output_path}")
print("\n  Tabs (simple names for easy copying):")
print("    Inputs  - Editable account data")
print("    Summary - Formula-driven (references Inputs)")
print("    Meetings - Raw meeting data for tracing")
print("    HowToUse - Instructions")

print("\n" + "="*80)
print("FORMULAS USE ONLY: AVERAGEIF, COUNTIF, MINIFS, MAXIFS, SUMIF")
print("NO array formulas - should work in any Excel version")
print("="*80)

