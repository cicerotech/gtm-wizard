"""
Sales Sub-Steps Analysis v8
===========================
- Account Owner pulled from source data
- Full Summary tab restored (like v6)
- SalesReview sorted by Owner
- Three-tab flow: SalesReview → Inputs → Summary
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

print("="*80)
print("SALES SUB-STEPS ANALYSIS v8")
print("With Account Owners & Full Summary")
print("="*80)

# =============================================================================
# LOAD DATA
# =============================================================================
print("\n[STEP 1] Loading data...")

audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
secondary_file = '/Users/keiganpesenti/Desktop/meetings with accounts.xlsx'

xl_audit = pd.ExcelFile(audit_file)
xl_secondary = pd.ExcelFile(secondary_file)

df_accts = pd.read_excel(xl_audit, sheet_name='all accts')
df_meetings_audit = pd.read_excel(xl_audit, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel(xl_secondary, sheet_name='all meetings')

# Build owner lookup from validated accounts
owner_lookup = {}
for _, row in df_accts.iterrows():
    if pd.notna(row.get('Account Name')) and pd.notna(row.get('Account Owner')):
        name = str(row['Account Name']).lower().strip()
        owner_lookup[name] = row['Account Owner']

print(f"  Owner data for {len(owner_lookup)} accounts")

# Combine meetings
df_meetings_raw = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings_raw['Date'] = pd.to_datetime(df_meetings_raw['Date'], errors='coerce')
df_meetings_raw['Date_Only'] = df_meetings_raw['Date'].dt.date

df_meetings_raw['_key'] = (
    df_meetings_raw['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings_raw['Date_Only'].astype(str) + '|' +
    df_meetings_raw['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings_raw.drop_duplicates(subset='_key', keep='first').copy()
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

test_patterns = ['event triage', 'johnson hana', 'jhi', 'eudia', 'test account']
df_meetings['_acct_lower'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
for pattern in test_patterns:
    df_meetings = df_meetings[~df_meetings['_acct_lower'].str.contains(pattern, na=False)]

print(f"  Meetings loaded: {len(df_meetings)}")

# =============================================================================
# PROCESS DATA
# =============================================================================
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).lower().strip()
    for suffix in [' inc', ' inc.', ' llc', ' ltd', ' limited', ' corp', ' corporation', ' plc', ' global', ' group']:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s

df_meetings['_norm_acct'] = df_meetings['Company / Account'].apply(normalize_name)
account_counts = df_meetings.groupby('_norm_acct').size().reset_index(name='count')
accounts_with_data = account_counts[account_counts['count'] >= 3]['_norm_acct'].tolist()

validated_info = {}
for _, row in df_accts.iterrows():
    norm = normalize_name(row['Account Name'])
    validated_info[norm] = {
        'first_meeting': pd.to_datetime(row['First Meeting Date']) if pd.notna(row['First Meeting Date']) else None,
        'first_close': pd.to_datetime(row['First Deal Closed']) if pd.notna(row['First Deal Closed']) else None,
        'owner': row.get('Account Owner', ''),
    }

def classify_substep(subject, meeting_number):
    if pd.isna(subject):
        subject = ''
    s = str(subject).lower().strip()
    
    if meeting_number == 1:
        return 'Intro'
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product', 'eudia ai']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['use case', 'use-case', 'requirements']):
        return 'UseCase'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing']):
        return 'Scoping'
    if any(kw in s for kw in ['proposal', 'delivery plan']):
        return 'Proposal'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation']):
        return 'Contracting'
    if any(kw in s for kw in ['pilot', 'poc', 'kickoff']):
        return 'Pilot'
    return 'Followup'

print("\n[STEP 2] Building account data...")

account_data = []
all_meetings = []

for norm_acct in accounts_with_data:
    acct_meetings = df_meetings[df_meetings['_norm_acct'] == norm_acct].copy()
    acct_meetings = acct_meetings.sort_values('Date').reset_index(drop=True)
    
    if len(acct_meetings) < 3:
        continue
    
    orig_name = acct_meetings['Company / Account'].iloc[0]
    first_meeting = acct_meetings['Date'].min()
    
    # Get owner from validated data
    owner = ''
    if norm_acct in validated_info:
        owner = validated_info[norm_acct].get('owner', '')
    if not owner:
        # Try fuzzy match
        for key in owner_lookup:
            if norm_acct in key or key in norm_acct:
                owner = owner_lookup[key]
                break
    
    meeting_dates = {}
    for idx, row in acct_meetings.iterrows():
        meeting_num = idx + 1
        substep = classify_substep(row['Subject'], meeting_num)
        
        all_meetings.append({
            'Account': orig_name,
            'Owner': owner,
            'MeetingNum': meeting_num,
            'Date': row['Date'],
            'Subject': row['Subject'],
            'Classification': substep,
        })
        
        if substep not in meeting_dates:
            meeting_dates[substep] = []
        meeting_dates[substep].append(row['Date'])
    
    record = {
        'Account': orig_name,
        'Owner': owner,
        'TotalMeetings': len(acct_meetings),
        'FirstMeetingDate': first_meeting.strftime('%m/%d/%Y') if pd.notna(first_meeting) else '',
    }
    
    if norm_acct in validated_info and validated_info[norm_acct]['first_close']:
        close_date = validated_info[norm_acct]['first_close']
        record['CloseDate'] = close_date.strftime('%m/%d/%Y')
        record['SalesCycleDays'] = (close_date - first_meeting).days
    else:
        record['CloseDate'] = ''
        record['SalesCycleDays'] = ''
    
    all_substeps = ['Intro', 'Followup', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Proposal', 'Compliance', 'Contracting', 'Pilot']
    for substep in all_substeps:
        if substep in meeting_dates and len(meeting_dates[substep]) > 0:
            first_date = min(meeting_dates[substep])
            record[f'DaysTo{substep}'] = (first_date - first_meeting).days
            record[f'{substep}Count'] = len(meeting_dates[substep])
        else:
            record[f'DaysTo{substep}'] = ''
            record[f'{substep}Count'] = 0
    
    if len(acct_meetings) >= 2:
        record['DaysFirstToSecond'] = (acct_meetings.iloc[1]['Date'] - acct_meetings.iloc[0]['Date']).days
    else:
        record['DaysFirstToSecond'] = ''
    
    account_data.append(record)

accounts_df = pd.DataFrame(account_data)
meetings_df = pd.DataFrame(all_meetings)

# Sort by Owner for easy filtering
accounts_df = accounts_df.sort_values(['Owner', 'Account']).reset_index(drop=True)

# Print owner distribution
owner_counts = accounts_df['Owner'].value_counts()
print(f"\n  Accounts by Owner:")
for owner, count in owner_counts.head(10).items():
    print(f"    {owner or '(unassigned)'}: {count}")

print(f"\n  Total accounts: {len(accounts_df)}")

# =============================================================================
# HELPER: Calculate stats for Summary
# =============================================================================
def calc_stats(df, col):
    valid = df[(df['Include'] == 'Y') & (df[col].apply(lambda x: isinstance(x, (int, float)) and x >= 0))]
    values = valid[col].dropna()
    if len(values) == 0:
        return {'avg': '', 'count': 0, 'min': '', 'max': ''}
    return {
        'avg': round(values.mean(), 1),
        'count': len(values),
        'min': int(values.min()),
        'max': int(values.max())
    }

# Add Include column for stats calculation
accounts_df['Include'] = 'Y'

# =============================================================================
# CREATE WORKBOOK
# =============================================================================
print("\n[STEP 3] Creating workbook...")

wb = Workbook()

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
editable_fill = PatternFill(start_color="FFFFD6", end_color="FFFFD6", fill_type="solid")
owner_fills = {
    'Julie Stefanich': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    'Asad Hussain': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    'Nathan Shine': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    'Alex Fox': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    'Tom Clancy': PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
}

# =============================================================================
# TAB 1: SalesReview (EDIT HERE)
# =============================================================================
ws_review = wb.active
ws_review.title = 'SalesReview'

review_cols = ['Owner', 'Account', 'Include', 'Verified', 'TotalMeetings', 'FirstMeetingDate', 
               'CloseDate', 'SalesCycleDays', 'DaysFirstToSecond', 'DaysToDemo', 'DemoCount', 
               'DaysToCAB', 'CABCount', 'DaysToScoping', 'ScopingCount', 'DaysToCompliance', 
               'ComplianceCount', 'DaysToContracting', 'ContractingCount', 'Notes']

# Headers
for col_idx, col_name in enumerate(review_cols, 1):
    cell = ws_review.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill

# Data - sorted by Owner
for row_idx, (_, row) in enumerate(accounts_df.iterrows(), 2):
    owner = row.get('Owner', '')
    
    for col_idx, col_name in enumerate(review_cols, 1):
        if col_name == 'Include':
            val = 'Y'
        elif col_name == 'Verified':
            val = ''
        elif col_name == 'Notes':
            val = ''
        else:
            val = row.get(col_name, '')
        
        cell = ws_review.cell(row=row_idx, column=col_idx, value=val)
        
        # Color code by owner
        if owner in owner_fills:
            cell.fill = owner_fills[owner]
        
        # Highlight key editable columns
        if col_name in ['Include', 'Verified', 'Notes']:
            cell.fill = editable_fill

num_accounts = len(accounts_df)
data_end_row = num_accounts + 1

# =============================================================================
# TAB 2: Inputs (formulas reference SalesReview)
# =============================================================================
ws_inputs = wb.create_sheet('Inputs')

# Map SalesReview columns to letters
sr_col_map = {col: get_column_letter(idx+1) for idx, col in enumerate(review_cols)}

input_cols = review_cols[:-1]  # Exclude Notes

for col_idx, col_name in enumerate(input_cols, 1):
    cell = ws_inputs.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

for row_idx in range(2, data_end_row + 1):
    for col_idx, col_name in enumerate(input_cols, 1):
        sr_col = sr_col_map[col_name]
        ws_inputs.cell(row=row_idx, column=col_idx, value=f'=SalesReview!{sr_col}{row_idx}')

# =============================================================================
# TAB 3: Summary (FULL - like v6)
# =============================================================================
ws_summary = wb.create_sheet('Summary')

headers = ['Category', 'Metric', 'Average', 'SampleN', 'Min', 'Max', 'Total', 'Notes']
for col_idx, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill

row_num = 3

# DATA OVERVIEW
ws_summary.cell(row=row_num, column=1, value='DATA OVERVIEW')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Total accounts (Include=Y)')
ws_summary.cell(row=row_num, column=3, value='=COUNTIF(Inputs!C:C,"Y")')
ws_summary.cell(row=row_num, column=8, value='Change Include to N to exclude')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Verified accounts')
ws_summary.cell(row=row_num, column=3, value='=COUNTIF(Inputs!D:D,"Y")')
ws_summary.cell(row=row_num, column=8, value='Accounts confirmed by sales')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='Accounts with close date')
ws_summary.cell(row=row_num, column=3, value='=COUNTIFS(Inputs!C:C,"Y",Inputs!H:H,">0")')
ws_summary.cell(row=row_num, column=8, value='Closed deals only')
row_num += 2

# SALES CYCLE
ws_summary.cell(row=row_num, column=1, value='SALES CYCLE (days)')
row_num += 1
stats = calc_stats(accounts_df, 'SalesCycleDays')
ws_summary.cell(row=row_num, column=2, value='First Meeting to Close')
ws_summary.cell(row=row_num, column=3, value='=AVERAGEIF(Inputs!C:C,"Y",Inputs!H:H)')
ws_summary.cell(row=row_num, column=4, value='=COUNTIFS(Inputs!C:C,"Y",Inputs!H:H,">0")')
ws_summary.cell(row=row_num, column=5, value=stats['min'])
ws_summary.cell(row=row_num, column=6, value=stats['max'])
ws_summary.cell(row=row_num, column=8, value='Min/Max are static')
row_num += 2

# TIME TO MILESTONE (FULL LIST)
ws_summary.cell(row=row_num, column=1, value='TIME TO MILESTONE (days from first meeting)')
row_num += 1

milestones = [
    ('DaysFirstToSecond', 'I', 'First to Second Meeting', 'Engagement velocity'),
    ('DaysToDemo', 'J', 'First to Demo', 'Key qualification step'),
    ('DaysToCAB', 'L', 'First to CAB Discussion', 'Champion identification'),
    ('DaysToScoping', 'N', 'First to Scoping', 'Pricing discussions'),
    ('DaysToCompliance', 'P', 'First to Compliance', 'Security review'),
    ('DaysToContracting', 'R', 'First to Contracting', 'Near close'),
]

for col_name, col_let, label, note in milestones:
    if col_name in accounts_df.columns:
        stats = calc_stats(accounts_df, col_name)
        ws_summary.cell(row=row_num, column=2, value=label)
        ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
        ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!C:C,"Y",Inputs!{col_let}:{col_let},">=0")')
        ws_summary.cell(row=row_num, column=5, value=stats['min'])
        ws_summary.cell(row=row_num, column=6, value=stats['max'])
        ws_summary.cell(row=row_num, column=8, value=note)
        row_num += 1

row_num += 1

# MEETING COUNTS (FULL LIST)
ws_summary.cell(row=row_num, column=1, value='MEETING COUNTS (per account)')
row_num += 1

counts = [
    ('DemoCount', 'K', 'Demo meetings'),
    ('CABCount', 'M', 'CAB discussions'),
    ('ScopingCount', 'O', 'Scoping meetings'),
    ('ComplianceCount', 'Q', 'Compliance meetings'),
    ('ContractingCount', 'S', 'Contracting meetings'),
]

for col_name, col_let, label in counts:
    if col_name in accounts_df.columns:
        ws_summary.cell(row=row_num, column=2, value=label)
        ws_summary.cell(row=row_num, column=3, value=f'=AVERAGEIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
        ws_summary.cell(row=row_num, column=4, value=f'=COUNTIFS(Inputs!C:C,"Y",Inputs!{col_let}:{col_let},">0")')
        ws_summary.cell(row=row_num, column=7, value=f'=SUMIF(Inputs!C:C,"Y",Inputs!{col_let}:{col_let})')
        row_num += 1

row_num += 1

# BY OWNER SUMMARY
ws_summary.cell(row=row_num, column=1, value='BY OWNER')
row_num += 1
ws_summary.cell(row=row_num, column=2, value='(Filter SalesReview by Owner column to see by rep)')
row_num += 2

# NOTES
ws_summary.cell(row=row_num, column=1, value='HOW THIS WORKS')
row_num += 1
notes = [
    'SalesReview tab: Reps verify their accounts (sorted by Owner)',
    'Inputs tab: Formulas pull from SalesReview',
    'Summary tab: Auto-calculates from Inputs',
    'To exclude: Set Include=N in SalesReview',
    'To correct: Edit value directly in SalesReview',
    'Average, SampleN, Total = formulas (auto-update)',
    'Min, Max = static values (update manually if needed)',
]
for note in notes:
    ws_summary.cell(row=row_num, column=2, value=note)
    row_num += 1

# =============================================================================
# TAB 4: Meetings
# =============================================================================
ws_meetings = wb.create_sheet('Meetings')

meeting_cols = ['Owner', 'Account', 'MeetingNum', 'Date', 'Subject', 'Classification']
for col_idx, col_name in enumerate(meeting_cols, 1):
    cell = ws_meetings.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Sort meetings by owner too
meetings_df = meetings_df.sort_values(['Owner', 'Account', 'MeetingNum']).reset_index(drop=True)

for row_idx, (_, row) in enumerate(meetings_df.iterrows(), 2):
    for col_idx, col_name in enumerate(meeting_cols, 1):
        val = row.get(col_name, '')
        if col_name == 'Date' and pd.notna(val):
            val = val.strftime('%m/%d/%Y')
        ws_meetings.cell(row=row_idx, column=col_idx, value=val)

# =============================================================================
# TAB 5: Instructions
# =============================================================================
ws_help = wb.create_sheet('Instructions')

instructions = [
    'SALES REVIEW WORKFLOW',
    '',
    'FOR SALES REPS:',
    '1. Open SalesReview tab',
    '2. Your accounts are grouped by Owner (your name)',
    '3. Review each account\'s data',
    '4. If data is wrong, edit it directly',
    '5. Set Verified = Y when confirmed',
    '6. Add Notes for any corrections needed',
    '',
    'DATA FLOW:',
    'SalesReview (you edit) → Inputs (formulas) → Summary (auto)',
    '',
    'COLUMNS TO REVIEW:',
    '- TotalMeetings: Count of meetings with this account',
    '- FirstMeetingDate: Date of first meeting',
    '- CloseDate: When deal closed (if applicable)',
    '- SalesCycleDays: Days from first meeting to close',
    '- DaysToDemo: Days from first meeting to first demo',
    '- DaysToCAB: Days to CAB discussion',
    '',
    'COLOR CODING:',
    '- Green: Julie Stefanich accounts',
    '- Orange: Asad Hussain accounts',
    '- Blue: Nathan Shine accounts',
    '- Yellow (cells): Editable fields',
    '',
    'FOR MANAGERS:',
    '- Summary shows aggregated metrics',
    '- Verified count = accounts confirmed by reps',
    '- Filter by Owner to review by rep',
]

for row_idx, text in enumerate(instructions, 1):
    ws_help.cell(row=row_idx, column=1, value=text)

# Column widths
ws_review.column_dimensions['A'].width = 18
ws_review.column_dimensions['B'].width = 25
ws_review.column_dimensions['T'].width = 30

# Save
output_path = '/Users/keiganpesenti/Desktop/sales_substeps_v8.xlsx'
wb.save(output_path)

print(f"\n  Workbook saved: {output_path}")
print("\n  TABS:")
print("    SalesReview - SORTED BY OWNER - reps find their section easily")
print("    Inputs      - Formulas from SalesReview")
print("    Summary     - FULL metrics (restored)")
print("    Meetings    - Raw data sorted by Owner")
print("    Instructions")

