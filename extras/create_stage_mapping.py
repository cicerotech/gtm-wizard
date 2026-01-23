"""
Create Stage Mapping for Detailed Process Map
==============================================
Maps our Inputs_formula milestones to sales stages
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("="*80)
print("CREATING STAGE MAPPING FOR PROCESS MAP")
print("="*80)

# Load our Inputs data to calculate averages
print("\n[1] Loading Inputs data...")

# We'll calculate from Source_Meetings to get accurate stage-by-stage breakdown
audit_file = '/Users/keiganpesenti/Desktop/latest audits v0.xlsx'
df_meetings_audit = pd.read_excel(audit_file, sheet_name='latest audits v0')
df_meetings_secondary = pd.read_excel('/Users/keiganpesenti/Desktop/meetings with accounts.xlsx', sheet_name='all meetings')

df_meetings = pd.concat([df_meetings_audit, df_meetings_secondary], ignore_index=True)
df_meetings['Date'] = pd.to_datetime(df_meetings['Date'], errors='coerce')
df_meetings = df_meetings[df_meetings['Date'] >= '2020-01-01'].copy()

# Dedupe
df_meetings['_key'] = (
    df_meetings['Company / Account'].fillna('').str.lower().str.strip() + '|' +
    df_meetings['Date'].dt.date.astype(str) + '|' +
    df_meetings['Subject'].fillna('').str.lower().str.strip()
)
df_meetings = df_meetings.drop_duplicates(subset='_key', keep='first')

# Classify
def classify_meeting(subject):
    if pd.isna(subject):
        return 'Followup'
    s = str(subject).lower().strip()
    
    if 'intro' in s or 'introduction' in s:
        return 'Intro'
    if any(kw in s for kw in ['demo', 'sigma', 'cortex', 'platform', 'walkthrough', 'product overview']):
        return 'Demo'
    if any(kw in s for kw in ['cab', 'customer advisory', 'advisory board']):
        return 'CAB'
    if any(kw in s for kw in ['scoping', 'scope', 'pricing', 'proposal']):
        return 'Scoping'
    if any(kw in s for kw in ['infosec', 'security', 'compliance']):
        return 'Compliance'
    if any(kw in s for kw in ['contract', 'redline', 'msa', 'negotiation', 'legal']):
        return 'Contracting'
    if any(kw in s for kw in ['use case', 'champion', 'landing path']):
        return 'UseCase'
    if any(kw in s for kw in ['kickoff', 'delivery', 'onboard']):
        return 'Delivery'
    return 'Followup'

df_meetings['Classification'] = df_meetings['Subject'].apply(classify_meeting)

# Calculate days from first meeting for each account
df_meetings['_acct'] = df_meetings['Company / Account'].fillna('').str.lower().str.strip()
first_meetings = df_meetings.groupby('_acct')['Date'].min().reset_index()
first_meetings.columns = ['_acct', 'FirstMeetingDate']
df_meetings = df_meetings.merge(first_meetings, on='_acct')
df_meetings['DaysFromFirst'] = (df_meetings['Date'] - df_meetings['FirstMeetingDate']).dt.days

print(f"  Loaded {len(df_meetings)} meetings")

# Calculate averages by classification
print("\n[2] Calculating stage timings...")

# For each classification, get the average days to first occurrence
stage_timings = {}

classifications = ['Intro', 'Followup', 'CAB', 'UseCase', 'Demo', 'Scoping', 'Compliance', 'Contracting', 'Delivery']

for cls in classifications:
    cls_meetings = df_meetings[df_meetings['Classification'] == cls]
    if len(cls_meetings) > 0:
        # Get first occurrence per account
        first_per_account = cls_meetings.groupby('_acct')['DaysFromFirst'].min()
        avg_days = first_per_account.mean()
        count = len(first_per_account)
        stage_timings[cls] = {'avg': avg_days, 'count': count}
        print(f"  {cls}: Avg {avg_days:.1f} days (n={count})")

# Also calculate second meeting timing
account_meetings = df_meetings.groupby('_acct').apply(
    lambda x: x.sort_values('Date').iloc[1]['DaysFromFirst'] if len(x) >= 2 else None
).dropna()
second_meeting_avg = account_meetings.mean()
print(f"  Second Meeting: Avg {second_meeting_avg:.1f} days (n={len(account_meetings)})")

# =============================================================================
# Create the Process Map timing data
# =============================================================================
print("\n[3] Creating process map timing data...")

# Mapping our data to the detailed process map stages
# Based on your screenshot structure:

process_stages = [
    # Stage, Step, Our Data Source, Time (Days)
    ('Stage 1 Discovery', 'Meeting 1 Eudia intro', 'First Meeting', 0),
    ('Stage 1 Discovery', 'Followup 1', 'Second Meeting', round(second_meeting_avg) if not pd.isna(second_meeting_avg) else 38),
    ('Stage 1 Discovery', 'Followup 2', 'Followup Avg', round(stage_timings.get('Followup', {}).get('avg', 71))),
    ('Stage 1 Discovery', 'Meeting 2a CAB discussion', 'CAB', round(stage_timings.get('CAB', {}).get('avg', 19))),
    ('Stage 1 discover/qualification', 'Meeting 2b "use case identification"', 'UseCase', round(stage_timings.get('UseCase', {}).get('avg', 5))),
    ('Stage 2 SQO', 'Meeting 3 Products overview', 'Demo', round(stage_timings.get('Demo', {}).get('avg', 71))),
    ('Stage 2 (internal)', 'Scoping, pricing, delivery assessment', 'Scoping', round(stage_timings.get('Scoping', {}).get('avg', 67))),
    ('Stage 4 proposal', 'Meeting 4 Proposal and delivery plan', 'Proposal', round(stage_timings.get('Scoping', {}).get('avg', 67)) + 5),
    ('Stage 4 proposal', 'Deal desk informal', 'Contracting', round(stage_timings.get('Contracting', {}).get('avg', 61))),
    ('Stage 4 proposal', 'Infosec review', 'Compliance', round(stage_timings.get('Compliance', {}).get('avg', 59))),
    ('Stage 4 proposal', 'Meeting 5 followups', 'Followup', round(stage_timings.get('Contracting', {}).get('avg', 61)) + 5),
    ('Stage 4 proposal', 'Share proposal with counterparty', 'Contracting+5', round(stage_timings.get('Contracting', {}).get('avg', 61)) + 10),
    ('Stage 4 proposal', 'Sign contract', 'Close', round(stage_timings.get('Contracting', {}).get('avg', 61)) + 20),
    ('Stage 5 Delivery', 'Internal kickoff meeting', 'Delivery', round(stage_timings.get('Delivery', {}).get('avg', 0)) if stage_timings.get('Delivery') else 5),
]

# =============================================================================
# Generate Excel output
# =============================================================================
print("\n[4] Generating Excel output...")

wb = Workbook()
ws = wb.active
ws.title = 'Process_Map_Timing'

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

headers = ['Stage', 'Step', 'Data_Source', 'Time_Days', 'Cumulative', 'Notes']
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill

cumulative = 0
for row_idx, (stage, step, source, days) in enumerate(process_stages, 2):
    ws.cell(row=row_idx, column=1, value=stage)
    ws.cell(row=row_idx, column=2, value=step)
    ws.cell(row=row_idx, column=3, value=source)
    ws.cell(row=row_idx, column=4, value=days)
    
    # Cumulative only for sequential stages
    if 'Stage 1' in stage and cumulative == 0:
        cumulative = days
    elif days > cumulative:
        cumulative = days
    ws.cell(row=row_idx, column=5, value=cumulative)

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 15

# === TAB 2: Summary mapping to your SF stages ===
ws2 = wb.create_sheet('Stage_Duration_Mapping')

# Map our milestones to SF stages
sf_stage_mapping = [
    ['SF Stage', 'Your Milestone', 'Avg Days From First Meeting', 'Sample N', 'Notes'],
    ['Stage 0 - Prospecting', 'Before First Meeting', 0, '', 'Pre-engagement'],
    ['Stage 1 - Discovery', 'First to Second Meeting', round(second_meeting_avg) if not pd.isna(second_meeting_avg) else 38, len(account_meetings), 'Initial engagement phase'],
    ['Stage 1 - Discovery', 'First to CAB', round(stage_timings.get('CAB', {}).get('avg', 0)), stage_timings.get('CAB', {}).get('count', 0), 'Champion identification'],
    ['Stage 2 - SQO', 'First to Demo', round(stage_timings.get('Demo', {}).get('avg', 0)), stage_timings.get('Demo', {}).get('count', 0), 'Qualification confirmed'],
    ['Stage 2 - SQO', 'First to Scoping', round(stage_timings.get('Scoping', {}).get('avg', 0)), stage_timings.get('Scoping', {}).get('count', 0), 'Pricing discussions'],
    ['Stage 4 - Proposal', 'First to Contracting', round(stage_timings.get('Contracting', {}).get('avg', 0)), stage_timings.get('Contracting', {}).get('count', 0), 'Deal negotiation'],
    ['Stage 4 - Proposal', 'First to Compliance', round(stage_timings.get('Compliance', {}).get('avg', 0)), stage_timings.get('Compliance', {}).get('count', 0), 'Security review'],
]

for row_idx, row_data in enumerate(sf_stage_mapping, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 30

# === TAB 3: Formula to paste into your Detailed process map ===
ws3 = wb.create_sheet('Formulas_For_ProcessMap')

formulas = [
    ['Step in Process Map', 'Formula to use in Time(Days) column'],
    ['Meeting 1 Eudia intro', '0'],
    ['Followup 1', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!I:I)'],
    ['Followup 2', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!I:I)*1.5'],
    ['Meeting 2a CAB discussion', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!L:L)'],
    ['Meeting 3 Products overview', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!J:J)'],
    ['Scoping, pricing', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!N:N)'],
    ['Contracting', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!R:R)'],
    ['Compliance/Infosec', f'=AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!P:P)'],
]

for row_idx, row_data in enumerate(formulas, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 60

# Save
output_path = '/Users/keiganpesenti/Desktop/STAGE_TIMING_MAPPING.xlsx'
wb.save(output_path)

print(f"\n  Output saved: {output_path}")
print("\n" + "="*80)
print("MAPPING YOUR DATA TO PROCESS MAP")
print("="*80)
print(f"""
Your Inputs_formula columns → Process Map Time(Days):

Column I (DaysFirstToSecond) → Followup 1 timing
Column J (DaysToDemo)        → Meeting 3 Products overview / Stage 2 SQO
Column L (DaysToCAB)         → Meeting 2a CAB discussion / Stage 1
Column N (DaysToScoping)     → Scoping, pricing / Stage 2
Column P (DaysToCompliance)  → Infosec review / Stage 4
Column R (DaysToContracting) → Sign contract / Stage 4

AVERAGEIF formulas to use:
  =AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!I:I)  → Avg days to second meeting
  =AVERAGEIF(Inputs_formula!C:C,"Y",Inputs_formula!J:J)  → Avg days to demo
  etc.
""")

